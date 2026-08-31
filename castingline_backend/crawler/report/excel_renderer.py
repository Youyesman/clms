# -*- coding: utf-8 -*-
"""P001: 영화 상영현황 보고서 — 엑셀 렌더러 (openpyxl).

첨부 샘플(sample_excel / sample_excel_주요작X)의 시트·표 구조·색상을 따른다.
숫자는 PDF와 동일한 ViewModel(aggregation.build_report_data)만 사용한다. (§23)
"""
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XlImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .common import (COLOR_HEADER_GREEN, COLOR_MAIN_CREAM, COLOR_RED, LOGO_PATH,
                     fmt_cmp, fmt_num, fmt_pct, fmt_rank_cmp)

FONT_NAME = "맑은 고딕"
GREEN_FILL = PatternFill(start_color=COLOR_HEADER_GREEN, end_color=COLOR_HEADER_GREEN, fill_type="solid")
CREAM_FILL = PatternFill(start_color=COLOR_MAIN_CREAM, end_color=COLOR_MAIN_CREAM, fill_type="solid")

TITLE_FONT = Font(name=FONT_NAME, size=16, bold=True)
SUB_FONT = Font(name=FONT_NAME, size=9, color="555555")
SECTION_FONT = Font(name=FONT_NAME, size=11, bold=True)
HEADER_FONT = Font(name=FONT_NAME, size=9, bold=True)
DATA_FONT = Font(name=FONT_NAME, size=9)
DATA_BOLD = Font(name=FONT_NAME, size=9, bold=True)
BIG_FONT = Font(name=FONT_NAME, size=14, bold=True)
RED_FONT = Font(name=FONT_NAME, size=9, bold=True, color=COLOR_RED)

_side = Side(style="thin", color="BFBFBF")
BORDER = Border(left=_side, right=_side, top=_side, bottom=_side)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=False)
LEFT = Alignment(horizontal="left", vertical="center")


def _cell(ws, row, col, value, font=DATA_FONT, fill=None, border=True, align=CENTER):
    c = ws.cell(row=row, column=col, value=value)
    c.font = font
    if fill:
        c.fill = fill
    if border:
        c.border = BORDER
    c.alignment = align
    return c


def _cmp_value(cmp_dict):
    """전주 비교 → (표시문자열, 폰트)"""
    txt, red = fmt_cmp(cmp_dict)
    return txt, (RED_FONT if red else DATA_FONT)


def _sheet_head(ws, title, subtitle, last_col):
    _cell(ws, 1, 1, title, font=TITLE_FONT, border=False, align=LEFT)
    _cell(ws, 2, 1, subtitle, font=SUB_FONT, border=False, align=LEFT)
    # 우측 상단 CASTING LINE 로고 (§22)
    try:
        img = XlImage(LOGO_PATH)
        h = 42
        img.width = int(img.width * h / img.height)
        img.height = h
        ws.add_image(img, f"{get_column_letter(max(last_col - 1, 1))}1")
    except Exception:
        pass


def _write_table(ws, start_row, headers, rows, highlight_rows=None):
    """headers: [str] / rows: [[(값, 폰트|None)]]. 반환: 다음 빈 행 번호"""
    highlight_rows = highlight_rows or set()
    for ci, h in enumerate(headers, 1):
        _cell(ws, start_row, ci, h, font=HEADER_FONT, fill=GREEN_FILL)
    r = start_row + 1
    for ri, row in enumerate(rows):
        fill = CREAM_FILL if ri in highlight_rows else None
        for ci, item in enumerate(row, 1):
            val, font = item if isinstance(item, tuple) else (item, DATA_FONT)
            _cell(ws, r, ci, val, font=font or DATA_FONT, fill=fill)
        r += 1
    return r


def _set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


_WEEKDAY_KOR = ["월", "화", "수", "목", "금", "토", "일"]


def _period_str(p):
    # C008: 특정 날짜(비연속) 선택 보고서는 범위(~) 대신 날짜를 나열해 표기한다
    if p.get("dates"):
        cur = ", ".join(d.strftime("%m.%d") for d in p["dates"])
        prev = ", ".join(d.strftime("%m.%d") for d in (p.get("prev_dates") or []))
        return (f"조사기간 {p['start'].strftime('%Y')}년 {cur}",
                f"전주 {prev}")
    return (f"조사기간 {p['start'].strftime('%Y.%m.%d')} - {p['end'].strftime('%Y.%m.%d')}",
            f"전주 {p['prev_start'].strftime('%Y.%m.%d')} - {p['prev_end'].strftime('%Y.%m.%d')}")


def _breakdown_sheet(ws, row, title, block):
    """A001(0829): 포맷별 / 지역별 상세 현황 — 2줄 헤더 + 일자별(총 좌석수·비중)."""
    _cell(ws, row, 1, title, font=SECTION_FONT, border=False, align=LEFT)
    dates = block["dates"]
    n = len(dates)
    hr1, hr2 = row + 1, row + 2

    _cell(ws, hr1, 1, "구분", font=HEADER_FONT, fill=GREEN_FILL)
    _cell(ws, hr2, 1, "", font=HEADER_FONT, fill=GREEN_FILL)
    ci = 2
    for d in dates:
        label = f"{d.month}/{d.day} ({_WEEKDAY_KOR[d.weekday()]})"
        _cell(ws, hr1, ci, label, font=HEADER_FONT, fill=GREEN_FILL)
        _cell(ws, hr1, ci + 1, "", font=HEADER_FONT, fill=GREEN_FILL)
        _cell(ws, hr2, ci, "총 좌석수", font=HEADER_FONT, fill=GREEN_FILL)
        _cell(ws, hr2, ci + 1, "비중", font=HEADER_FONT, fill=GREEN_FILL)
        ws.merge_cells(start_row=hr1, start_column=ci, end_row=hr1, end_column=ci + 1)
        ci += 2
    _cell(ws, hr1, ci, "합계", font=HEADER_FONT, fill=GREEN_FILL)
    _cell(ws, hr1, ci + 1, "", font=HEADER_FONT, fill=GREEN_FILL)
    _cell(ws, hr2, ci, "총 좌석수", font=HEADER_FONT, fill=GREEN_FILL)
    _cell(ws, hr2, ci + 1, "비중", font=HEADER_FONT, fill=GREEN_FILL)
    ws.merge_cells(start_row=hr1, start_column=ci, end_row=hr1, end_column=ci + 1)
    _cell(ws, hr1, ci + 2, block["count_label"], font=HEADER_FONT, fill=GREEN_FILL)
    _cell(ws, hr2, ci + 2, "", font=HEADER_FONT, fill=GREEN_FILL)
    _cell(ws, hr1, ci + 3, "회차수", font=HEADER_FONT, fill=GREEN_FILL)
    _cell(ws, hr2, ci + 3, "", font=HEADER_FONT, fill=GREEN_FILL)
    ws.merge_cells(start_row=hr1, start_column=1, end_row=hr2, end_column=1)
    ws.merge_cells(start_row=hr1, start_column=ci + 2, end_row=hr2, end_column=ci + 2)
    ws.merge_cells(start_row=hr1, start_column=ci + 3, end_row=hr2, end_column=ci + 3)

    r = hr2 + 1
    for item in block["rows"]:
        # '합계' 행만 연한 노랑 (헤더는 다른 표와 같은 연한 연두)
        fill = CREAM_FILL if item.get("is_total") else None
        font = DATA_BOLD if item.get("is_total") else DATA_FONT
        _cell(ws, r, 1, item["label"], font=DATA_BOLD, fill=fill)
        c = 2
        for day in item["days"]:
            _cell(ws, r, c, fmt_num(day["seats"], "석"), font=font, fill=fill)
            _cell(ws, r, c + 1, fmt_pct(day["share"]), font=font, fill=fill)
            c += 2
        _cell(ws, r, c, fmt_num(item["total_seats"], "석"), font=font, fill=fill)
        _cell(ws, r, c + 1, fmt_pct(item["total_share"]), font=font, fill=fill)
        _cell(ws, r, c + 2, fmt_num(item["count"]), font=font, fill=fill)
        _cell(ws, r, c + 3, fmt_num(item["shows"]), font=font, fill=fill)
        r += 1
    return r + 1


def _daily_kpi_block(ws, row, title, kpi_rows):
    """일별 4대 지표 표 (행=지표 / 열=일자별 값·전주 대비)."""
    _cell(ws, row, 1, title, font=SECTION_FONT, border=False, align=LEFT)
    headers = ["지표"]
    for r in kpi_rows:
        headers += [f"{r['date'].month}/{r['date'].day}",
                    f"전주 {r['prev_date'].month}/{r['prev_date'].day} 대비"]
    metric_defs = [
        ("총 좌석수", lambda r: fmt_num(r["total_seats"], "석"), "seats_cmp"),
        ("예매좌석수", lambda r: fmt_num(r["reserved"], "석"), "reserved_cmp"),
        ("좌석점유율", lambda r: fmt_pct(r["occupancy"]), "occ_cmp"),
        ("회차수", lambda r: fmt_num(r["shows"], "회"), "shows_cmp"),
    ]
    rows = []
    for label, fn, ck in metric_defs:
        line = [(label, DATA_BOLD)]
        for r in kpi_rows:
            line.append(fn(r))
            line.append(_cmp_value(r[ck]))
        rows.append(line)
    return _write_table(ws, row + 1, headers, rows)


def _top_daily_block(ws, row, title, top, dates, star):
    """TOP10(+주요작) 일별 전주 비교."""
    _cell(ws, row, 1, title, font=SECTION_FONT, border=False, align=LEFT)
    headers = ["#", "작품명"]
    for d in dates:
        headers += [f"{d.month}/{d.day} 총 좌석수(좌점율)", "전주比"]
    rows, hl = [], set()
    for i, m in enumerate(top):
        is_star = star and m["is_main"]
        line = [m["rank"],
                (("★ " + m["title"]) if is_star else m["title"],
                 DATA_BOLD if is_star else DATA_FONT)]
        for day in m["days"]:
            line.append(f"{fmt_num(day['total_seats'])} ({fmt_pct(day['occupancy'])})")
            line.append(_cmp_value(day["seats_cmp"]))
        rows.append(line)
        if is_star:
            hl.add(i)
    return _write_table(ws, row + 1, headers, rows, highlight_rows=hl)


def _slot_block(ws, row, slots):
    """주요작 일자별 주요 시간대 회차 배정 비중."""
    _cell(ws, row, 1, "주요작 일자별 주요 시간대 회차 배정 비중",
          font=SECTION_FONT, border=False, align=LEFT)
    headers = ["구분"] + slots["labels"] + ["총 회차"]
    rows = []
    for r in slots["rows"]:
        d = r["date"]
        line = [(f"{d.month}/{d.day} ({_WEEKDAY_KOR[d.weekday()]})", DATA_BOLD)]
        for c in r["counts"]:
            pct = (c / r["total"] * 100) if r["total"] else 0.0
            line.append(f"{fmt_num(c)}회 ({fmt_pct(pct)})")
        line.append((f"{fmt_num(r['total'])}회 (100%)", DATA_BOLD))
        rows.append(line)
    rows.append([("기간 평균 비중", DATA_BOLD)]
                + [fmt_pct(x) for x in slots["avg_pct"]] + ["100.0%"])
    return _write_table(ws, row + 1, headers, rows)


def build_excel(data, out_path):
    """A001(0829): PDF와 같은 페이지 구성·같은 표로 시트를 만든다."""
    p = data["period"]
    cur_s, prev_s = _period_str(p)
    daily = data.get("daily") or {}
    wb = Workbook()
    wb.remove(wb.active)

    n = len(daily.get("dates") or [])

    if data["mode"] == "main":
        main = data["main"]

        # ================= P.1 주요작 상영 현황 =================
        ws = wb.create_sheet("P1 주요작 상영 현황")
        _set_widths(ws, [16, 18, 24] + [18, 22] * max(n, 3))
        _sheet_head(ws, "P.1 주요작 상영 현황",
                    f"작품명 {main['title']} | {cur_s} | {prev_s}", last_col=max(2 * n + 2, 8))
        r = 4
        if daily.get("kpi_rows"):
            r = _daily_kpi_block(ws, r, f"주요작 일별 상영 현황 - {main['title']}",
                                 daily["kpi_rows"]) + 1

        _cell(ws, r, 1, "멀티별 현황", font=SECTION_FONT, border=False, align=LEFT)
        multi_rows = []
        for m in data["main_multis"]:
            multi_rows.append([
                m["name"], fmt_num(m["total_seats"]), _cmp_value(m["seats_cmp"]),
                fmt_num(m["reserved"]), fmt_pct(m["occupancy"]),
                fmt_num(m["screens"]), fmt_num(m["theaters"]), fmt_num(m["shows"]),
            ])
        r = _write_table(ws, r + 1,
                         ["멀티", "총 좌석수", "전주比", "예매좌석수", "점유율", "스크린", "극장", "회차"],
                         multi_rows) + 1

        if daily.get("slots"):
            r = _slot_block(ws, r, daily["slots"]) + 1

        if data.get("main_formats", {}).get("rows"):
            _breakdown_sheet(ws, r, "포맷별 상세 현황", data["main_formats"])

        # ================= P.2 주요작 상세 현황 =================
        ws = wb.create_sheet("P2 주요작 상세 현황")
        _set_widths(ws, [14, 16, 30] + [16, 22] * max(n, 3))
        _sheet_head(ws, "P.2 주요작 상세 현황", f"{cur_s} | {prev_s}",
                    last_col=max(2 * n + 5, 9))
        r = 4
        if data.get("main_regions", {}).get("rows"):
            r = _breakdown_sheet(ws, r, "지역별 상세 현황", data["main_regions"]) + 1

        _cell(ws, r, 1, "총 좌석수 기준 극장 TOP 10", font=SECTION_FONT, border=False, align=LEFT)
        top_rows = []
        for i, th in enumerate(data["main_theaters_top10"], 1):
            top_rows.append([
                i, th["multi"], th["name"], fmt_num(th["total_seats"]),
                _cmp_value(th["seats_cmp"]), fmt_num(th["reserved"]),
                fmt_pct(th["occupancy"]), fmt_num(th["screens"]), fmt_num(th["shows"]),
            ])
        _write_table(ws, r + 1,
                     ["#", "멀티", "극장명", "총 좌석수", "전주比", "예매좌석수", "점유율", "스크린", "회차"],
                     top_rows)

        # ================= P.3 주요작 vs 경쟁작 =================
        ws = wb.create_sheet("P3 주요작 vs 경쟁작")
        _set_widths(ws, [6, 30] + [22, 24] * max(n, 3))
        _sheet_head(ws, "P.3 주요작 vs 경쟁작", f"{cur_s} | {prev_s}",
                    last_col=max(2 * n + 2, 6))
        _cell(ws, 4, 1, f"주요작 경쟁 포지션  (전체 {data['movie_count']}개 작품 중)",
              font=SECTION_FONT, border=False, align=LEFT)
        rk = main["ranks"]
        cards = [("좌석수 순위", "seats", fmt_num(main["total_seats"], "석")),
                 ("예매좌석 순위", "reserved", fmt_num(main["reserved"], "석")),
                 ("점유율 순위", "occupancy", fmt_pct(main["occupancy"])),
                 ("회차 순위", "shows", fmt_num(main["shows"], "회"))]
        for ci, (label, _, _) in enumerate(cards, 1):
            _cell(ws, 5, ci, label, font=HEADER_FONT, fill=GREEN_FILL)
        for ci, (_, k, _) in enumerate(cards, 1):
            _cell(ws, 6, ci, f"{rk[k]['cur']}위", font=BIG_FONT)
        for ci, (_, _, v) in enumerate(cards, 1):
            _cell(ws, 7, ci, v)
        ws.row_dimensions[6].height = 26

        if daily.get("top"):
            _top_daily_block(ws, 9,
                             f"주요작 일별 전주 비교  (경쟁작 총 {data['movie_count']}개 작품)",
                             daily["top"], daily["dates"], star=True)

    else:
        totals = data["totals"]

        # ================= P.1 경쟁작 전체 상영 요약 =================
        ws = wb.create_sheet("P1 경쟁작 전체 상영 요약")
        _set_widths(ws, [6, 30] + [22, 24] * max(n, 3))
        _sheet_head(ws, "P.1 경쟁작 전체 상영 요약",
                    f"주요 상영작 {data['movie_count']}개 작품 | {cur_s} | {prev_s}",
                    last_col=max(2 * n + 2, 6))
        r = 4
        if daily.get("kpi_rows"):
            r = _daily_kpi_block(ws, r, "전체 시장 일별 현황", daily["kpi_rows"]) + 1

        _cell(ws, r, 1, "경쟁작 경쟁 포지션", font=SECTION_FONT, border=False, align=LEFT)
        L = data["leaders"]
        cards = [("좌석수 1위", L["seats"]), ("예매좌석 1위", L["reserved"]),
                 ("점유율 1위", L["occupancy"]), ("회차 1위", L["shows"])]
        for ci, (label, _) in enumerate(cards, 1):
            _cell(ws, r + 1, ci, label, font=HEADER_FONT, fill=GREEN_FILL)
        for ci, (_, c) in enumerate(cards, 1):
            _cell(ws, r + 2, ci, c["title"], font=DATA_BOLD)
        for ci, (_, c) in enumerate(cards, 1):
            val = fmt_pct(c["value"]) if c["unit"] == "%" else fmt_num(c["value"], c["unit"])
            _cell(ws, r + 3, ci, val)
        for ci, (_, c) in enumerate(cards, 1):
            txt, red = fmt_rank_cmp(c["prev_rank"], c["cur_rank"])
            _cell(ws, r + 4, ci, txt, font=RED_FONT if red else DATA_FONT)
        r += 6

        if daily.get("top"):
            _top_daily_block(ws, r, "총 좌석수 기준 TOP 10 작품 일별 전주 비교",
                             daily["top"], daily["dates"], star=False)

        # ================= P.2 경쟁작 일별 상세 현황 (일자마다 시트 하나) =================
        headers = ["#", "작품명", "총 좌석수", "전주比", "점유비중", "예매좌석수",
                   "극장수", "스크린수", "회차수"]
        for day in daily.get("by_date_movies") or []:
            d = day["date"]
            ws = wb.create_sheet(f"P2 {d.strftime('%m.%d')}({_WEEKDAY_KOR[d.weekday()]})")
            _set_widths(ws, [5, 30, 13, 22, 11, 13, 10, 11, 10])
            _sheet_head(ws, f"P.2 경쟁작 일별 상세 현황 — {d.strftime('%Y.%m.%d')}"
                            f"({_WEEKDAY_KOR[d.weekday()]})",
                        f"주요 상영작 {data['movie_count']}개 작품 | {cur_s} | 총 좌석수 기준 내림차순",
                        last_col=9)
            rows = []
            for s in day["rows"]:
                rows.append([
                    s["rank"], s["title"], fmt_num(s["total_seats"]),
                    _cmp_value(s["seats_cmp"]), fmt_pct(s["share"]),
                    fmt_num(s["reserved"]), fmt_num(s["theaters"]),
                    fmt_num(s["screens"]), fmt_num(s["shows"]),
                ])
            _write_table(ws, 4, headers, rows)

    for ws in wb.worksheets:
        ws.sheet_view.showGridLines = False

    wb.save(out_path)
    return out_path
