"""KOBIS(영화관입장권통합전산망) 회원용통계(영화사별)상세 수집기.

수동 흐름(배급사 계정 로그인 → 회원용통계보기 → 영화별 상세 → 엑셀저장)을
서버 사이드로 재현한다.
  1) GET  /kobis/business/comm/user/openLogin.do        : 세션/CSRF 확보
  2) GET  /kobis/business/comm/user/findSmsNo.do        : 인증번호 검증 (best-effort)
  3) POST /kobis/j_login?j_username=..&j_password=..    : 스프링 시큐리티 로그인
  4) GET  /kobis/business/mast/thea/findCompanyStat.do  : 회원용통계 페이지(CSRF)
  5) POST findCompanyStat.do {sStartDt,sEndDt}          : 기간 내 영화 목록
  6) POST findCompanyStatDetailXls.do {movieCd,...}     : 영화별 상세 엑셀
     (SpreadsheetML XML → openpyxl xlsx 변환. 변환된 파일은 스코어 업로더의
      영진위(일반극장) 파서 preview_kofic_format 이 그대로 읽는 양식)

제약: 조회 기간 최대 1개월(사이트 제한). 파일이 크면(월 단위 인기작) 다운로드에
수십 초 걸릴 수 있다.
"""

import io
import re
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime
from xml.etree import ElementTree as ET

import requests
from openpyxl import Workbook

BASE = "https://www.kobis.or.kr"
_UA = "Mozilla/5.0 (Windows NT 10.0; Win64; x64)"
_SS = "{urn:schemas-microsoft-com:office:spreadsheet}"


def _login(session, user, password, aprv_no=""):
    """KOBIS 로그인. 실패 시 RuntimeError."""
    session.get(f"{BASE}/kobis/business/comm/user/openLogin.do", timeout=30)
    if aprv_no:
        # 로그인 화면의 인증번호 검증 — 실패해도 j_login 은 시도한다.
        try:
            session.get(
                f"{BASE}/kobis/business/comm/user/findSmsNo.do",
                params={"userId": user, "userPw": password, "aprvNo": aprv_no},
                headers={"Accept": "application/json, text/javascript"},
                timeout=20,
            )
        except requests.RequestException:
            pass
    r = session.post(
        f"{BASE}/kobis/j_login",
        params={"j_username": user, "j_password": password},
        headers={"Accept": "application/extJs+sua",
                 "Content-Type": "application/extJs+sua"},
        data='{test:"test"}',
        timeout=30,
    )
    if '"userId"' not in r.text:
        m = re.search(r'"value"\s*:\s*"([^"]+)"', r.text)
        raise RuntimeError(f"로그인 실패: {m.group(1) if m else '아이디/비밀번호 확인 필요'}")


def _fetch_csrf(session):
    """회원용통계 페이지에서 CSRF 토큰 확보."""
    r = session.get(f"{BASE}/kobis/business/mast/thea/findCompanyStat.do", timeout=30)
    m = re.search(r'name="CSRFToken" value="([^"]+)"', r.text)
    if not m:
        raise RuntimeError("회원용통계 페이지 접근 실패(권한/세션 확인)")
    return m.group(1)


def _num(s):
    try:
        return int(str(s).replace(",", "").strip() or 0)
    except ValueError:
        return 0


def fetch_movie_list(session, csrf, start_dt, end_dt):
    """기간 조회 → 영화 목록 [{movieCd, movieNm, theaters, visitors}].

    목록 행: <a onclick="mstView('movie','CD')"><strong>영화명</strong></a>
            + 극장수|스크린수|발권금액|매출액|관객수|누적관객수 셀
    """
    r = session.post(
        f"{BASE}/kobis/business/mast/thea/findCompanyStat.do",
        data={"CSRFToken": csrf, "movieCd": "", "loadEnd": "0",
              "sStartDt": start_dt, "sEndDt": end_dt},
        timeout=60,
    )
    movies = []
    for m in re.finditer(
        r"mstView\('movie','(\d+)'\).*?<strong>([^<]+)</strong>(.*?)</tr>",
        r.text, re.S,
    ):
        cells = [_num(c) for c in
                 re.findall(r'<td[^>]*class="total"[^>]*>\s*([\d,]+)\s*</td>', m.group(3))]
        movies.append({
            "movieCd": m.group(1),
            "movieNm": m.group(2).strip(),
            "theaters": cells[0] if len(cells) > 0 else 0,
            "visitors": cells[4] if len(cells) > 4 else 0,
        })
    return movies


def _xml_to_xlsx(xml_bytes):
    """KOBIS 상세 엑셀(SpreadsheetML XML) → xlsx bytes (레이아웃 1:1 유지).

    시트 구조가 영진위 업로드 양식과 동일하므로 그대로 변환만 한다.
    SpreadsheetML 은 빈 셀을 생략하고 ss:Index 로 건너뛰므로 컬럼 포인터를 관리한다.
    """
    root = ET.fromstring(xml_bytes)
    wb = Workbook()
    wb.remove(wb.active)
    for wi, ws_el in enumerate(root.findall(f"{_SS}Worksheet")):
        name = ws_el.get(f"{_SS}Name") or f"sheet{wi}"
        ws = wb.create_sheet(title=name[:31])
        ri = 0
        for row_el in ws_el.iter(f"{_SS}Row"):
            ri += 1
            ci = 0
            for cell_el in row_el.findall(f"{_SS}Cell"):
                idx = cell_el.get(f"{_SS}Index")
                ci = int(idx) if idx else ci + 1
                data_el = cell_el.find(f"{_SS}Data")
                if data_el is None or data_el.text is None:
                    continue
                text = data_el.text
                if data_el.get(f"{_SS}Type") == "Number":
                    try:
                        num = float(text)
                        value = int(num) if num == int(num) else num
                    except ValueError:
                        value = text
                else:
                    value = text
                ws.cell(row=ri, column=ci, value=value)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def download_movie_detail_xlsx(session, csrf, movie_cd, start_dt, end_dt):
    """영화 하나의 상세 통계 엑셀 다운로드 → xlsx bytes."""
    r = session.post(
        f"{BASE}/kobis/business/mast/thea/findCompanyStatDetailXls.do",
        data={"CSRFToken": csrf, "movieCd": movie_cd, "loadEnd": "0",
              "sStartDt": start_dt, "sEndDt": end_dt},
        timeout=300,
    )
    if b"<Workbook" not in r.content[:2000]:
        raise RuntimeError("상세 엑셀 응답 형식이 예상과 다릅니다(세션 만료 가능)")
    return _xml_to_xlsx(r.content)


def _filter_movies(movies, includes, excludes):
    includes = [i for i in (includes or []) if i]
    excludes = [e for e in (excludes or []) if e]
    out = []
    for mv in movies:
        nm = mv["movieNm"]
        if includes and not any(kw in nm for kw in includes):
            continue
        if any(ex in nm for ex in excludes):
            continue
        out.append(mv)
    return out


def _safe_name(s):
    return re.sub(r'[\\/:*?"<>|]', "_", s).strip()


def _validate_period(start_dt, end_dt):
    d0 = datetime.strptime(start_dt, "%Y-%m-%d").date()
    d1 = datetime.strptime(end_dt, "%Y-%m-%d").date()
    if d1 < d0:
        raise RuntimeError("기간이 잘못 입력되었습니다.")
    if (d1 - d0).days > 31:
        raise RuntimeError("조회 기간은 1달을 초과할 수 없습니다(KOBIS 제한).")


def get_accounts():
    """수집에 사용할 배급사 계정 목록 (DB 활성 계정 우선, 없으면 기본 목록)."""
    try:
        from .models import KobisDistributorAccount
        rows = list(KobisDistributorAccount.objects.filter(is_active=True))
        if rows:
            return [{"name": r.name, "user": r.user, "password": r.password,
                     "aprv_no": r.aprv_no} for r in rows]
    except Exception:
        pass
    from .kobis_accounts import KOBIS_ACCOUNTS
    return KOBIS_ACCOUNTS


def crawl_one_account(name, user, password, aprv_no, start_dt, end_dt,
                      includes, excludes):
    """한 배급사 계정 수집 → {name, ok, error, movies[{..filename, xlsx}]}."""
    res = {"name": name, "ok": False, "error": "", "movies": []}
    try:
        s = requests.Session()
        s.headers["User-Agent"] = _UA
        _login(s, user, password, aprv_no)
        csrf = _fetch_csrf(s)
        movies = _filter_movies(
            fetch_movie_list(s, csrf, start_dt, end_dt), includes, excludes)
        for mv in movies:
            item = dict(mv)
            try:
                item["xlsx"] = download_movie_detail_xlsx(
                    s, csrf, mv["movieCd"], start_dt, end_dt)
                item["filename"] = (
                    f"영진위_{_safe_name(name)}_{_safe_name(mv['movieNm'])}"
                    f"_{start_dt}_{end_dt}.xlsx")
                item["error"] = ""
            except Exception as e:
                item["xlsx"] = None
                item["filename"] = None
                item["error"] = str(e)[:200]
            res["movies"].append(item)
        res["ok"] = True
    except Exception as e:
        res["error"] = str(e)[:200]
    return res


def crawl_all_accounts(start_dt, end_dt, includes, excludes, max_workers=4):
    """모든 배급사 계정 수집 → summary 리스트 (계정 순서 유지).

    각 항목: {name, ok, error, movies:[{movieCd, movieNm, theaters, visitors,
             filename, xlsx(bytes|None), error}]}
    """
    end_dt = end_dt or start_dt
    _validate_period(start_dt, end_dt)
    accounts = get_accounts()
    order = {a["name"]: i for i, a in enumerate(accounts)}
    results = []
    with ThreadPoolExecutor(max_workers=max_workers) as ex:
        futs = [
            ex.submit(crawl_one_account, a["name"], a["user"], a["password"],
                      a.get("aprv_no") or "", start_dt, end_dt, includes, excludes)
            for a in accounts
        ]
        for f in as_completed(futs):
            results.append(f.result())
    results.sort(key=lambda r: order.get(r["name"], 999))
    return results
