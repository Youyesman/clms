import os
import re
import logging
from datetime import datetime
from collections import defaultdict
from django.utils import timezone as dj_timezone

import pandas as pd
from django.conf import settings
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter

from crawler.models import MovieSchedule
from client.models import Client

logger = logging.getLogger(__name__)


def export_schedules_to_excel(start_date_str, end_date_str, companies=None, target_titles=None, failures=None):
    """
    Exports MovieSchedule data and Failure logs to an Excel file.
    """
    save_dir = os.path.join(settings.BASE_DIR, 'media', 'crawler_exports')
    os.makedirs(save_dir, exist_ok=True)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"crawler_result_{timestamp}.xlsx"
    file_path = os.path.join(save_dir, filename)

    with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
        if failures:
            df_all = pd.DataFrame(failures)
            desired_order = ['brand', 'region', 'theater', 'date', 'reason', 'worker']
            cols = [c for c in desired_order if c in df_all.columns] + [c for c in df_all.columns if c not in desired_order]
            df_all = df_all[cols]
            unique_dates = df_all['date'].unique()
            for d_val in unique_dates:
                sheet_name_safe = str(d_val).replace('/', '').replace('\\', '').replace(':', '')
                sheet_title = f"Fail_{sheet_name_safe}"[:30]
                df_sub = df_all[df_all['date'] == d_val]
                df_sub.to_excel(writer, sheet_name=sheet_title, index=False)
        else:
            df_summary = pd.DataFrame([{
                '결과': '수집 완료',
                '수집 기간': f"{start_date_str} ~ {end_date_str}",
                '수집 대상': ', '.join(companies) if companies else '전체',
                '실패 건수': 0,
                '비고': '모든 극장 데이터 정상 수집됨'
            }])
            df_summary.to_excel(writer, sheet_name='수집결과', index=False)

    return file_path


# ===== Styles =====
BLUE_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
YELLOW_LIGHT = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
BLACK_FILL = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
PINK_FILL = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
TOTAL_BLUE_FILL = PatternFill(start_color="0033CC", end_color="0033CC", fill_type="solid")

THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

CENTER = Alignment(horizontal='center', vertical='center')
LEFT = Alignment(horizontal='left', vertical='center')

# 소계·총계 줄은 반드시 '숫자'로 기록한다(텍스트로 쓰면 엑셀 수식/자동합계에서 집계되지 않음)
NUM_FMT = '#,##0'
PCT_FMT = '0.0%'

TITLE_FONT = Font(name='맑은 고딕', size=14, bold=True, color="0000FF")
INFO_FONT = Font(name='맑은 고딕', size=10, color="FF0000")
HEADER_FONT = Font(name='맑은 고딕', size=10, bold=True)
WHITE_FONT = Font(name='맑은 고딕', size=10, bold=True, color="FFFFFF")
DATA_FONT = Font(name='맑은 고딕', size=10)
BOLD_FONT = Font(name='맑은 고딕', size=10, bold=True)
# 일반극장 매핑 안 됨 강조 (빨강)
UNMAPPED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
UNMAPPED_FONT = Font(name='맑은 고딕', size=10, bold=True, color="9C0006")


def _safe_int(val):
    try:
        return int(val)
    except:
        return 0


def _fmt_number(n):
    """Format number with comma separator as string."""
    if isinstance(n, float):
        return f"{n:,.1f}"
    return f"{int(n):,}"


def _auto_width(ws, min_row=1):
    """Auto-adjust column widths based on content."""
    import re as _re

    def _fmt_suffix_width(number_format):
        """C005: 표시 서식에 붙는 리터럴 텍스트('#,##0" 극장"'의 ' 극장' 등)의
        표시 폭. 숫자 자릿수만으로 폭을 잡으면 접미어가 붙은 '합' 라인 수치가
        ######으로 잘리므로, 따옴표 안 리터럴의 폭을 더한다."""
        if not number_format or number_format in ("General",):
            return 0
        w = 0
        for lit in _re.findall(r'"([^"]*)"', str(number_format)):
            w += sum(2 if ord(c) > 127 else 1 for c in lit)
        return w

    for i, col_cells in enumerate(ws.columns, 1):
        max_len = 0
        for cell in col_cells:
            if cell.row < min_row:
                continue
            try:
                v = cell.value
                if v is None or v == "":
                    continue
                # M004: 실제 표시 폭 기준으로 계산 — 수식 문자열('=SUM…')과
                # 백분율 float의 원시값 길이가 폭을 부풀리지 않게 한다
                if isinstance(v, str) and v.startswith("="):
                    continue
                if isinstance(v, float):
                    width = 7  # 0.0% / #,##0 표시 서식 기준
                    width += _fmt_suffix_width(cell.number_format)
                elif isinstance(v, int):
                    width = len(f"{v:,}")
                    width += _fmt_suffix_width(cell.number_format)
                else:
                    val = str(v)
                    width = sum(2 if ord(c) > 127 else 1 for c in val)
                max_len = max(max_len, width)
            except:
                pass
        # M004: 페이지가 한눈에 들어오게 — 여유는 살리되 과하게 넓어지지 않게 조정
        # C005: 굵은 글씨·서식 여유분으로 한 글자만큼 더 준다 (#### 방지)
        adj = (max_len + 2) * 1.08
        ws.column_dimensions[get_column_letter(i)].width = min(max(adj, 6), 36)


def _build_region_map():
    """Build (brand, theater_name) -> region mapping from Client model."""
    clients = Client.objects.filter(excel_theater_name__isnull=False).values(
        'theater_kind', 'excel_theater_name', 'theater_name', 'client_name', 'region_code'
    )

    region_map = {}

    def normalize_brand(kind):
        if not kind:
            return None
        k = kind.upper()
        if 'CGV' in k: return 'CGV'
        if 'LOTTE' in k or '롯데' in k: return 'LOTTE'
        if 'MEGA' in k or '메가' in k: return 'MEGABOX'
        if 'CINEQ' in k or '씨네큐' in k: return 'CINEQ'
        return kind

    for c in clients:
        brand = normalize_brand(c['theater_kind'])
        if not brand or not c['region_code']:
            continue
        region = c['region_code']
        if c['excel_theater_name']:
            region_map[(brand, c['excel_theater_name'].replace(" ", ""))] = region
        if c['theater_name']:
            region_map[(brand, c['theater_name'].replace(" ", ""))] = region

    return region_map


def _resolve_region(brand, theater, region_map):
    """Resolve region for a theater using fuzzy matching."""
    if not brand or not theater:
        return "-"
    clean = theater.replace(" ", "")
    if (brand, clean) in region_map:
        return region_map[(brand, clean)]

    def strip_brand(s):
        return s.replace("CGV", "").replace("롯데시네마", "").replace("롯데", "").replace("메가박스", "").replace("씨네큐", "").replace(" ", "")

    crawl_pure = strip_brand(theater)
    for (m_brand, m_name), m_region in region_map.items():
        if m_brand != brand:
            continue
        client_pure = strip_brand(m_name)
        if crawl_pure == client_pure:
            return m_region
        if len(client_pure) >= 2 and len(crawl_pure) >= 2:
            if client_pure in crawl_pure or crawl_pure in client_pure:
                return m_region
    return "-"


def _format_theater_name(brand, theater):
    """Format theater name with brand prefix."""
    clean = theater.replace("CGV", "").replace("롯데시네마", "").replace("롯데", "").replace("메가박스", "").replace("씨네큐", "").strip()
    if brand == 'CGV': return f"CGV {clean}"
    if brand == 'LOTTE': return f"롯데{clean}"
    if brand == 'MEGABOX': return f"메가박스{clean}"
    if brand == 'CINEQ': return f"씨네큐{clean}"
    return f"{brand} {clean}"


def _norm_aud_key(name):
    """
    상영관 이름 정규화 키 (영진위관 ↔ DB관 매칭용).
    - 크롤된 screen_name 과 동일하게 괄호/메타 제거 + 'N관' 추출
      (예: 'DMZ관 [상서면관]' -> 'DMZ관', '5관(리클라이너)' -> '5관')
    - 선행 0 제거 ('04관' -> '4관') 후 공백제거·소문자
    """
    s = MovieSchedule.normalize_screen_name(name)
    m = re.match(r'0*(\d+)\s*관$', s)
    if m:
        s = f"{int(m.group(1))}관"
    return re.sub(r'\s+', '', s).lower()


def _build_normal_theater_index():
    """
    일반극장(KOBIS) 행 보강용 인덱스.
    - name_to_client: 정규화된 (영진위극장명 우선 → 극장명) -> Client
    - seat_idx: client_id -> {kofic/name/num: 관 -> 좌석수}
    """
    from client.models import Client, Theater

    qs = Client.objects.exclude(theater_kind__in=['CGV', '메가박스', '롯데']).only(
        'id', 'client_name', 'kofic_theater_name', 'kofic_theater_name2', 'region_code'
    )
    name_to_client = {}
    # 1) 영진위극장명(1·2) 우선
    for c in qs:
        for nm in (c.kofic_theater_name, c.kofic_theater_name2):
            if nm:
                name_to_client.setdefault(re.sub(r'\s+', '', nm).lower(), c)
    # 2) 그 다음 극장명
    for c in qs:
        if c.client_name:
            name_to_client.setdefault(re.sub(r'\s+', '', c.client_name).lower(), c)

    seat_idx = {}
    for t in Theater.objects.all().only(
        'client_id', 'auditorium_name', 'kofic_auditorium_name', 'seat_count'
    ):
        d = seat_idx.setdefault(t.client_id, {'kofic': {}, 'name': {}, 'num': {}})
        # 영진위관이름·DB관이름 모두 동일 정규화로 색인 + 'N관' 숫자 색인(둘 다)
        for raw, bucket in ((t.kofic_auditorium_name, 'kofic'), (t.auditorium_name, 'name')):
            if not raw:
                continue
            key = _norm_aud_key(raw)
            if key:
                d[bucket].setdefault(key, t.seat_count)
            m = re.search(r'(\d+)', raw)
            if m:
                d['num'].setdefault(int(m.group(1)), t.seat_count)
    return name_to_client, seat_idx


def _resolve_normal_theater(theater_name, screen_name, normal_index):
    """일반극장 행: 영진위극장명→극장명 순으로 Client 매칭. 미매칭이면 None."""
    name_to_client, seat_idx = normal_index
    norm = re.sub(r'\s+', '', str(theater_name or '')).lower()
    c = name_to_client.get(norm)
    if not c:
        return None
    region = c.region_code or '-'
    seat = 0
    d = seat_idx.get(c.id, {})
    sn = _norm_aud_key(screen_name)
    if sn in d.get('kofic', {}):           # 1) 영진위관이름 매칭 우선
        seat = d['kofic'][sn]
    elif sn in d.get('name', {}):          # 2) DB관이름 매칭
        seat = d['name'][sn]
    else:                                  # 3) 'N관' 숫자 매칭 (영진위/DB관 공통)
        m = re.search(r'(\d+)', str(screen_name or ''))
        if m and int(m.group(1)) in d.get('num', {}):
            seat = d['num'][int(m.group(1))]
    return {'region': region, 'seat': _safe_int(seat), 'client_name': c.client_name}


# E004: 특별 상영 포맷 canonical 태그 표시 순서 (MovieSchedule.extract_format_tags 가 저장)
SPECIAL_FORMAT_ORDER = ["IMAX", "4DX", "SUPER-4D", "MX4D", "SCREENX", "DOLBY", "ATMOS", "3D"]


def schedule_matches_formats(tags, wanted):
    """U002: 선택한 특별 상영 포맷에 이 회차가 해당하는지.

    E004 canonical 태그(IMAX·4DX·DOLBY …) 정확 일치를 먼저 보고,
    E004 이전에 수집돼 원문이 그대로 남은 태그('DOLBYCINEMA', '디지털 3D' 등)도
    걸러지도록 부분 문자열까지 본다. (_extract_format_and_type 과 같은 2단 규칙)
    """
    tag_uppers = [str(t).upper() for t in (tags or [])]
    if not tag_uppers:
        return False
    return any(w in tag_uppers or any(w in t for t in tag_uppers) for w in wanted)


def _extract_format_and_type(tags):
    """Extract format (2D/IMAX/DOLBY...) and sub_type (일반/자막/더빙) from tags."""
    tag_uppers = [str(t).upper() for t in (tags or [])]
    tag_set = set(tag_uppers)
    # canonical 포맷 태그(E004) 우선 — 복합 포맷은 'IMAX 3D'처럼 이어 붙인다
    fmt_parts = [f for f in SPECIAL_FORMAT_ORDER if f in tag_set]
    if not fmt_parts:
        # 구버전 태그(제목 괄호에서 추출된 원문 등) 호환: 키워드 포함 검사
        for t in tag_uppers:
            if any(f in t for f in ("IMAX", "4DX", "SCREENX", "DOLBY", "ATMOS")):
                fmt_parts = [t]
                break
    fmt = " ".join(fmt_parts) if fmt_parts else "2D"
    sub_type = _extract_sub_type(tags)
    return fmt, sub_type


def _extract_sub_type(tags):
    """C005: 구분(더빙/자막/일반) 판별.

    canonical 태그('더빙'/'자막') 외에 영진위식 복합 표기('디지털 더빙',
    '디지털 영문자막')도 부분 문자열로 판별한다.
    """
    for t in (tags or []):
        if "더빙" in str(t):
            return "더빙"
    for t in (tags or []):
        if "자막" in str(t):
            return "자막"
    return "일반"


def _brand_priority(b):
    b = (b or '').upper()
    if 'CGV' in b: return 1
    if 'LOTTE' in b: return 2
    if 'MEGA' in b: return 3
    if '일반극장' in b: return 4
    return 99


def _brand_display(b):
    b = (b or '').upper()
    if 'CGV' in b: return 'CGV 계'
    if 'LOTTE' in b: return 'LOTTE 계'
    if 'MEGA' in b: return 'MEGA 계'
    if '일반극장' in b: return '일반극장 계'
    return b


# 요약·비교 시트에 행으로 표기되는 멀티(계열사) 순서. 일반극장을 빼면 각 시트의
# 행 합계와 '합' 행이 어긋나므로 반드시 포함한다.
BRAND_ORDER = ['CGV', 'LOTTE', 'MEGABOX', '일반극장']

# 상영시간표 시트의 멀티별 소계 라벨
SCHEDULE_SUBTOTAL_LABEL = {
    'CGV': 'CGV소계',
    'LOTTE': 'LOTTE소계',
    'MEGABOX': 'MEGA소계',
    '일반극장': '일반극장 소계',
}


CINE_DE_CHEF_MAP = {
    "압구정": "압구정",
    "용산": "용산아이파크몰",
    "센텀": "센텀시티",
}


def _filter_cine_de_chef(schedules):
    """씨네드쉐프 극장 처리: 원본 극장이 있으면 씨네드쉐프 제거, 없으면 원본명으로 치환."""
    all_theaters = set()
    cine_de_chef_items = []
    normal_items = []

    for sch in schedules:
        if sch.brand and sch.brand.upper() == "CGV" and "씨네드쉐프" in (sch.theater_name or ""):
            cine_de_chef_items.append(sch)
        else:
            normal_items.append(sch)
            if sch.brand and sch.brand.upper() == "CGV":
                all_theaters.add(sch.theater_name or "")

    for sch in cine_de_chef_items:
        # "CGV 씨네드쉐프 센텀" → "센텀" 추출
        raw_name = (sch.theater_name or "").replace("CGV", "").replace("씨네드쉐프", "").strip()
        mapped_name = CINE_DE_CHEF_MAP.get(raw_name)
        if not mapped_name:
            # 매핑에 없으면 그대로 유지
            normal_items.append(sch)
            continue

        # 원본 극장이 데이터에 존재하면 씨네드쉐프 제거 (skip)
        original_candidates = [f"CGV {mapped_name}", f"CGV{mapped_name}", mapped_name]
        if any(t.replace(" ", "") in {c.replace(" ", "") for c in all_theaters} for t in original_candidates):
            continue  # 원본 있음 → 씨네드쉐프 제거

        # 원본 없음 → 극장명을 원본으로 치환
        sch.theater_name = f"CGV {mapped_name}"
        normal_items.append(sch)

    # C003: 같은 특별관이 'CINE de CHEF관'/'CINE de CHEF B관' 표기(좌석 정보 없음)로도
    # 내려와 '스트레스리스 시네마'/'템퍼 시네마' 행과 중복된다.
    # 좌석수를 읽을 수 없는 CINE de CHEF 표기 회차는 표에서 제거한다.
    return [
        sch for sch in normal_items
        if not (sch.brand and sch.brand.upper() == "CGV"
                and re.search(r'cine\s*de\s*chef', str(sch.screen_name or ''), re.I)
                and not _safe_int(sch.total_seats))
    ]


def _process_to_rows(schedules, region_map, normal_index=None):
    """Process schedule queryset into structured rows for display and aggregation."""
    schedules = _filter_cine_de_chef(list(schedules))
    if normal_index is None:
        normal_index = _build_normal_theater_index()

    grouped = {}
    for sch in schedules:
        s_date = sch.play_date or sch.start_time.date()
        # C005: 같은 관이라도 더빙/자막 구분이 다르면 별도 행으로 분리한다
        key = (sch.brand, sch.theater_name, sch.screen_name, s_date,
               _extract_sub_type(sch.tags))
        if key not in grouped:
            grouped[key] = []
        grouped[key].append(sch)

    rows = []
    max_shows = 0

    for (brand, theater, screen, s_date, sub_type), items in grouped.items():
        items.sort(key=lambda x: x.start_time)

        region = _resolve_region(brand, theater, region_map)
        display_theater = _format_theater_name(brand, theater)
        fmt, _ = _extract_format_and_type(items[0].tags if items else [])

        total_capacity = _safe_int(items[0].total_seats) if items else 0
        show_count = len(items)

        # 일반극장(KOBIS): '일반극장' 접두 제거 + 영진위극장명/극장명 매칭으로 지역·좌석수 보강
        unmapped = False
        if brand == '일반극장':
            info = _resolve_normal_theater(theater, screen, normal_index)
            if info:
                display_theater = info['client_name']
                region = info['region']
                if info['seat']:
                    total_capacity = info['seat']
            else:
                display_theater = theater  # 접두어 없이 원본 극장명
                region = '매핑안됨'
                unmapped = True

        total_seats_sum = 0
        sold_seats_sum = 0
        show_times = []
        # V001/V002(0828): 회차 단위 지표(골든타임·특별관 등)를 쓰는 화면이
        # 자체 계산을 따로 하지 않도록 회차별 상세도 함께 내보낸다.
        shows_detail = []

        for item in items:
            hhmm = dj_timezone.localtime(item.start_time).strftime("%H:%M")
            show_times.append(hhmm)
            raw_seat = _safe_int(item.total_seats)
            r_seat = _safe_int(item.remaining_seats)
            # 좌석 정보가 없는 회차(일반극장/KOBIS 등)는 극장 정원으로 총좌석수만 채우고,
            # 잔여좌석을 모르므로 판매좌석수는 0으로 둔다. (예전엔 정원 전체가 판매된 것으로
            # 집계돼 소계·총계의 판매좌석수와 좌판율이 부풀려졌다.)
            t_seat = raw_seat if raw_seat > 0 else total_capacity
            i_sold = max(0, t_seat - r_seat) if raw_seat > 0 else 0
            total_seats_sum += t_seat
            sold_seats_sum += i_sold
            shows_detail.append({
                'time': hhmm,
                'hour': dj_timezone.localtime(item.start_time).hour,
                'seats': t_seat,
                'sold': i_sold,
                'tags': list(item.tags or []),
            })

        max_shows = max(max_shows, len(show_times))

        rows.append({
            'brand': brand,
            'region': region,
            'theater': display_theater,
            'format': fmt,
            'sub_type': sub_type,
            'screen': screen,
            'capacity': total_capacity,
            'show_times': show_times,
            'shows': shows_detail,
            'show_count': show_count,
            'total_seats': total_seats_sum,
            'sold_seats': sold_seats_sum,
            'unmapped': unmapped,
        })

    # C005: 더빙과 자막이 함께 상영되는 영화는, 구분 표기가 없는 회차를
    # '일반' 대신 '자막'으로 표기한다 (더빙이 아예 없는 영화는 '일반' 유지).
    # _process_to_rows 는 항상 한 작품 단위로 호출되므로 rows 전체 기준으로 판단한다.
    if any(r['sub_type'] == '더빙' for r in rows):
        for r in rows:
            if r['sub_type'] == '일반':
                r['sub_type'] = '자막'

    rows.sort(key=lambda x: (
        _brand_priority(x['brand']),
        x['region'],
        x['theater'],
        x['screen'],
        x['sub_type']
    ))
    return rows, max_shows


def _calc_summary(rows):
    """Calculate summary statistics from a list of rows."""
    theaters = set((r['brand'], r['theater']) for r in rows)
    theater_count = len(theaters)
    show_count = sum(r['show_count'] for r in rows)
    screen_count = len(rows)
    capacity = sum(r['capacity'] for r in rows)
    total_seats = sum(r['total_seats'] for r in rows)
    sold_seats = sum(r['sold_seats'] for r in rows)

    avg_shows = round(show_count / theater_count, 1) if theater_count else 0
    avg_seats = round(total_seats / show_count, 1) if show_count else 0
    avg_sold_rate = round(sold_seats / total_seats * 100, 1) if total_seats else 0

    return {
        'theater_count': theater_count,
        'show_count': show_count,
        'screen_count': screen_count,
        'capacity': capacity,
        'total_seats': total_seats,
        'sold_seats': sold_seats,
        'avg_shows': avg_shows,
        'avg_seats': avg_seats,
        'avg_sold_rate': avg_sold_rate,
    }


def _write_subtotal_row(ws, ri, max_col, label, fill, font,
                        data_rows=None, subtotal_rows=None):
    """상영시간표 시트의 소계/총계 행.

    값이 아니라 **엑셀 수식**으로 기록한다. 사용자가 시트에서 바로 검산할 수 있고,
    행을 지우거나 필터를 걸어도 합계가 따라간다. (예전엔 '1,234' 문자열이라
    엑셀이 텍스트로 취급해 수식·자동합계에서 아예 집계되지 않았다.)

    data_rows     : (시작행, 끝행) — 브랜드 데이터 행 범위 → 소계
    subtotal_rows : [소계행, ...]  — 소계 행들을 더해 총계
    """
    # Column indices (1-based): C1=지역(극장수), C2=극장명(label), then stats at end
    stat_start = max_col - 3  # 총회차, 총스크린, 총좌석수, 판매좌석수

    for ci in range(1, max_col + 1):
        cell = ws.cell(row=ri, column=ci)
        cell.fill = fill
        cell.font = font
        cell.border = THIN_BORDER
        cell.alignment = CENTER

    ws.cell(row=ri, column=2, value=label)

    def _ref(letter):
        if data_rows:
            start, end = data_rows
            return f"{letter}{start}:{letter}{end}"
        return ",".join(f"{letter}{r}" for r in subtotal_rows)

    # 극장수: 극장명(B) 열은 같은 극장끼리 병합돼 그룹당 한 칸만 값이 남으므로
    # COUNTA 가 곧 극장 수. 총계 행은 소계들의 합.
    theater_cell = ws.cell(row=ri, column=1)
    theater_cell.value = (f"=COUNTA({_ref('B')})" if data_rows
                          else f"=SUM({_ref('A')})")
    theater_cell.number_format = NUM_FMT

    for off in range(4):
        col = stat_start + off
        cell = ws.cell(row=ri, column=col,
                       value=f"=SUM({_ref(get_column_letter(col))})")
        cell.number_format = NUM_FMT


def _write_schedule_sheet(ws, rows, proc_date, movie_title, display_max_shows, gen_info):
    """Write a single schedule sheet (상영시간표)."""
    date_str = proc_date.strftime("%Y-%m-%d")

    # Column definitions
    base_cols = ['지역', '극장명', '포맷', '구분', '관', '좌석수']
    show_cols = [f"{i+1}회" for i in range(display_max_shows)]
    stat_cols = ['총회차', '총스크린', '총좌석수', '판매좌석수']
    all_cols = base_cols + show_cols + stat_cols
    max_col = len(all_cols)

    SUBTOTAL_FONT = Font(name='맑은 고딕', size=10, bold=True, color="000000")

    # Row 2: Movie Title (Row 1 empty for spacing)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max_col)
    title_cell = ws.cell(row=2, column=1, value=movie_title or "전체 영화")
    title_cell.font = TITLE_FONT
    title_cell.alignment = CENTER
    for ci in range(1, max_col + 1):
        c = ws.cell(row=2, column=ci)
        c.border = THIN_BORDER
        c.fill = YELLOW_FILL

    # Row 3: 날짜만 가운데 표기 (M004: 생성 시간 표기 삭제)
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=max_col)
    date_cell = ws.cell(row=3, column=1, value=date_str)
    date_cell.font = SUBTOTAL_FONT
    date_cell.alignment = CENTER

    # Row 4: Headers
    for ci, col_name in enumerate(all_cols, 1):
        cell = ws.cell(row=4, column=ci, value=col_name)
        cell.border = THIN_BORDER
        cell.alignment = CENTER
        if col_name in ['총좌석수', '판매좌석수']:
            cell.fill = BLACK_FILL
            cell.font = WHITE_FONT
        else:
            cell.fill = YELLOW_FILL
            cell.font = HEADER_FONT

    # Group rows by brand for subtotals
    brand_groups = {}
    for row in rows:
        b = row['brand']
        if b not in brand_groups:
            brand_groups[b] = []
        brand_groups[b].append(row)

    # Write data rows grouped by brand, with subtotal after each brand
    ri = 5
    ordered_brands = [b for b in BRAND_ORDER if b in brand_groups]
    other_brands = [b for b in brand_groups if b not in BRAND_ORDER]

    subtotal_row_nums = []  # 총 계 행에서 참조할 소계 행 번호

    for brand in ordered_brands + other_brands:
        brand_rows = brand_groups[brand]
        merge_start = ri

        for row in brand_rows:
            vals = [
                row['region'], row['theater'], row['format'], row['sub_type'],
                row['screen'], row['capacity']
            ]
            for i in range(display_max_shows):
                vals.append(row['show_times'][i] if i < len(row['show_times']) else "")
            vals.extend([row['show_count'], 1, row['total_seats'], row['sold_seats']])

            for ci, val in enumerate(vals, 1):
                cell = ws.cell(row=ri, column=ci, value=val)
                cell.border = THIN_BORDER
                cell.alignment = CENTER
                cell.font = DATA_FONT
                # 좌석수 / 총회차·총스크린·총좌석수·판매좌석수 → 합계줄과 동일한 숫자 서식
                if ci == 6 or ci > max_col - 4:
                    cell.number_format = NUM_FMT


            # 매핑 안 된 일반극장 행: 지역·극장명 셀을 빨갛게 강조
            if row.get('unmapped'):
                for ci in (1, 2):
                    c = ws.cell(row=ri, column=ci)
                    c.fill = UNMAPPED_FILL
                    c.font = UNMAPPED_FONT

            ri += 1

        # Merge theater name cells (column B) within this brand group
        if len(brand_rows) > 1:
            merge_b_start = merge_start
            for check_ri in range(merge_start + 1, ri + 1):
                current = ws.cell(row=check_ri, column=2).value if check_ri < ri else None
                prev = ws.cell(row=check_ri - 1, column=2).value
                if current != prev or check_ri >= ri:
                    if check_ri - 1 > merge_b_start:
                        ws.merge_cells(start_row=merge_b_start, start_column=2, end_row=check_ri - 1, end_column=2)
                    merge_b_start = check_ri

        # 브랜드 소계 — 해당 브랜드 데이터 행 범위를 참조하는 수식
        label = SCHEDULE_SUBTOTAL_LABEL.get(brand, "기타 소계")
        _write_subtotal_row(ws, ri, max_col, label,
                            PINK_FILL, SUBTOTAL_FONT,
                            data_rows=(merge_start, ri - 1))
        subtotal_row_nums.append(ri)

        ri += 1

    # 총 계 — 소계 행들을 더한다 (데이터 범위를 다시 합산하면 소계까지 이중 집계됨)
    _write_subtotal_row(ws, ri, max_col, "총 계",
                        TOTAL_BLUE_FILL, WHITE_FONT,
                        subtotal_rows=subtotal_row_nums)

    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A5"  # M004: 4행(헤더)까지 틀고정
    _auto_width(ws, min_row=4)


def _write_comparison_sheet(ws, main_data, competitor_data_dict, movie_title, gen_info, brand_presence=None):
    """Write 비교표(집계작 및 경쟁작 멀티별 비교) sheet.

    brand_presence: {(date, brand): bool} — 해당 날짜×계열사에 크롤 데이터가
    한 건이라도 존재하는지. 없으면(수집 시차) '미수집'으로 표기해
    '상영 없음'과 구분한다 (C001).
    """
    # Movie order: main first, then competitors (E006: 주요작 없으면 경쟁작만)
    movie_titles = ([movie_title] if main_data else []) + list(competitor_data_dict.keys())
    # U001: 모든 작품(주요작·경쟁작)에 극장수 칸을 두고, '상영관'은 '스크린수'로 표기한다.
    cols_per_movie = 5  # 극장수, 스크린수, 회차, 총좌석수, 평균좌판율
    fixed_cols = 2  # 상영일자, 영화명(계열사)

    total_cols = fixed_cols + len(movie_titles) * cols_per_movie

    # Row 1: Movie title headers (merged per movie block)
    # M004: 1행 A열은 비우고 B열에 '영화명' 표기
    ws.cell(row=1, column=1, value="").fill = BLUE_FILL
    ws.cell(row=1, column=1).border = THIN_BORDER
    ws.cell(row=1, column=2, value="영화명").fill = BLUE_FILL
    ws.cell(row=1, column=2).font = HEADER_FONT
    ws.cell(row=1, column=2).border = THIN_BORDER
    ws.cell(row=1, column=2).alignment = CENTER

    for mi, mt in enumerate(movie_titles):
        start_col = fixed_cols + 1 + mi * cols_per_movie
        end_col = start_col + cols_per_movie - 1
        ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)
        cell = ws.cell(row=1, column=start_col, value=mt)
        cell.fill = BLUE_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        for ci in range(start_col, end_col + 1):
            ws.cell(row=1, column=ci).border = THIN_BORDER
            ws.cell(row=1, column=ci).fill = BLUE_FILL

    # Row 2: Sub-headers (repeating per movie)
    sub_headers = ['극장수', '스크린수', '회차', '총좌석수', '평균좌판율']
    ws.cell(row=2, column=1, value="상영일자").fill = BLUE_FILL
    ws.cell(row=2, column=1).font = HEADER_FONT
    ws.cell(row=2, column=1).border = THIN_BORDER
    ws.cell(row=2, column=1).alignment = CENTER
    # M004: 2행 B열은 '멀티'(계열사 구분) 표기
    ws.cell(row=2, column=2, value="멀티").fill = BLUE_FILL
    ws.cell(row=2, column=2).font = HEADER_FONT
    ws.cell(row=2, column=2).border = THIN_BORDER
    ws.cell(row=2, column=2).alignment = CENTER

    for mi in range(len(movie_titles)):
        for si, sh in enumerate(sub_headers):
            col = fixed_cols + 1 + mi * cols_per_movie + si
            cell = ws.cell(row=2, column=col, value=sh)
            cell.fill = BLUE_FILL
            cell.font = HEADER_FONT
            cell.border = THIN_BORDER
            cell.alignment = CENTER

    # Build data: all_movie_data[movie_title] -> {date -> rows}
    all_movie_data = {}
    if main_data:
        all_movie_data[movie_title] = main_data
    for ct, c_data in competitor_data_dict.items():
        all_movie_data[ct] = c_data

    # 날짜: 주요작·경쟁작 전체의 날짜 합집합 (E006: 주요작 없어도 동작)
    dates = sorted(set(d for m_data in all_movie_data.values() for d in m_data.keys()))

    row_idx = 3
    for proc_date in dates:
        date_str = proc_date.strftime("%Y-%m-%d")
        date_start_row = row_idx

        for brand in BRAND_ORDER:
            # Write date + brand
            ws.cell(row=row_idx, column=1, value=date_str if row_idx == date_start_row else '')
            ws.cell(row=row_idx, column=1).border = THIN_BORDER
            ws.cell(row=row_idx, column=1).alignment = CENTER
            if row_idx == date_start_row:
                ws.cell(row=row_idx, column=1).fill = BLUE_FILL
                ws.cell(row=row_idx, column=1).font = BOLD_FONT

            ws.cell(row=row_idx, column=2, value=_brand_display(brand))
            ws.cell(row=row_idx, column=2).border = THIN_BORDER
            ws.cell(row=row_idx, column=2).alignment = CENTER
            ws.cell(row=row_idx, column=2).font = BOLD_FONT

            # For each movie, write stats
            for mi, mt in enumerate(movie_titles):
                m_data = all_movie_data.get(mt, {})
                m_rows = m_data.get(proc_date, [])
                brand_rows = [r for r in m_rows if r['brand'] == brand]

                base_col = fixed_cols + 1 + mi * cols_per_movie
                if brand_rows:
                    s = _calc_summary(brand_rows)
                    vals = [
                        (s['theater_count'], NUM_FMT),
                        (s['screen_count'], NUM_FMT),
                        (s['show_count'], NUM_FMT),
                        (s['total_seats'], NUM_FMT),
                        (s['sold_seats'] / s['total_seats'] if s['total_seats'] else 0, PCT_FMT),
                    ]
                elif brand_presence is not None and not brand_presence.get((proc_date, brand), True):
                    # 크롤 데이터 자체가 없는 날짜×계열사 (C001)
                    # M004: '미수집' 글자 대신 숫자 0으로 표기
                    vals = [(0, NUM_FMT)] + [('', None)] * (cols_per_movie - 1)
                else:
                    vals = [('', None)] * cols_per_movie

                for si, (v, fmt) in enumerate(vals):
                    cell = ws.cell(row=row_idx, column=base_col + si, value=v)
                    cell.border = THIN_BORDER
                    cell.alignment = CENTER
                    cell.font = DATA_FONT
                    if fmt:
                        cell.number_format = fmt

            row_idx += 1

        # 합 row
        ws.cell(row=row_idx, column=1, value='')
        ws.cell(row=row_idx, column=1).border = THIN_BORDER
        ws.cell(row=row_idx, column=2, value='합')
        ws.cell(row=row_idx, column=2).border = THIN_BORDER
        ws.cell(row=row_idx, column=2).alignment = CENTER
        ws.cell(row=row_idx, column=2).fill = YELLOW_LIGHT
        ws.cell(row=row_idx, column=2).font = BOLD_FONT

        for mi, mt in enumerate(movie_titles):
            m_data = all_movie_data.get(mt, {})
            m_rows = m_data.get(proc_date, [])
            base_col = fixed_cols + 1 + mi * cols_per_movie

            if m_rows:
                s = _calc_summary(m_rows)
                # 단위(극장/스크린/회/석)는 표시서식으로 붙인다 — 값 자체는 숫자여야 수식이 먹는다
                vals = [
                    (s['theater_count'], '#,##0" 극장"'),
                    (s['screen_count'], '#,##0" 스크린"'),
                    (s['show_count'], '#,##0" 회"'),
                    (s['total_seats'], '#,##0" 석"'),
                    (s['sold_seats'] / s['total_seats'] if s['total_seats'] else 0, PCT_FMT),
                ]
            else:
                vals = [('', None)] * cols_per_movie

            for si, (v, fmt) in enumerate(vals):
                cell = ws.cell(row=row_idx, column=base_col + si, value=v)
                cell.border = THIN_BORDER
                cell.alignment = CENTER
                cell.fill = YELLOW_LIGHT
                cell.font = BOLD_FONT
                if fmt:
                    cell.number_format = fmt

        # Merge date column
        if row_idx > date_start_row:
            ws.merge_cells(start_row=date_start_row, start_column=1, end_row=row_idx, end_column=1)

        row_idx += 1

    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "C3"  # M004: 제목 2행 + 상영일자·멀티 열 틀고정
    _auto_width(ws, min_row=2)


def _write_competitor_detail_sheet(ws, main_data, competitor_data_dict, movie_title, gen_info):
    """Write 경쟁작 detail sheet - per-screen data for all movies side by side."""
    # E006: 주요작 없으면 경쟁작만 나열
    movie_titles = ([movie_title] if main_data else []) + list(competitor_data_dict.keys())
    # V003(0828): 소계·총계의 극장수를 **작품별**로 표기한다. 예전에는 A열에 기준
    # 작품(첫 작품)의 극장수 하나만 적어서, 기준 작품이 안 걸린 계열사는 0으로 나오고
    # 총계도 비교표와 어긋났다. 비교표와 같은 열 구성으로 맞춘다.
    cols_per_movie = 7  # 극장수, 상영관, 회차, 좌석수, 총좌석수, 판매좌석수, 판매좌석율
    fixed_cols = 2  # 상영일자, 극장명

    all_movie_data = {}
    if main_data:
        all_movie_data[movie_title] = main_data
    for ct, c_data in competitor_data_dict.items():
        all_movie_data[ct] = c_data

    # M004: 생성 시간 행 삭제 — 1행은 공백, 2행 작품명, 3행 서브헤더, 4행부터 데이터
    total_cols = fixed_cols + len(movie_titles) * cols_per_movie

    # Row 2: Movie title headers (M004: A·B열의 '상영일자'/'극장명' 글자 삭제)
    ws.cell(row=2, column=1, value="").fill = BLUE_FILL
    ws.cell(row=2, column=1).border = THIN_BORDER
    ws.cell(row=2, column=2, value="").fill = BLUE_FILL
    ws.cell(row=2, column=2).border = THIN_BORDER

    for mi, mt in enumerate(movie_titles):
        start_col = fixed_cols + 1 + mi * cols_per_movie
        end_col = start_col + cols_per_movie - 1
        ws.merge_cells(start_row=2, start_column=start_col, end_row=2, end_column=end_col)
        cell = ws.cell(row=2, column=start_col, value=mt)
        cell.fill = BLUE_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        for ci in range(start_col, end_col + 1):
            ws.cell(row=2, column=ci).border = THIN_BORDER
            ws.cell(row=2, column=ci).fill = BLUE_FILL

    # Row 3: Sub-headers
    sub_headers = ['극장수', '상영관', '회차', '좌석수', '총좌석수', '판매좌석수', '판매좌석율']
    ws.cell(row=3, column=1, value="상영일자").fill = BLUE_FILL
    ws.cell(row=3, column=1).font = HEADER_FONT
    ws.cell(row=3, column=1).border = THIN_BORDER
    ws.cell(row=3, column=1).alignment = CENTER
    ws.cell(row=3, column=2, value="극장명").fill = BLUE_FILL
    ws.cell(row=3, column=2).font = HEADER_FONT
    ws.cell(row=3, column=2).border = THIN_BORDER
    ws.cell(row=3, column=2).alignment = CENTER

    for mi in range(len(movie_titles)):
        for si, sh in enumerate(sub_headers):
            col = fixed_cols + 1 + mi * cols_per_movie + si
            cell = ws.cell(row=3, column=col, value=sh)
            cell.fill = BLUE_FILL
            cell.font = HEADER_FONT
            cell.border = THIN_BORDER
            cell.alignment = CENTER

    # Collect all unique (date, theater) combinations across ALL movies
    # Build index: (date, theater) -> {movie_title -> [rows]}
    theater_movie_index = {}
    dates = sorted(set(d for m_data in all_movie_data.values() for d in m_data.keys()))

    for mt, m_data in all_movie_data.items():
        for proc_date, rows in m_data.items():
            for r in rows:
                key = (proc_date, r['theater'])
                if key not in theater_movie_index:
                    theater_movie_index[key] = {}
                if mt not in theater_movie_index[key]:
                    theater_movie_index[key][mt] = []
                theater_movie_index[key][mt].append(r)

    # Sort by date, brand priority, theater name
    sorted_keys = sorted(theater_movie_index.keys(), key=lambda k: (
        k[0],  # date
        _brand_priority(theater_movie_index[k].get(movie_title, [{}])[0].get('brand', '') if theater_movie_index[k].get(movie_title) else
                        next(iter(theater_movie_index[k].values()), [{}])[0].get('brand', '')),
        k[1]   # theater name
    ))

    # ----- helpers for brand subtotals / grand total -----
    def _key_brand(k):
        """이 (날짜, 극장) 행이 속한 브랜드. 집계작 우선, 없으면 첫 영화 기준."""
        mv = theater_movie_index[k]
        main_rows = mv.get(movie_title)
        if main_rows:
            return main_rows[0].get('brand', '')
        first = next(iter(mv.values()), None)
        return first[0].get('brand', '') if first else ''

    BRAND_SUBTOTAL_LABEL = {
        'CGV': 'CGV 소계',
        'LOTTE': '롯데 소계',
        'MEGABOX': '메가박스 소계',
        '일반극장': '일반극장 소계',
    }

    # 소계 누적기는 **블록 인덱스**로 키를 잡는다.
    # 같은 작품이 주요작·경쟁작으로 중복 등록되면 movie_titles 에 같은 제목이 두 번
    # 들어오는데, 제목을 키로 쓰면 두 블록이 누적기 하나를 공유해 같은 행을 두 번
    # 더한다 → 소계가 정확히 2배로 부풀었다(비교표와 불일치).
    def _new_agg():
        return {mi: {'screens': 0, 'shows': 0, 'capacity': 0, 'total': 0, 'sold': 0,
                     'theaters': set()}
                for mi in range(len(movie_titles))}

    def _accumulate(agg, movie_rows):
        for mi, mt in enumerate(movie_titles):
            for r in movie_rows.get(mt, []):
                a = agg[mi]
                a['screens'] += 1
                a['shows'] += r['show_count']
                a['capacity'] += r['capacity']
                a['total'] += r['total_seats']
                a['sold'] += r['sold_seats']
                # V003: 비교표(_calc_summary)와 같은 기준 — (브랜드, 극장) 고유 수
                a['theaters'].add((r['brand'], r['theater']))

    def _write_total_row(rr, label, agg, fill, font):
        for ci in range(1, total_cols + 1):
            cell = ws.cell(row=rr, column=ci)
            cell.border = THIN_BORDER
            cell.alignment = CENTER
            cell.fill = fill
            cell.font = font
        ws.cell(row=rr, column=2, value=label)
        for mi, mt in enumerate(movie_titles):
            base_col = fixed_cols + 1 + mi * cols_per_movie
            a = agg[mi]
            if a['screens'] == 0:
                continue
            # 극장수, 상영관, 회차, 좌석수, 총좌석수, 판매좌석수, 판매좌석율
            vals = [len(a['theaters']), a['screens'], a['shows'],
                    a['capacity'], a['total'], a['sold']]
            for si, v in enumerate(vals):
                cell = ws.cell(row=rr, column=base_col + si, value=int(v))
                cell.number_format = NUM_FMT
            rate_cell = ws.cell(row=rr, column=base_col + 6,
                                value=(a['sold'] / a['total']) if a['total'] > 0 else 0)
            rate_cell.number_format = PCT_FMT

    # Write data rows — 브랜드(멀티)별 소계(분홍) + 날짜별 총계(파랑, E003)
    row_idx = 4  # M004: 생성 시간 행 삭제로 한 행 위부터 시작
    brand_agg = None
    date_agg = None
    current_group = None  # (날짜, 브랜드) — 날짜가 바뀌면 같은 브랜드라도 소계를 끊는다

    def _flush_brand(rr):
        g_date, g_brand = current_group
        _write_total_row(rr, BRAND_SUBTOTAL_LABEL.get(g_brand, f"{g_brand} 소계"),
                         brand_agg, PINK_FILL, BOLD_FONT)
        return rr + 1

    def _flush_date(rr):
        # E003: 날짜별 세 멀티 합계 — 기존 최하단 총계와 같은 파란색
        _write_total_row(rr, f"{current_group[0].strftime('%Y-%m-%d')} 총계",
                         date_agg, TOTAL_BLUE_FILL, WHITE_FONT)
        return rr + 1

    for key in sorted_keys:
        (proc_date, theater) = key
        kb = _key_brand(key)
        group = (proc_date, kb)

        if current_group is None:
            brand_agg, date_agg = _new_agg(), _new_agg()
            current_group = group
        elif group != current_group:
            row_idx = _flush_brand(row_idx)
            brand_agg = _new_agg()
            if proc_date != current_group[0]:
                # 날짜가 바뀌면 이전 날짜의 총계를 쓴다 (E003)
                row_idx = _flush_date(row_idx)
                date_agg = _new_agg()
            current_group = group

        movie_rows = theater_movie_index[key]
        _accumulate(brand_agg, movie_rows)
        _accumulate(date_agg, movie_rows)

        # Find max screens for this theater across all movies
        max_screens = max(len(rows) for rows in movie_rows.values())

        for screen_idx in range(max_screens):
            date_str = proc_date.strftime("%Y-%m-%d")
            ws.cell(row=row_idx, column=1, value=date_str)
            ws.cell(row=row_idx, column=1).border = THIN_BORDER
            ws.cell(row=row_idx, column=1).alignment = CENTER
            ws.cell(row=row_idx, column=1).font = DATA_FONT

            ws.cell(row=row_idx, column=2, value=theater)
            ws.cell(row=row_idx, column=2).border = THIN_BORDER
            ws.cell(row=row_idx, column=2).alignment = CENTER
            ws.cell(row=row_idx, column=2).font = DATA_FONT

            for mi, mt in enumerate(movie_titles):
                base_col = fixed_cols + 1 + mi * cols_per_movie
                m_rows = movie_rows.get(mt, [])

                if screen_idx < len(m_rows):
                    r = m_rows[screen_idx]
                    sold_rate = (r['sold_seats'] / r['total_seats']) if r['total_seats'] > 0 else 0
                    # V003: 극장수 칸은 소계·총계 전용이라 데이터 행에서는 비워 둔다
                    vals = [
                        '',
                        r['screen'],
                        r['show_count'],
                        r['capacity'],
                        r['total_seats'],
                        r['sold_seats'],
                        sold_rate
                    ]
                else:
                    vals = ['', '', '', '', '', '', '']

                for si, v in enumerate(vals):
                    cell = ws.cell(row=row_idx, column=base_col + si, value=v)
                    cell.border = THIN_BORDER
                    cell.alignment = CENTER
                    cell.font = DATA_FONT
                    # 합계줄과 동일한 숫자/백분율 서식 (si 0=극장수 공란, 1=상영관명은 문자)
                    if si in (2, 3, 4, 5) and isinstance(v, (int, float)):
                        cell.number_format = NUM_FMT
                    elif si == 6 and isinstance(v, (int, float)):
                        cell.number_format = PCT_FMT

            row_idx += 1

    # flush 마지막 (날짜, 브랜드) 소계 + 마지막 날짜 총계
    # (E003: 시트 전체 총계 대신 날짜별 총계로 마무리한다)
    if current_group is not None:
        row_idx = _flush_brand(row_idx)
        row_idx = _flush_date(row_idx)

    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "C4"  # M004: 제목·헤더 3행 + 상영일자·극장명 열 틀고정
    _auto_width(ws, min_row=3)


def export_transformed_schedules(queryset, movie_title=None, start_date=None, end_date=None, competitor_querysets=None):
    """
    Exports MovieSchedule QuerySet to Excel matching the standard schedule format.

    Sheets (E001 순서):
    1. 상영시간표_YYYY-MM-DD (per date schedule) — 주요작이 있을 때만 (E006)
    2. 경쟁작 (competitor detail)
    3. 비교표 (집계작 및 경쟁작 멀티별 비교)
    """
    # E006: 주요작 없이 경쟁작만으로도 다운로드 가능
    has_main = queryset is not None and queryset.exists()
    if not has_main and not competitor_querysets:
        return None

    # ========== Region Mapping ==========
    region_map = _build_region_map()
    normal_index = _build_normal_theater_index()  # 일반극장 지역·좌석수 보강용

    # ========== Collect Main Data Per Date ==========
    # 날짜 목록: 주요작 기준, 주요작이 없으면(E006) 경쟁작 전체 날짜
    if has_main:
        available_dates = list(
            queryset.filter(play_date__isnull=False)
            .values_list('play_date', flat=True).distinct().order_by('play_date')
        )
    else:
        date_set = set()
        for comp_qs in (competitor_querysets or {}).values():
            date_set.update(
                comp_qs.filter(play_date__isnull=False)
                .values_list('play_date', flat=True).distinct()
            )
        available_dates = sorted(date_set)

    if not available_dates:
        return None

    all_data = {}
    global_max_shows = 0

    if has_main:
        for d in available_dates:
            sub_qs = queryset.filter(play_date=d)
            rows, max_shows = _process_to_rows(sub_qs, region_map, normal_index)
            if rows:
                all_data[d] = rows
                global_max_shows = max(global_max_shows, max_shows)

        if not all_data:
            return None

    display_max_shows = max(12, global_max_shows)

    # ========== Collect Competitor Data ==========
    competitor_all_data = {}  # {comp_title: {date: rows}}
    if competitor_querysets:
        for comp_title, comp_qs in competitor_querysets.items():
            comp_data = {}
            for d in available_dates:
                sub_qs = comp_qs.filter(play_date=d)
                rows, _ = _process_to_rows(sub_qs, region_map, normal_index)
                if rows:
                    comp_data[d] = rows
            if comp_data:
                competitor_all_data[comp_title] = comp_data

        # 경쟁작은 '크롤 대상 영화' 등록 순이 아니라 총좌석수가 많은 순으로 나열한다.
        # (집계작은 비교 기준이므로 항상 맨 앞 고정)
        def _total_seats(comp_data):
            return sum(r['total_seats'] for rows in comp_data.values() for r in rows)

        competitor_all_data = dict(
            sorted(competitor_all_data.items(), key=lambda kv: -_total_seats(kv[1]))
        )

    # ========== File Setup ==========
    save_dir = os.path.join(settings.BASE_DIR, 'media', 'crawler_exports')
    os.makedirs(save_dir, exist_ok=True)

    now = datetime.now()
    gen_date = now.strftime("%Y-%m-%d")
    gen_time = now.strftime("%H_%M_%S")
    weekday_kor = ["월", "화", "수", "목", "금", "토", "일"]
    day_of_week = weekday_kor[now.weekday()]
    gen_info = f"({gen_date} [{gen_time}])({day_of_week})"

    if not movie_title:
        # E006: 주요작 없이 경쟁작만 내보내는 경우의 파일명 표기
        movie_title = "경쟁작" if not has_main else "Overall"
    safe_title = re.sub(r'[\\/*?:"<>|]', "", movie_title).replace(" ", "")

    if not start_date:
        start_date = gen_date
    if not end_date:
        end_date = start_date

    def fmt_date(d):
        if isinstance(d, datetime):
            return d.strftime("%Y-%m-%d")
        return str(d).split(' ')[0]

    # M003: 파일명 규칙 변경
    #  - 주요작 선택: '시간표_주요작제목.xlsx' (예: 시간표_비광)
    #  - 주요작 없음: '경쟁작 좌석수_MM.DD~MM.DD.xlsx' (예: 경쟁작 좌석수_08.20~08.26)
    def _mmdd(x):
        return f"{x[5:7]}.{x[8:10]}" if len(x) >= 10 else x

    if has_main:
        filename = f"시간표_{safe_title}.xlsx"
    else:
        filename = (f"경쟁작 좌석수_{_mmdd(fmt_date(start_date))}"
                    f"~{_mmdd(fmt_date(end_date))}.xlsx")
    file_path = os.path.join(save_dir, filename)

    # ========== Write Excel ==========
    # E001: 계열사별·지역별·포맷별 요약표 시트는 제거.
    # 시트 순서는 상영시간표 → 경쟁작 → 비교표.
    wb = Workbook()
    wb.remove(wb.active)

    # 1. Schedule Sheets (상영시간표) — 주요작이 있을 때만 (E006)
    for proc_date, rows in all_data.items():
        # M004: 시트명은 시간표의 해당 날짜만 (예: '0902')
        sheet_name = proc_date.strftime('%m%d')
        if sheet_name in wb.sheetnames:  # 연도가 다른 같은 월일 충돌 방지
            sheet_name = proc_date.strftime('%y%m%d')
        ws = wb.create_sheet(sheet_name)
        _write_schedule_sheet(ws, rows, proc_date, movie_title, display_max_shows, gen_info)

    # 2. 경쟁작
    if competitor_all_data:
        ws = wb.create_sheet("경쟁작")
        _write_competitor_detail_sheet(ws, all_data, competitor_all_data, movie_title, gen_info)

    # 3. 비교표 (구 '집계작 및 경쟁작 멀티별 비교' — E001에서 시트명 변경)
    if competitor_all_data:
        # C001: 날짜×계열사에 크롤 데이터가 아예 없는 경우(수집 시차)를
        # '상영 없음'과 구분해 표기하기 위한 존재 여부 맵 (영화 필터와 무관하게 전체 조회)
        presence_dates = sorted(
            set(all_data.keys())
            | {d for c_data in competitor_all_data.values() for d in c_data.keys()}
        )
        brand_presence = {
            (d, b): MovieSchedule.objects.filter(play_date=d, brand=b).exists()
            for d in presence_dates
            for b in BRAND_ORDER
        }
        ws = wb.create_sheet("비교표")
        _write_comparison_sheet(
            ws, all_data, competitor_all_data, movie_title, gen_info,
            brand_presence=brand_presence,
        )

    if not wb.sheetnames:
        return None

    wb.save(file_path)
    logger.info(f"Schedule Excel exported: {file_path}")
    return file_path


# =============================================================================
# 특수상영(무대인사 등) 키워드 기반 평면 엑셀 내보내기
# =============================================================================

_SPECIAL_FORMAT_KEYWORDS = [
    "2D", "3D", "4D", "디지털", "DIGITAL", "자막", "더빙", "IMAX", "4DX", "SCREENX",
    "SCREEN-X", "DOLBY", "ATMOS", "LASER", "SPHEREX", "리클라이너",
    # E004 canonical 포맷 태그 — 구분1 컬럼에서 제외 (포맷1 컬럼에 표기됨)
    "SUPER-4D", "MX4D", "아이맥스", "돌비", "애트모스", "수퍼4D", "슈퍼4D",
]


def _special_category_tags(tags):
    """포맷(2D/자막/IMAX 등)을 제외한 구분 태그(무대인사·GV 등)만 반환."""
    out = []
    for t in (tags or []):
        tu = str(t).upper()
        if any(f in tu for f in [k.upper() for k in _SPECIAL_FORMAT_KEYWORDS]):
            continue
        out.append(str(t))
    return out


def _brand_to_multi(brand):
    """MovieSchedule.brand -> 엑셀 '멀티' 표기 (예시 파일 기준: MEGABOX->MEGA)."""
    if brand == "MEGABOX":
        return "MEGA"
    return brand or ""


def export_special_screenings(queryset, keyword=None, start_date=None, end_date=None):
    """
    특수상영(무대인사 등) 평면 엑셀 내보내기.
    컬럼: 영화명·날짜·지역·멀티·극장명·관·포맷1·구분1·상영시간·총회차·총스크린·총좌석수·판매좌석수·잔여좌석수·좌석판매율
    """
    save_dir = os.path.join(settings.BASE_DIR, 'media', 'crawler_exports')
    os.makedirs(save_dir, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    kw = re.sub(r'[^0-9A-Za-z가-힣]', '', str(keyword or "특수상영"))
    filename = f"특수상영_{kw}_{timestamp}.xlsx"
    file_path = os.path.join(save_dir, filename)

    region_map = _build_region_map()

    headers = [
        "영화명", "날짜", "지역", "멀티", "극장명", "관", "포맷1", "구분1",
        "상영시간", "총회차", "총스크린", "총좌석수", "판매좌석수", "잔여좌석수", "좌석판매율",
    ]

    rows = list(queryset)

    # 멀티(브랜드) 순서: CGV → MEGA → LOTTE → 일반극장 (예시 파일 기준)
    BRAND_ORD = {'CGV': 0, 'MEGABOX': 1, 'LOTTE': 2, '일반극장': 3}
    BRAND_LABEL = {'CGV': 'CGV', 'MEGABOX': 'MEGA', 'LOTTE': 'LOTTE', '일반극장': '일반극장'}

    def sort_key(s):
        return (
            BRAND_ORD.get(s.brand, 9),
            str(s.play_date or (s.start_time.date() if s.start_time else "")),
            s.theater_name or "",
            s.screen_name or "",
            s.start_time or dj_timezone.now(),
        )
    rows.sort(key=sort_key)

    wb = Workbook()
    ws = wb.active
    ws.title = "특수상영"

    # 예시 파일 색상: 헤더 회색 / 소계 분홍 / 총계 파랑+흰글씨
    HEAD_FILL = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    SUB_FILL = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
    TOTAL_FILL = PatternFill(start_color="0033CC", end_color="0033CC", fill_type="solid")
    HEAD_FONT = Font(bold=True, size=10)
    SUB_FONT = Font(bold=True, size=10)
    TOTAL_FONT = Font(bold=True, size=10, color="FFFFFF")
    DATA_FONT_ = Font(size=10)
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")
    NUMFMT = '#,##0'

    def write_cells(rr, values, fill=None, font=None):
        for c, v in enumerate(values, 1):
            cell = ws.cell(row=rr, column=c, value=v)
            cell.border = border
            cell.alignment = center
            if fill:
                cell.fill = fill
            cell.font = font or DATA_FONT_
            if c in (10, 11, 12, 13, 14) and isinstance(v, (int, float)):
                cell.number_format = NUMFMT

    write_cells(1, headers, fill=HEAD_FILL, font=HEAD_FONT)

    r = 2
    grand = [0, 0, 0, 0, 0]  # 총회차, 총스크린, 총좌석수, 판매좌석수, 잔여좌석수
    i, n = 0, len(rows)
    while i < n:
        brand = rows[i].brand
        agg = [0, 0, 0, 0, 0]
        # 같은 브랜드 데이터 행 작성
        while i < n and rows[i].brand == brand:
            s = rows[i]
            total = int(s.total_seats or 0)
            remain = int(s.remaining_seats or 0)
            sold = max(0, total - remain)
            rate = f"{(sold / total * 100):.1f}%" if total > 0 else ""
            fmt, sub_type = _extract_format_and_type(s.tags)
            format1 = f"{fmt}({sub_type})" if sub_type and sub_type != "일반" else fmt
            category1 = " ".join(_special_category_tags(s.tags))
            play_date = s.play_date or (s.start_time.date() if s.start_time else None)
            region = _resolve_region(brand, s.theater_name, region_map)
            write_cells(r, [
                s.movie_title,
                str(play_date) if play_date else "",
                region,
                _brand_to_multi(brand),
                s.theater_name,
                s.screen_name,
                format1,
                category1,
                dj_timezone.localtime(s.start_time).strftime("%H:%M") if s.start_time else "",
                1, 1, total, sold, remain, rate,
            ])
            r += 1
            agg[0] += 1; agg[1] += 1; agg[2] += total; agg[3] += sold; agg[4] += remain
            i += 1
        # 멀티별 소계
        s_rate = f"{round(agg[3] / agg[2] * 100)}%" if agg[2] else ""
        write_cells(r, ["", "", "", "", f"{BRAND_LABEL.get(brand, brand)} 소계", "", "", "", "",
                        agg[0], agg[1], agg[2], agg[3], agg[4], s_rate], fill=SUB_FILL, font=SUB_FONT)
        r += 1
        for k in range(5):
            grand[k] += agg[k]

    # 총 계
    g_rate = f"{round(grand[3] / grand[2] * 100)}%" if grand[2] else ""
    write_cells(r, ["", "", "", "", "총 계", "", "", "", "",
                    grand[0], grand[1], grand[2], grand[3], grand[4], g_rate], fill=TOTAL_FILL, font=TOTAL_FONT)

    _auto_width(ws, min_row=1)
    ws.freeze_panes = "A2"
    wb.save(file_path)
    logger.info(f"특수상영 Excel exported: {file_path} ({len(rows)} rows)")
    return file_path
