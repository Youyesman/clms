"""메가박스 윙업(M SCORE) 관람객현황(배급사) 크롤러.

수동 흐름(검색 → 영화 체크 → 조회 → 엑셀다운로드)을 서버 사이드로 재현한다.
  1) login.do            : JSON 로그인 (세션 쿠키 획득)
  2) searchMovie.do      : 기간 내 영화 목록 조회
  3) (포함/제외 키워드로 영화 필터)
  4) selectShb2DataList.do: 선택 영화들의 관람객현황 데이터(JSON)
  5) 메일 첨부와 동일한 메가박스 스코어 업로드 양식 xlsx 로 변환

비밀번호 변경 모달은 순수 프론트 JS(비번 변경일 기준)라 무시한다.
"""

import html as _html
import io
import json
from datetime import date

import requests
from django.conf import settings
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

BASE = "https://wingup.megabox.co.kr"
_UA = "Mozilla/5.0 (Windows NT 10.0; Win64; x64)"

# 고정 앞 컬럼(메가박스 실제 다운로드 양식: A열은 빈 여백)
_FIXED_HEADERS = ["지점", "상영일", "관", "상영영화", "상영종류", "티켓가"]
_FONT_NAME = "맑은 고딕"
_HEADER_FILL = "FFDBDBDB"  # 헤더 회색


def _clean(v):
    """HTML 엔티티 디코딩(예: '2D&#40;자막&#41;' → '2D(자막)')."""
    return _html.unescape(str(v)) if v is not None else ""


def _login(session, user=None, password=None):
    if user is None:
        cfg = settings.MEGABOX_WINGUP
        user, password = cfg["USER"], cfg["PASSWORD"]
    session.get(f"{BASE}/loginView.do", timeout=20)
    r = session.post(
        f"{BASE}/login.do",
        data=json.dumps({"userNo": user, "passwd": password}),
        headers={
            "Content-Type": "application/json;charset=UTF-8",
            "X-Requested-With": "XMLHttpRequest",
            "Referer": f"{BASE}/loginView.do",
        },
        timeout=20,
    )
    try:
        j = r.json()
    except Exception:
        raise RuntimeError("로그인 응답 파싱 실패(계정/잠금 확인)")
    if str(j.get("statCd")) != "0":
        raise RuntimeError(f"로그인 실패: {j.get('msg')}")
    info = j.get("dlUserInfo", {})
    if info.get("succFailAt") != "Y":
        raise RuntimeError("로그인 실패: 아이디/비밀번호 확인 필요")
    # 관람객현황 페이지 진입(세션 활성화) — 비번 변경 모달은 무시
    session.get(f"{BASE}/wp/shb/BoxoBrchTermPrcoL.do", timeout=20)


def _ajax_headers():
    return {
        "Content-Type": "application/json;charset=UTF-8",
        "X-Requested-With": "XMLHttpRequest",
        "Referer": f"{BASE}/wp/shb/BoxoBrchTermPrcoL.do",
    }


def search_movies(session, start_de, end_de, movie_nm=""):
    """기간 내 영화 목록. [{movieNo, kofMovieCd, movieNm, movieKindCd, ...}]"""
    r = session.post(
        f"{BASE}/wp/sha/searchMovie.do",
        data=json.dumps({"playStartDe": start_de, "playEndDe": end_de, "movieNm": movie_nm}),
        headers=_ajax_headers(),
        timeout=20,
    )
    return r.json().get("movieList", [])


def filter_movies(movies, include="", excludes=None):
    """영화명에 include 키워드 포함 + excludes 키워드 모두 미포함인 영화만."""
    excludes = [e for e in (excludes or []) if e]
    inc = (include or "").strip()
    picked = []
    for m in movies:
        nm = _clean(m.get("movieNm"))
        if inc and inc not in nm:
            continue
        if any(ex in nm for ex in excludes):
            continue
        picked.append(m)
    return picked


def fetch_score_rows(session, start_de, end_de, movie_nos, brch_no="", brch_div_cd=""):
    """선택 영화들의 관람객현황 데이터 행(JSON) 반환."""
    payload = {
        "brchNo": brch_no,
        "movieNm": "",
        "playStartDe": start_de,
        "playEndDe": end_de,
        "movieNo": ",".join(str(n) for n in movie_nos),
        "brchDivCd": brch_div_cd,
    }
    r = session.post(
        f"{BASE}/wp/shb/selectShb2DataList.do",
        data=json.dumps(payload),
        headers=_ajax_headers(),
        timeout=60,
    )
    return r.json().get("dataList", [])


def rows_to_xlsx(rows, start_de="", end_de=""):
    """관람객현황 JSON → 메가박스 실제 다운로드 양식과 동일한 xlsx(bytes).

    레이아웃(A열은 여백):
      row1 빈줄 / row2~4 병합 제목 / row6 조회기간·출력일시 /
      row7~10 병합 헤더(회색) / row11~ 데이터 / 마지막 합계행.
    회차 컬럼은 1회~N회(N=최대 상영회차) 동적. (파서 skiprows=6 과 호환)
    """
    # 최대 회차 N 계산 (데이터에 값이 있는 최고 회차)
    n_show = 1
    for row in rows:
        mx = row.get("maxPlaySeq") or 0
        for k in range(1, 16):
            if (row.get(f"play{k}Seq") or 0) and k > mx:
                mx = k
        n_show = max(n_show, mx)

    headers = _FIXED_HEADERS + [f"{i}회" for i in range(1, n_show + 1)] + ["합계", "매출액"]
    n_cols = len(headers)              # B열부터의 컬럼 수
    first_col = 2                       # B (A열은 여백)
    last_col = first_col + n_cols - 1
    L = get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    thin = Side(style="thin", color="FF000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    f_title = Font(name=_FONT_NAME, size=20, bold=True)
    f_period = Font(name=_FONT_NAME, size=11, bold=True)
    f_header = Font(name=_FONT_NAME, size=11)
    f_data = Font(name=_FONT_NAME, size=10)
    fill_header = PatternFill("solid", fgColor=_HEADER_FILL)
    al_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    al_left = Alignment(vertical="center")
    al_right = Alignment(horizontal="right", vertical="center")

    # 열 너비 (실제 양식과 동일)
    ws.column_dimensions["A"].width = 2.625
    for col, w in (("B", 15), ("C", 15), ("D", 15), ("E", 37.5)):
        ws.column_dimensions[col].width = w

    # row2~4 병합 제목
    ws.merge_cells(f"B2:{L(last_col)}4")
    for r in range(2, 5):
        for c in range(first_col, last_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = f_title
            cell.alignment = al_center
            cell.border = border
    ws.cell(row=2, column=first_col, value="관람객현황(배급사)")

    # row6 조회기간 / 출력일시
    period = f"조회기간 : {start_de} ~ {end_de}" if start_de else "조회기간 :"
    ws.cell(row=6, column=first_col, value=period).font = f_period
    ws.cell(row=6, column=first_col + 3, value=f"출력일시 : {date.today().isoformat()}").font = f_period

    # row7~10 병합 헤더 (회색)
    for j, h in enumerate(headers):
        c = first_col + j
        ws.merge_cells(start_row=7, start_column=c, end_row=10, end_column=c)
        for r in range(7, 11):
            cell = ws.cell(row=r, column=c)
            cell.font = f_header
            cell.fill = fill_header
            cell.border = border
            cell.alignment = al_center
        ws.cell(row=7, column=c, value=h)

    # row11~ 데이터
    show_keys = [f"play{i}Seq" for i in range(1, n_show + 1)]
    sums = [0] * n_show
    sum_tot = 0
    sum_amt = 0
    r = 11
    for row in rows:
        plays = [int(row.get(k) or 0) for k in show_keys]
        tot = int(row.get("totSum") or sum(plays))
        amt = int(row.get("sellTotAmt") or 0)
        for i, p in enumerate(plays):
            sums[i] += p
        sum_tot += tot
        sum_amt += amt
        values = [
            _clean(row.get("brchNm")),
            _clean(row.get("playDe")),
            _clean(row.get("theabNm")),
            _clean(row.get("movieNm")),
            _clean(row.get("movieKindCd")),
            int(row.get("ticketAmt") or 0),
        ] + plays + [tot, amt]
        for j, v in enumerate(values):
            cell = ws.cell(row=r, column=first_col + j, value=v)
            cell.font = f_data
            cell.border = border
            cell.alignment = al_right if isinstance(v, int) else al_left
        r += 1

    # 마지막 합계행
    total_row = ["", "", "", "", "", "합계"] + sums + [sum_tot, sum_amt]
    for j, v in enumerate(total_row):
        cell = ws.cell(row=r, column=first_col + j, value=(v if v != "" else None))
        cell.font = f_header
        cell.border = border
        cell.alignment = al_right if isinstance(v, int) else al_center

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def crawl_megabox_score(start_de, end_de, include="", excludes=None):
    """전체 흐름 실행 → (filename, xlsx_bytes, meta dict)."""
    start_de = start_de.replace("-", "")
    end_de = end_de.replace("-", "")
    session = requests.Session()
    session.headers["User-Agent"] = _UA

    _login(session)
    movies = search_movies(session, start_de, end_de, include or "")
    picked = filter_movies(movies, include=include, excludes=excludes)
    if not picked:
        raise RuntimeError("조건에 맞는 영화가 없습니다. (키워드/기간 확인)")
    movie_nos = [m.get("movieNo") for m in picked]
    rows = fetch_score_rows(session, start_de, end_de, movie_nos)
    if not rows:
        raise RuntimeError("관람객현황 데이터가 없습니다.")

    xlsx = rows_to_xlsx(rows, start_de, end_de)
    filename = f"메가박스_관람객현황_{start_de}_{end_de}.xlsx"
    meta = {
        "movies": [
            {"movieNo": m.get("movieNo"), "movieNm": _clean(m.get("movieNm"))}
            for m in picked
        ],
        "row_count": len(rows),
    }
    return filename, xlsx, meta


# ── 멀티 배급사(전 계정) 크롤 ──
from concurrent.futures import ThreadPoolExecutor, as_completed  # noqa: E402

from .megabox_accounts import MEGABOX_ACCOUNTS  # noqa: E402


def get_accounts():
    """크롤에 사용할 배급사 계정 목록.

    DB(MegaboxDistributorAccount)의 활성 계정을 우선 사용하고,
    아직 등록된 계정이 없으면 하드코딩 기본 목록으로 폴백한다.
    """
    try:
        from .models import MegaboxDistributorAccount
        rows = list(MegaboxDistributorAccount.objects.filter(is_active=True))
        if rows:
            return [{"name": r.name, "user": r.user, "password": r.password} for r in rows]
    except Exception:
        pass
    return MEGABOX_ACCOUNTS


def crawl_one_account(name, user, password, start_de, end_de, includes, excludes):
    """한 배급사 계정 크롤 → {name, ok, error, movies, row_count, rows}."""
    res = {"name": name, "ok": False, "error": "", "movies": [], "row_count": 0, "rows": []}
    try:
        s = requests.Session()
        s.headers["User-Agent"] = _UA
        _login(s, user, password)

        # 포함 키워드별로 검색(서버측 movieNm 부분일치) → movieNo 로 중복 제거
        found = {}
        for kw in (includes or [""]):
            for m in search_movies(s, start_de, end_de, kw):
                found[m.get("movieNo")] = m
        picked = filter_movies(list(found.values()), include="", excludes=excludes)
        res["movies"] = [
            {"movieNo": m.get("movieNo"), "movieNm": _clean(m.get("movieNm"))}
            for m in picked
        ]
        if picked:
            rows = fetch_score_rows(s, start_de, end_de, [m.get("movieNo") for m in picked])
            res["rows"] = rows
            res["row_count"] = len(rows)
        res["ok"] = True
    except Exception as e:
        res["error"] = str(e)[:200]
    return res


def _safe_name(s):
    """파일명에 못 쓰는 문자 제거."""
    import re
    return re.sub(r'[\\/:*?"<>|]', "_", s).strip()


def crawl_all_accounts(start_de, end_de, includes, excludes, max_workers=10):
    """모든 배급사 계정 크롤 → 배급사별 개별 엑셀.

    반환: summary 리스트. 각 항목:
      {name, ok, error, row_count, movies, filename, xlsx(bytes 또는 None)}
    """
    start_de = start_de.replace("-", "")
    end_de = end_de.replace("-", "")
    includes = [i for i in (includes or []) if i]
    excludes = [e for e in (excludes or []) if e]

    accounts = get_accounts()
    order = {a["name"]: i for i, a in enumerate(accounts)}
    results = []
    with ThreadPoolExecutor(max_workers=max_workers) as ex:
        futs = [
            ex.submit(crawl_one_account, a["name"], a["user"], a["password"],
                      start_de, end_de, includes, excludes)
            for a in accounts
        ]
        for f in as_completed(futs):
            results.append(f.result())
    results.sort(key=lambda r: order.get(r["name"], 999))

    summary = []
    for r in results:
        item = {
            "name": r["name"],
            "ok": r["ok"],
            "error": r["error"],
            "row_count": r["row_count"],
            "movies": r["movies"],
            "filename": None,
            "xlsx": None,
        }
        rows = r.get("rows", [])
        if rows:
            rows.sort(key=lambda x: (
                _clean(x.get("brchNm")), _clean(x.get("movieNm")),
                _clean(x.get("theabNm")), x.get("ticketAmt") or 0,
            ))
            item["xlsx"] = rows_to_xlsx(rows, start_de, end_de)
            item["filename"] = f"메가박스_{_safe_name(r['name'])}_{start_de}_{end_de}.xlsx"
        summary.append(item)
    return summary
