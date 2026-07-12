"""씨네큐 스코어(score.cineq.co.kr) 관객현황 크롤러.

수동 흐름(로그인 → 관객현황(Admin002) 극장별 조회 → 엑셀저장)을 서버 사이드로 재현한다.
  1) POST /Login/LoginProcess      : 폼 로그인 (성공 시 302 → /Admin002, LoginData 쿠키)
  2) GET  /Admin002                : 극장 목록(select[name=TheaterCode]) 파싱
  3) GET  /Admin002/Index          : 극장/상영일별 관객현황 테이블(HTML) 조회
     (TheaterCode, MovieCode=all, playDate=YYYYMMDD — 상영일 당일 조회 불가, 전일부터)
  4) 사이트 '엑셀저장'과 동일한 씨네큐 스코어 업로드 양식 xlsx 로 변환
     (score.score_parsers.preview_cineq_format 이 그대로 파싱 가능한 레이아웃)

테이블 구조: 영화관/영화명/상영일/상영관 rowspan 그룹 + 회차 상영시간 행
+ 가격대별 인원 행(class="seat") + 상영관 소계(total-screen)/극장 합계(total-theater) 행.
"""

import html as _html
import io
import re
from datetime import date, datetime, timedelta

import requests
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

BASE = "https://score.cineq.co.kr"
_UA = "Mozilla/5.0 (Windows NT 10.0; Win64; x64)"
N_SHOW = 13  # 씨네큐 관객현황 회차 컬럼 수 (1회~13회 고정)

_FONT_NAME = "맑은 고딕"
_HEADER_FILL = "FFDBDBDB"

HEADERS = ["영화관", "영화명", "상영일", "상영관", "가격(원)"] \
    + [f"{i}회" for i in range(1, N_SHOW + 1)] + ["합계"]


def _text(cell_html):
    """셀 HTML → 태그 제거 + 엔티티 디코딩된 텍스트."""
    return _html.unescape(re.sub(r"<[^>]+>", "", cell_html)).strip()


def _login(session, user, password):
    session.get(f"{BASE}/Login", timeout=20)
    r = session.post(
        f"{BASE}/Login/LoginProcess",
        data={"userid": user, "userpass": password},
        timeout=20,
        allow_redirects=False,
    )
    # 성공: 302 → /Admin002 (+ LoginData 쿠키).
    # 실패: 200 + alert('사유') 스크립트 (예: 정보 불일치, 휴면 회원).
    if r.status_code != 302:
        m = re.search(r"alert\('([^']+)'\)", r.text)
        reason = m.group(1) if m else "아이디/비밀번호 확인 필요"
        raise RuntimeError(f"로그인 실패: {reason}")


def fetch_theaters(session):
    """관객현황 페이지의 극장 select 파싱 → [(code, name)]."""
    r = session.get(f"{BASE}/Admin002", timeout=20)
    r.encoding = "utf-8"
    m = re.search(r'<select name="TheaterCode">(.*?)</select>', r.text, re.S)
    if not m:
        raise RuntimeError("극장 목록 파싱 실패(세션 만료/페이지 변경 확인)")
    return re.findall(r'<option value="(\d+)"[^>]*>\s*([^<]+?)\s*</option>', m.group(1))


def _parse_score_table(page_html):
    """Admin002 조회 결과 HTML → 가격대별 인원 행 리스트.

    반환 행: {theater, movie, play_date, screen, price(int), plays[13], total(int)}
    소계(total-screen)·합계(total-theater)·회차시간 행은 제외한다.
    """
    m = re.search(r'<table class="table-list[^"]*"[^>]*>(.*?)</table>', page_html, re.S)
    if not m:
        return []

    rows = []
    ctx = {"theater": "", "movie": "", "play_date": "", "screen": ""}
    for tr_attrs, tr_html in re.findall(r"<tr([^>]*)>(.*?)</tr>", m.group(1), re.S):
        if "head" in tr_attrs:
            continue
        cells = re.findall(r"<td([^>]*)>(.*?)</td>", tr_html, re.S)
        if not cells:
            continue

        # rowspan 그룹 셀(영화관/영화명/상영일/상영관)로 현재 문맥 갱신
        for attrs, content in cells:
            for cls, key in (("theater", "theater"), ("movie", "movie"),
                             ("date", "play_date"), ("screen", "screen")):
                if f'class="{cls}"' in attrs:
                    val = _text(content)
                    # 순수 숫자 상영관은 DB 관 명칭("N관")으로 정규화.
                    # "(리클라이너)1" 등 접두 형식은 스코어 파서의 씨네큐 룰이 처리한다.
                    if key == "screen" and re.fullmatch(r"\d+", val):
                        val = f"{val}관"
                    ctx[key] = val

        if "total" in tr_attrs:          # 상영관 소계 / 극장 합계 행
            continue
        if "seat" not in tr_attrs:       # 회차 상영시간 행(가격 없음)
            continue

        # seat 행: [가격, 1회..13회 인원, 합계]
        plain = [_text(c) for a, c in cells if 'class="price"' in a or not a.strip()]
        if len(plain) != N_SHOW + 2:
            continue
        try:
            price = int(plain[0].replace(",", "") or 0)
            plays = [int(v.replace(",", "") or 0) for v in plain[1:1 + N_SHOW]]
            total = int(plain[-1].replace(",", "") or 0)
        except ValueError:
            continue
        rows.append({**ctx, "price": price, "plays": plays, "total": total})
    return rows


def fetch_score_rows(session, theater_code, play_date):
    """극장/상영일 하나의 관객현황 조회 → 파싱된 행 리스트."""
    r = session.get(
        f"{BASE}/Admin002/Index",
        params={"TheaterCode": theater_code, "MovieCode": "all", "playDate": play_date},
        timeout=30,
    )
    r.encoding = "utf-8"
    return _parse_score_table(r.text)


def filter_rows(rows, includes=None, excludes=None):
    """영화명 키워드 필터. includes 비면 전체, excludes 는 모두 미포함."""
    includes = [i for i in (includes or []) if i]
    excludes = [e for e in (excludes or []) if e]
    out = []
    for row in rows:
        nm = row["movie"]
        if includes and not any(kw in nm for kw in includes):
            continue
        if any(ex in nm for ex in excludes):
            continue
        out.append(row)
    return out


def rows_to_xlsx(rows):
    """파싱 행 → 씨네큐 스코어 업로드 양식 xlsx(bytes).

    사이트 '엑셀저장' 표와 동일한 평면 헤더(영화관/영화명/상영일/상영관/가격(원)/1회~13회/합계).
    score_parsers.preview_cineq_format 이 forward-fill 로 읽으므로 모든 행에 그룹값을 채운다.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    thin = Side(style="thin", color="FF000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    f_header = Font(name=_FONT_NAME, size=11, bold=True)
    f_data = Font(name=_FONT_NAME, size=10)
    fill_header = PatternFill("solid", fgColor=_HEADER_FILL)
    al_center = Alignment(horizontal="center", vertical="center")
    al_left = Alignment(vertical="center")
    al_right = Alignment(horizontal="right", vertical="center")

    for col, w in (("A", 14), ("B", 30), ("C", 11), ("D", 16), ("E", 10)):
        ws.column_dimensions[col].width = w

    for j, h in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=j, value=h)
        cell.font = f_header
        cell.fill = fill_header
        cell.border = border
        cell.alignment = al_center

    r = 2
    for row in rows:
        values = [row["theater"], row["movie"], str(row["play_date"]),
                  row["screen"], row["price"]] + row["plays"] + [row["total"]]
        for j, v in enumerate(values, start=1):
            cell = ws.cell(row=r, column=j, value=v)
            cell.font = f_data
            cell.border = border
            cell.alignment = al_right if isinstance(v, int) else al_left
        r += 1

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def _date_range(start_de, end_de, max_days=31):
    """YYYYMMDD 범위 → 날짜 문자열 리스트 (최대 max_days)."""
    d0 = datetime.strptime(start_de, "%Y%m%d").date()
    d1 = datetime.strptime(end_de, "%Y%m%d").date()
    if d1 < d0:
        d0, d1 = d1, d0
    days = (d1 - d0).days + 1
    if days > max_days:
        raise RuntimeError(f"조회 기간이 너무 깁니다(최대 {max_days}일).")
    return [(d0 + timedelta(days=i)).strftime("%Y%m%d") for i in range(days)]


# ── 멀티 배급사(전 계정) 크롤 ──
from concurrent.futures import ThreadPoolExecutor, as_completed  # noqa: E402

from .cineq_accounts import CINEQ_ACCOUNTS  # noqa: E402


def get_accounts():
    """크롤에 사용할 배급사 계정 목록.

    DB(CineQDistributorAccount)의 활성 계정을 우선 사용하고,
    아직 등록된 계정이 없으면 하드코딩 기본 목록으로 폴백한다.
    """
    try:
        from .models import CineQDistributorAccount
        rows = list(CineQDistributorAccount.objects.filter(is_active=True))
        if rows:
            return [{"name": r.name, "user": r.user, "password": r.password} for r in rows]
    except Exception:
        pass
    return CINEQ_ACCOUNTS


def crawl_one_account(name, user, password, dates, includes, excludes):
    """한 배급사 계정 크롤 → {name, ok, error, movies, row_count, rows}."""
    res = {"name": name, "ok": False, "error": "", "movies": [], "row_count": 0, "rows": []}
    try:
        s = requests.Session()
        s.headers["User-Agent"] = _UA
        _login(s, user, password)

        theaters = fetch_theaters(s)
        rows = []
        for play_de in dates:
            for code, _tname in theaters:
                rows.extend(fetch_score_rows(s, code, play_de))
        rows = filter_rows(rows, includes, excludes)
        res["rows"] = rows
        res["row_count"] = len(rows)
        res["movies"] = [{"movieNm": nm} for nm in
                         sorted({row["movie"] for row in rows})]
        res["ok"] = True
    except Exception as e:
        res["error"] = str(e)[:200]
    return res


def _safe_name(s):
    """파일명에 못 쓰는 문자 제거."""
    return re.sub(r'[\\/:*?"<>|]', "_", s).strip()


def crawl_all_accounts(start_de, end_de, includes, excludes, max_workers=10):
    """모든 배급사 계정 크롤 → 배급사별 개별 엑셀.

    반환: summary 리스트. 각 항목:
      {name, ok, error, row_count, movies, filename, xlsx(bytes 또는 None)}
    """
    start_de = start_de.replace("-", "")
    end_de = (end_de or start_de).replace("-", "")
    includes = [i for i in (includes or []) if i]
    excludes = [e for e in (excludes or []) if e]
    dates = _date_range(start_de, end_de)

    accounts = get_accounts()
    order = {a["name"]: i for i, a in enumerate(accounts)}
    results = []
    with ThreadPoolExecutor(max_workers=max_workers) as ex:
        futs = [
            ex.submit(crawl_one_account, a["name"], a["user"], a["password"],
                      dates, includes, excludes)
            for a in accounts
        ]
        for f in as_completed(futs):
            results.append(f.result())
    results.sort(key=lambda r: order.get(r["name"], 999))

    summary = []
    for r in results:
        item = {
            "name": r["name"],
            "ok": r["ok"],
            "error": r["error"],
            "row_count": r["row_count"],
            "movies": r["movies"],
            "filename": None,
            "xlsx": None,
        }
        rows = r.get("rows", [])
        if rows:
            rows.sort(key=lambda x: (x["play_date"], x["theater"], x["movie"],
                                     x["screen"], x["price"]))
            item["xlsx"] = rows_to_xlsx(rows)
            # 파일명의 '씨네큐' 키워드로 스코어 업로더가 씨네큐 양식 파서를 선택한다.
            item["filename"] = f"씨네큐_{_safe_name(r['name'])}_{start_de}_{end_de}.xlsx"
        summary.append(item)
    return summary
