"""롯데컬처웍스 moviesales 리포트 링크 → 회차별 판매현황 엑셀 추출.

메일 본문의 Linker.aspx 링크는 UbiReport 뷰어(iframe)로 연결되고, 그 리포트 HTML
안에는 전체 판매 데이터가 `streamdata:` 문자열(^t 탭 / ^n 줄바꿈 구분)로 통째로 들어
있다. 이를 파싱해 메일 첨부와 동일한 양식의 xlsx 로 재구성한다.
(별도 export 서버 호출 없이 리포트 HTML 한 번만 받으면 된다.)
"""

import io
import re
import time
import urllib.parse

import pandas as pd
import requests
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# SSRF 방지: 이 호스트의 링크만 허용한다.
LOTTE_HOST = "moviesales.lottecultureworks.co.kr"

_LINK_RE = re.compile(
    r"https?://moviesales\.lottecultureworks\.co\.kr/[^\s\"'<>]+", re.I
)
_IFRAME_RE = re.compile(r"<iframe[^>]+src=['\"]([^'\"]+)['\"]", re.I)
_STREAM_RE = re.compile(r"streamdata:\s*'(.*?)'\s*,", re.S)
_PLAYDT_RE = re.compile(r'PlayDt["\'\s+]*:["\'\s+]*(\d{8})')
_SUBJECT_DATE_RE = re.compile(r"(\d{4})\s*년\s*(\d{1,2})\s*월\s*(\d{1,2})\s*일")

_UA = "Mozilla/5.0 (Windows NT 10.0; Win64; x64)"

# 同 Linker 링크가 호출마다 다른 날짜 리포트를 반환할 수 있어(로드밸런싱 캐시),
# 메일이 가리키는 날짜를 만날 때까지 재시도하는 최대 횟수.
_MAX_TRIES = 15

# 메일 첨부(Excel Base Type)와 동일한 헤더/컬럼 순서
_HEADER = ["상영일자", "입회사", "영화", "대표영화관", "상영관", "상영회차", "발권금액", "매수", "합계"]

# ── UbiReport Excel(Base Type) 서식 재현용 상수 ──
_FONT_NAME = "맑은 고딕"
_COL_WIDTHS = [14.16, 15.54, 23.31, 13.81, 15.54, 16.4, 15.54, 11.22, 14.85]
_HEADER_FILL = "FFCCCCFF"
# 소계 단계별 배경색 (낮은 단계 → 높은 단계)
_SUBTOTAL_LABELS = [
    "상영회차별 소계", "상영관별 소계", "영화관별 소계", "영화별 소계", "입회사별 소계", "전체총계",
]
_LEVEL_FILLS = ["FFFAFAEA", "FFEDFAEA", "FFE9FAF4", "FFE9F4FA", "FFE9EAF9", "FFE4EDFD"]
# 우측 정렬 컬럼(0-based): 상영회차, 발권금액, 매수, 합계 (단, 소계 라벨이 든 칸은 좌측)
_RIGHT_COLS = {5, 6, 7, 8}


def parse_subject_date(subject):
    """메일 제목에서 'YYYY년 M월 D일' → 'YYYYMMDD'. 없으면 None."""
    if not subject:
        return None
    m = _SUBJECT_DATE_RE.search(subject)
    if not m:
        return None
    y, mo, d = m.groups()
    return f"{int(y):04d}{int(mo):02d}{int(d):02d}"


def find_report_links(subject, *texts):
    """본문에서 롯데 리포트 Linker 링크를 중복 없이 추출. 제목의 날짜를 함께 부착."""
    play_date = parse_subject_date(subject)
    found = []
    seen = set()
    for t in texts:
        if not t:
            continue
        for url in _LINK_RE.findall(t):
            url = url.replace("&amp;", "&")
            # 추적 픽셀(Check.html 등)이 아닌 Linker/Report 링크만
            if "Linker.aspx" not in url and "ReportViewer.aspx" not in url:
                continue
            if url in seen:
                continue
            seen.add(url)
            label = "회차별 판매현황 (엑셀 추출)"
            found.append({"url": url, "label": label, "play_date": play_date or ""})
    return found


def is_allowed_url(url):
    return bool(url) and LOTTE_HOST in url and url.lower().startswith("http")


def _resolve_report(session, linker_url):
    """Linker → (report_html_or_None, report_url, play_date).

    Linker 페이지면 iframe(ReportViewer) URL과 그 PlayDt 를 돌려준다.
    이미 리포트 페이지면 html 을 직접 돌려준다.
    """
    r = session.get(linker_url, timeout=30)
    r.raise_for_status()
    if "streamdata" in r.text:
        m = _PLAYDT_RE.search(urllib.parse.unquote_plus(r.text[:4000]))
        return r.text, linker_url, (m.group(1) if m else None)
    m = _IFRAME_RE.search(r.text)
    if not m:
        raise RuntimeError("리포트(iframe) URL을 찾지 못했습니다.")
    report_url = m.group(1)
    if LOTTE_HOST not in report_url:
        raise RuntimeError("허용되지 않은 리포트 호스트입니다.")
    dm = _PLAYDT_RE.search(urllib.parse.unquote_plus(report_url))
    return None, report_url, (dm.group(1) if dm else None)


def _parse_streamdata(html):
    m = _STREAM_RE.search(html)
    if not m:
        raise RuntimeError("리포트에서 데이터(streamdata)를 찾지 못했습니다.")
    _, _, body = m.group(1).partition("#")
    body = body.replace('\\"', '"')

    def clean(cell):
        return cell.strip().strip('"').strip("#").strip('"').strip()

    rows = []
    for line in body.split("^n"):
        if line.strip():
            rows.append([clean(x) for x in line.split("^t")])
    return rows


def _money(v):
    """숫자면 천단위 콤마, 빈값이면 '', 그 외('소계' 등 텍스트)는 원본 유지."""
    v = str(v).strip()
    if v == "":
        return ""
    try:
        return f"{int(float(v)):,}"
    except ValueError:
        return v


def _to_dataframe(rows):
    # streamdata: 입회사, 상영일자, 영화, 대표영화관, 상영관, 상영회차, 발권금액, 매수, 합계
    # 출력(원본):  상영일자, 입회사, 영화, 대표영화관, 상영관, 상영회차, 발권금액, 매수, 합계
    out = []
    for c in rows:
        if len(c) < 9:
            continue
        ipsa, date_raw, movie, theater, screen, show, fare, cnt, total = c[:9]
        date = date_raw.split(" ")[0].strip()
        show = str(show).strip()
        show_v = f"{show}회" if show.isdigit() else show
        cnt_raw = str(cnt).strip()
        if cnt_raw in ("", "0"):  # 매수 0 행은 원본처럼 매수/합계 공란
            cnt_v, total_v = "", ""
        else:
            cnt_v, total_v = _money(cnt), _money(total)
        out.append([date, ipsa, movie, theater, screen, show_v, _money(fare), cnt_v, total_v])
    return pd.DataFrame(out, columns=_HEADER)


def _build_xlsx_bytes(df):
    """메일 첨부(UbiReport Excel Base Type)와 동일한 서식(색/폰트/테두리/정렬/열너비)으로 생성.

    레이아웃: row1 제목(A1:I1 병합) / row2 공백 / row3 헤더 / row4~ 데이터·소계.
    (스코어 파서는 skiprows=2 로 읽음)
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    thin = Side(style="thin", color="FF000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    title_font = Font(name=_FONT_NAME, bold=True, underline="single", size=15, color="FF000000")
    bold_font = Font(name=_FONT_NAME, bold=True, size=8, color="FF000000")
    reg_font = Font(name=_FONT_NAME, size=8, color="FF000000")
    al_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    al_left = Alignment(vertical="center", wrap_text=True)
    al_right = Alignment(horizontal="right", vertical="center", wrap_text=True)
    header_fill = PatternFill("solid", fgColor=_HEADER_FILL)
    level_fills = [PatternFill("solid", fgColor=c) for c in _LEVEL_FILLS]

    for i, w in enumerate(_COL_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.sheet_format.defaultRowHeight = 16.5

    # row1: 제목
    ws.merge_cells("A1:I1")
    for c in range(1, 10):
        cell = ws.cell(row=1, column=c)
        cell.font = title_font
        cell.alignment = al_center
    ws["A1"] = "회차별 부금 계산서"

    # row3: 헤더
    for c, h in enumerate(_HEADER, start=1):
        cell = ws.cell(row=3, column=c, value=h)
        cell.font = bold_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = al_center

    # row4~: 데이터/소계
    r = 4
    for vals in df.itertuples(index=False, name=None):
        values = ["" if (v is None or (isinstance(v, float) and pd.isna(v))) else str(v) for v in vals]
        level = None
        label_col = None
        for i, v in enumerate(values):
            sv = v.strip()
            if sv in _SUBTOTAL_LABELS:
                level = _SUBTOTAL_LABELS.index(sv)
                label_col = i
                break
        font = reg_font if level is None else bold_font
        fill = None if level is None else level_fills[level]
        for c in range(9):
            cell = ws.cell(row=r, column=c + 1, value=values[c])
            cell.font = font
            cell.border = border
            if fill is not None:
                cell.fill = fill
            is_right = (c in _RIGHT_COLS) and (c != label_col)
            cell.alignment = al_right if is_right else al_left
        r += 1

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


_BROWSER_UA = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
    "(KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"
)


def _streamdata_via_browser(linker_url, expected_date=None, total_timeout=60):
    """헤드리스 브라우저(Playwright)로 Linker 를 열어 리포트의 streamdata 를 읽는다.

    롯데 서버는 헤드리스 GET 에는 캐시(엉뚱한 날짜)를 주지만, 실제 브라우저 엔진으로
    열면 요청한 날짜의 리포트를 정상 렌더한다. 그 iframe 의 inline streamdata 를 읽는다.
    expected_date(YYYYMMDD)가 주어지면 본문 첫 행 날짜가 일치할 때까지 대기한다.
    """
    import os
    from playwright.sync_api import sync_playwright

    os.environ["DJANGO_ALLOW_ASYNC_UNSAFE"] = "true"
    expected_date = (expected_date or "").strip() or None

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True)
        try:
            ctx = browser.new_context(user_agent=_BROWSER_UA)
            page = ctx.new_page()
            page.goto(linker_url, timeout=60000)

            deadline = time.time() + total_timeout
            last_sd = None
            while time.time() < deadline:
                frame = None
                for f in page.frames:
                    if "ReportViewer" in (f.url or ""):
                        frame = f
                        break
                if frame is not None:
                    try:
                        html = frame.content()
                    except Exception:
                        html = ""
                    m = _STREAM_RE.search(html)
                    if m and len(m.group(1)) > 10:
                        sd = m.group(1)
                        last_sd = sd
                        body = sd.split("#", 1)[1] if "#" in sd else ""
                        first = (
                            body.split("^n", 1)[0].replace('\\"', '"').split("^t")
                            if body else []
                        )
                        actual = (first[1][:10].replace("-", "") if len(first) > 1 else "")
                        if (not expected_date) or actual == expected_date:
                            return sd
                page.wait_for_timeout(1500)
            if last_sd is not None:
                return last_sd  # 날짜 일치 못 했지만 받은 데이터라도 반환
            raise RuntimeError("리포트 데이터를 불러오지 못했습니다(시간 초과).")
        finally:
            browser.close()


def extract_xlsx(linker_url, expected_date=None):
    """Linker URL → (filename, xlsx_bytes, row_count).

    헤드리스 브라우저로 리포트를 렌더해 streamdata 를 읽어 엑셀로 변환한다.
    """
    sd = _streamdata_via_browser(linker_url, expected_date)
    rows = _parse_streamdata("streamdata: '%s'," % sd)
    df = _to_dataframe(rows)
    xlsx = _build_xlsx_bytes(df)
    play_date = ""
    if not df.empty:
        play_date = str(df.iloc[0]["상영일자"]).replace("-", "")
    filename = f"롯데_회차별판매현황_{play_date or '추출'}.xlsx"
    return filename, xlsx, len(df)
