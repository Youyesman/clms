import re

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from django.http import HttpResponse
from decimal import Decimal  # Decimal 타입 체크를 위해 추가

# 엑셀 시트명에 쓸 수 없는 문자 (openpyxl이 ValueError를 던진다)
_INVALID_SHEET_CHARS = re.compile(r"[\\/*?:\[\]]")


class ExcelGenerator:
    def __init__(self, sheet_name="Data"):
        self.wb = openpyxl.Workbook()
        self.ws = self.wb.active
        # 영화 제목 등에 콜론(:)이 들어오면 시트명 규칙 위반으로 저장이 깨지므로
        # 금지 문자를 치환하고 최대 길이(31자)로 자른다
        safe_name = _INVALID_SHEET_CHARS.sub(" ", str(sheet_name)).strip() or "Data"
        self.ws.title = safe_name[:31]

        # 공통 스타일 정의
        self.header_fill = PatternFill(
            start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")  # 연한 파란색
        self.header_fill_green = PatternFill(
            start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")  # 연한 초록색 (총 좌석수, 총 상영관수용)
        self.header_font = Font(color="000000", bold=True, size=10)  # 검은색, 굵게, 10pt
        self.data_font = Font(color="000000", bold=False, size=10)  # 데이터 행용: 검은색, 일반, 10pt
        self.data_font_bold = Font(color="000000", bold=True, size=10)  # 극장명 컬럼용: 검은색, 굵게, 10pt
        self.center_align = Alignment(horizontal="center", vertical="center")
        self.left_align = Alignment(horizontal="left", vertical="center")  # 데이터 행용 왼쪽 정렬
        self.right_align = Alignment(
            horizontal="right", vertical="center")  # 숫자용 우측 정렬
        self.border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

    def add_header(self, headers, special_header_indices=None):
        """
        headers: 헤더 리스트
        special_header_indices: 특별한 배경색을 적용할 헤더 인덱스 리스트 (0-based)
        """
        self.ws.append(headers)
        for idx, cell in enumerate(self.ws[1]):
            # 특별한 헤더 인덱스가 지정되어 있고 현재 인덱스가 포함되어 있으면 초록색 배경
            if special_header_indices and idx in special_header_indices:
                cell.fill = self.header_fill_green
            else:
                cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = self.center_align
            cell.border = self.border

    def add_rows(self, data_list, bold_column_indices=None):
        """
        data_list: 데이터 행 리스트
        bold_column_indices: 굵게 표시할 컬럼 인덱스 리스트 (0-based, 예: [0]은 첫 번째 컬럼)
        """
        for row in data_list:
            self.ws.append(row)
            # 현재 추가된 마지막 행의 셀들을 순회
            for idx, cell in enumerate(self.ws[self.ws.max_row]):
                cell.border = self.border
                
                # 폰트 설정: 굵게 표시할 컬럼인지 확인
                if bold_column_indices and idx in bold_column_indices:
                    cell.font = self.data_font_bold
                else:
                    cell.font = self.data_font

                # ✅ 데이터가 숫자 타입(int, float, Decimal)인지 확인
                if isinstance(cell.value, (int, float, Decimal)):
                    # 천 단위 콤마 서식 적용 (#,##0)
                    cell.number_format = '#,##0'
                    # 숫자는 오른쪽 정렬이 가독성이 좋음
                    cell.alignment = self.right_align
                else:
                    # 일반 텍스트는 왼쪽 정렬, 수직 중앙 정렬 (예시 파일과 동일)
                    cell.alignment = self.left_align

    def auto_fit_columns(self):
        """콘텐츠 길이에 맞춰 열 너비 자동 조절.

        미리 지정해 둔 열 너비가 있으면 그보다 줄이지 않는다 — 수식 셀(SUM 총계 등)은
        계산값 길이를 여기서 알 수 없으므로 호출측이 최소 폭을 지정하는 용도.
        """
        for col in self.ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    # 수식 셀은 화면에 계산값(숫자)이 보이므로 수식 문자열 길이로
                    # 폭을 잡으면 열만 쓸데없이 넓어진다 → 너비 계산에서 제외
                    if isinstance(cell.value, str) and cell.value.startswith("="):
                        continue
                    # 숫자 셀은 #,##0 서식으로 콤마가 붙어 표시되므로 콤마 포함
                    # 길이로 계산해야 ####(폭 부족)이 안 생긴다
                    if isinstance(cell.value, (int, float, Decimal)):
                        length = len(f"{cell.value:,.0f}")
                    else:
                        length = len(str(cell.value))
                    if length > max_length:
                        max_length = length
                except:
                    pass
            # 한글/숫자 폰트 크기를 고려해 여유공간(+3) 추가
            prev = self.ws.column_dimensions[column].width or 0
            self.ws.column_dimensions[column].width = max(max_length + 3, prev)

    def to_response(self, filename, auto_fit=True):
        # auto_fit=False: 호출부에서 열 너비를 양식대로 지정한 경우 덮어쓰지 않는다
        if auto_fit:
            self.auto_fit_columns()
        # 다운로드 파일명도 OS 금지 문자(: 등)를 치환 — 영화 제목이 그대로 들어온다
        safe_filename = re.sub(r'[\\/:*?"<>|]', "_", str(filename))
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{safe_filename}.xlsx"'
        self.wb.save(response)
        return response
