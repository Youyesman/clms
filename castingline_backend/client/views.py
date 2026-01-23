from django.db.models.functions import Collate
from .models import Client, Theater
from .serializers import *
from rest_framework import viewsets, permissions
from rest_framework.authentication import TokenAuthentication
from rest_framework.exceptions import PermissionDenied
from rest_framework.pagination import PageNumberPagination
from rest_framework import filters
from rest_framework.permissions import AllowAny
from django.db.models import F, Sum, Count, Case, When, IntegerField, Value
from django.db.models.functions import Cast, Coalesce
from datetime import datetime
from rest_framework.views import APIView
from django.http import JsonResponse
from difflib import get_close_matches
from .models import Client
from django.contrib.postgres.search import TrigramSimilarity
from django_filters.rest_framework import DjangoFilterBackend
from .filters import ClientFilter
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework import status
from castingline_backend.utils.ordering import KoreanOrderingFilter
from django.db.models import Max, Subquery, OuterRef
from castingline_backend.utils.excel_helper import ExcelGenerator
from openpyxl.styles import Alignment
from decimal import Decimal


class DefaultPagination(PageNumberPagination):
    page_size = 20  # 한 페이지에 보여질 항목 수 설정
    page_size_query_param = "page_size"
    max_page_size = 100  # 최대 몇개 항목까지 보여줄건지?


class ClientViewSet(viewsets.ModelViewSet):
    queryset = Client.objects.all()
    serializer_class = ClientSerializer
    authentication_classes = []
    permission_classes = [AllowAny]
    pagination_class = DefaultPagination
    filter_backends = [
        DjangoFilterBackend,
        filters.SearchFilter,
        KoreanOrderingFilter,
    ]
    filterset_class = ClientFilter
    search_fields = ["client_code", "client_name", "operational_status"]
    ordering_fields = "__all__"

    ordering = ["theater_kind", "classification", "region_code", "client_name"]

    @action(detail=False, methods=["patch"])
    def bulk_update_settlement(self, request):
        target_dept = request.data.get("target_department")

        if not target_dept:
            return Response(
                {"error": "부금처 정보가 없습니다."}, status=status.HTTP_400_BAD_REQUEST
            )

        # 업데이트할 필드 목록
        fields_to_update = [
            "settlement_contact",
            "settlement_phone_number",
            "settlement_mobile_number",
            "settlement_email",
            "invoice_email_address",
            "invoice_email_address2",
            "settlement_remarks",
        ]

        # 전송된 데이터 중 위 필드에 해당하는 것만 추출
        update_data = {
            field: request.data.get(field)
            for field in fields_to_update
            if field in request.data
        }

        # 해당 부금처를 가진 모든 거래처 업데이트
        updated_count = Client.objects.filter(settlement_department=target_dept).update(
            **update_data
        )

        return Response(
            {
                "message": f"{updated_count}개의 거래처 정보가 업데이트되었습니다.",
                "updated_count": updated_count,
            },
            status=status.HTTP_200_OK,
        )

    def create(self, request, *args, **kwargs):
        # 현재 연도 가져오기 (예: 2025)
        current_year = datetime.now().year

        # 해당 연도의 가장 큰 code 값을 찾기
        last_theater = (
            Client.objects.filter(client_code__startswith=str(current_year))
            .order_by("-client_code")
            .first()
        )

        if last_theater:
            # 가장 큰 code에서 시퀀스 번호 추출 후 1 증가
            last_sequence = int(
                str(last_theater.client_code)[4:]
            )  # 예: "20250001" -> 0001
            new_sequence = last_sequence + 1
        else:
            # 해당 연도에 극장이 없으면 0001로 시작
            new_sequence = 1

        # 새로운 code 생성 (예: 20250001)
        new_code = int(
            f"{current_year}{new_sequence:04d}"
        )  # :04d는 4자리 숫자로 포맷팅

        # 요청 데이터에 code 추가
        request.data["client_code"] = new_code

        return super().create(request, *args, **kwargs)


class TheaterViewSet(viewsets.ModelViewSet):
    queryset = Theater.objects.all()
    serializer_class = TheaterSerializer
    authentication_classes = []
    permission_classes = [AllowAny]
    pagination_class = DefaultPagination
    filter_backends = [filters.SearchFilter]
    search_fields = ["number", "name"]  # 검색 필드 추가

    def get_queryset(self):
        queryset = super().get_queryset()
        client_id = self.request.query_params.get("client_id", None)
        if client_id:
            queryset = queryset.filter(client__id=client_id)
        return queryset


class FareViewSet(viewsets.ModelViewSet):
    queryset = Fare.objects.all()
    serializer_class = FareSerializer
    authentication_classes = []
    permission_classes = [AllowAny]
    pagination_class = DefaultPagination
    filter_backends = [filters.SearchFilter]
    search_fields = ["rate"]  # 검색 필드 추가

    def get_queryset(self):
        queryset = super().get_queryset()
        client_id = self.request.query_params.get("client_id", None)
        if client_id:
            queryset = queryset.filter(client__id=client_id)
        return queryset


class FindSimilarClient(APIView):
    def get(self, request):
        name = request.GET.get("name", "").strip()

        if not name:
            return JsonResponse({"error": "No name provided"}, status=400)

        # 정확히 일치하는 경우 우선 반환
        exact_match = Client.objects.filter(client_name=name).first()
        if exact_match:
            return JsonResponse(
                {
                    "id": exact_match.id,
                    "client_name": exact_match.client_name,
                    "client_type": exact_match.client_type,
                }
            )

        # TrigramSimilarity를 통한 유사도 검색
        similar_client = (
            Client.objects.annotate(
                similarity=TrigramSimilarity("client_name", name))
            .filter(similarity__gt=0.1)  # 유사도 임계값 (0~1)
            .order_by("-similarity")
            .first()
        )

        if similar_client:
            return JsonResponse(
                {
                    "id": similar_client.id,
                    "client_name": similar_client.client_name,
                    "client_type": similar_client.client_type,
                }
            )

        return JsonResponse({"error": "No similar client found"}, status=404)


class TheaterMapDistributorListView(APIView):
    """
    극장명 매핑 관리를 위한 배급사 목록 전용 API
    """

    authentication_classes = []
    permission_classes = []

    def get(self, request):
        # 1. distributor_theater_name이 "배급사별 극장명"이고
        # 2. client_type이 ["제작사", "배급사"] 리스트에 포함된 객체들만 필터링
        distributors = (
            Client.objects.filter(
                distributor_theater_name="배급사별 극장명",
                client_type__in=[
                    "제작사",
                    "배급사",
                ],  # ✅ IN 문법 사용 (제작사 OR 배급사)
            )
            .values("id", "client_name")
            .order_by("client_name")
        )

        # 리스트로 변환하여 JSON 응답
        return Response(list(distributors), status=status.HTTP_200_OK)


class TheaterMapViewSet(viewsets.ModelViewSet):
    queryset = DistributorTheaterMap.objects.all()
    serializer_class = TheaterMapSerializer
    pagination_class = DefaultPagination
    filter_backends = [
        DjangoFilterBackend,
        filters.SearchFilter,
        KoreanOrderingFilter,
    ]

    # 기본 필터 허용 필드
    filterset_fields = ["distributor", "theater", "apply_date"]
    search_fields = [
        "distributor_theater_name",
        "theater__client_name",
        "theater__client_code",
    ]
    ordering_fields = [
        "id",
        "apply_date",
        "distributor_theater_name",
        "theater__client_name",
        "theater__client_code",
    ]

    ordering = ["-apply_date", "id"]

    def get_queryset(self):
        queryset = super().get_queryset()

        # 1. 배급사 ID 필터링 (기본)
        dist_id = self.request.query_params.get("distributor")
        if dist_id:
            queryset = queryset.filter(distributor_id=dist_id)

        # 2. 추가 필터링 (극장 상세 정보 기반)
        # 극장 Autocomplete 검색 (시스템 극장 ID)
        theater_id = self.request.query_params.get("theater")
        if theater_id:
            queryset = queryset.filter(theater_id=theater_id)

        # 상태 (client_status)
        status = self.request.query_params.get("operational_status")
        if status:

            status = status == "true"
            print(status)
            queryset = queryset.filter(theater__operational_status=status)

        # 구분 (classification)
        classification = self.request.query_params.get("classification")
        if classification:
            queryset = queryset.filter(theater__classification=classification)

        # 멀티 (theater_kind)
        multi = self.request.query_params.get("theater_kind")
        if multi:
            queryset = queryset.filter(theater__theater_kind=multi)

        # 3. 최신 이력만 보기 (?latest=true)
        latest_param = self.request.query_params.get("latest")
        if latest_param == "true":
            latest_subquery = DistributorTheaterMap.objects.filter(
                distributor=OuterRef("distributor"), theater=OuterRef("theater")
            ).order_by("-apply_date", "-id")
            queryset = queryset.filter(id=Subquery(
                latest_subquery.values("id")[:1]))

        return queryset


class ClientExcelExportView(APIView):
    def get(self, request):
        """
        프론트엔드에서 보낸 쿼리 파라미터를 그대로 사용하여 필터링
        페이지네이션은 무시하고 필터된 전체 데이터를 엑셀로 출력
        """
        # 리스트 뷰와 동일한 필터 조건 재사용
        viewset = ClientViewSet()
        viewset.request = request
        viewset.format_kwarg = None

        # 기본 queryset 가져오기 (페이지네이션 파라미터는 무시됨)
        queryset = viewset.get_queryset()

        # 프론트엔드에서 보낸 쿼리 파라미터로 필터 적용 (ClientFilter + SearchFilter)
        queryset = viewset.filter_queryset(queryset)

        # Annotate total_seats and total_screens to avoid N+1 queries
        queryset = queryset.annotate(
            total_seats=Coalesce(
                Sum(
                    Case(
                        When(
                            theater_client__seat_count__regex=r"^\d+$",
                            then=Cast("theater_client__seat_count", IntegerField()),
                        ),
                        default=Value(0),
                        output_field=IntegerField(),
                    )
                ),
                Value(0),
            ),
            total_screens=Count("theater_client"),
        )

        # 3. 엑셀 생성 (ExcelGenerator 활용)
        excel = ExcelGenerator(sheet_name="전국 극장 자료")
        
        # [Sheet 1] 전국 극장 자료
        # 예시 파일과 동일한 헤더 정의 (23개 컬럼)
        headers = [
            "극장명", "바이포엠 극장코드", "직위", "지역", "멀티",
            "종사업자", "사업자번호", "사업자명(정식명)", "업태", "업종",
            "사업장소재지", "대표자명", "부금처", "부금담당자 휴대폰",
            "전화번호(부금)", "팩스번호", "담당자(부금)", "전화번호(대표)",
            "세금계산서 발행메일", "세금계산서 발행메일2", "부금 특이사항",
            "총 좌석수", "총 상영관수"
        ]
        # 총 좌석수(21번째, 인덱스 21)와 총 상영관수(22번째, 인덱스 22) 헤더는 초록색 배경
        excel.add_header(headers, special_header_indices=[21, 22])

        # queryset을 리스트로 변환하여 정렬
        clients_list = list(queryset)
        
        # CGV -> 롯데 -> 메가 -> 나머지극장 순서로 정렬
        def get_theater_sort_key(client):
            theater_kind = (client.theater_kind or "").lower()
            if "cgv" in theater_kind:
                return (0, theater_kind, client.client_name or "")
            elif "롯데" in theater_kind or "lotte" in theater_kind:
                return (1, theater_kind, client.client_name or "")
            elif "메가" in theater_kind or "mega" in theater_kind:
                return (2, theater_kind, client.client_name or "")
            else:
                return (3, theater_kind, client.client_name or "")
        
        clients_list.sort(key=get_theater_sort_key)

        data_rows = []
        for c in clients_list:
            # Theater 정보 집계 (총 좌석수, 총 상영관수) - Already annotated
            total_seats = c.total_seats
            total_screens = c.total_screens

            # 부금담당자 휴대폰 (settlement_mobile_number)
            settlement_mobile = c.settlement_mobile_number or ""
            
            data_rows.append([
                c.client_name or "",  # 극장명
                c.by4m_theater_code or "",  # 바이포엠 극장코드
                c.classification or "",  # 직위
                c.region_code or "",  # 지역
                c.theater_kind or "",  # 멀티
                c.business_operator or "",  # 종사업자
                c.business_registration_number or "",  # 사업자번호
                c.business_name or "",  # 사업자명(정식명)
                c.business_category or "",  # 업태
                c.business_industry or "",  # 업종
                c.business_address or "",  # 사업장소재지
                c.representative_name or "",  # 대표자명
                c.settlement_department or "",  # 부금처
                settlement_mobile,  # 부금담당자 휴대폰 (합쳐서 표시)
                c.settlement_phone_number or "",  # 전화번호(부금)
                c.fax_number or "",  # 팩스번호
                c.settlement_contact or "",  # 담당자(부금)
                c.representative_phone_number or "",  # 전화번호(대표)
                c.invoice_email_address or "",  # 세금계산서 발행메일 (이메일 주소)
                c.invoice_email_address2 or "",  # 세금계산서 발행메일2 (이메일 주소)
                c.settlement_remarks or "",  # 부금 특이사항
                total_seats,  # 총 좌석수
                total_screens,  # 총 상영관수
            ])

        # 극장명 컬럼(첫 번째, 인덱스 0)은 굵게 표시
        excel.add_rows(data_rows, bold_column_indices=[0])
        
        # 데이터 행의 처음 6개 컬럼(극장명~종사업자)은 가운데 정렬
        for row_idx in range(2, excel.ws.max_row + 1):
            for col_idx in range(1, 7):  # A~F 열 (1~6번째 컬럼)
                cell = excel.ws.cell(row=row_idx, column=col_idx)
                if not isinstance(cell.value, (int, float, Decimal)):
                    # 숫자가 아닌 경우만 가운데 정렬
                    cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # A열 고정 (freeze_panes)
        excel.ws.freeze_panes = 'B2'  # B2를 기준으로 A열과 1행 고정

        # [Sheet 2] 전국 극장 관 현황
        ws2 = excel.wb.create_sheet("전국 극장 관 현황")
        excel.ws = ws2  # Helper가 이제 두 번째 시트를 참조하도록 변경
        
        headers2 = ["극장명", "관(번호)", "좌석수", "관이름"]
        excel.add_header(headers2)
        
        # 검색된 거래처(clients_list)에 속한 상영관들만 조회
        # queryset은 annotate된 상태이므로, 단순 ID 리스트로 변환하여 필터링
        client_ids = [c.id for c in clients_list]
        theaters = Theater.objects.filter(client_id__in=client_ids).select_related('client')
        
        # DB 정렬 대신 Python 정렬 사용 (글자수 뭉침 방지)
        theaters_list = list(theaters)
        import re
        def theater_sort_key(t):
            name = t.client.client_name if t.client else ""
            auditorium = t.auditorium or ""
            
            # 우선순위: 0=특수문자, 1=한글, 2=영어/숫자
            if re.match(r'^[^0-9a-zA-Z가-힣]', name):
                prio = 0
            elif re.match(r'^[가-힣]', name):
                prio = 1
            else:
                prio = 2
            return (prio, name, auditorium)
            
        theaters_list.sort(key=theater_sort_key)
        
        data_rows2 = []
        for t in theaters_list:
            data_rows2.append([
                t.client.client_name if t.client else "미등록 극장",
                t.auditorium,
                int(t.seat_count) if t.seat_count and t.seat_count.isdigit() else 0,
                t.auditorium_name
            ])
        excel.add_rows(data_rows2)

        filename = f"Client_Export_{datetime.now().strftime('%Y%m%d')}"
        return excel.to_response(filename)


class TheaterExcelExportView(APIView):
    def get(self, request):
        """
        프론트엔드에서 보낸 쿼리 파라미터를 그대로 사용하여 필터링
        페이지네이션은 무시하고 필터된 전체 데이터를 엑셀로 출력
        """
        # 리스트 뷰와 동일한 필터 조건 재사용
        viewset = TheaterViewSet()
        viewset.request = request
        viewset.format_kwarg = None

        # 기본 queryset 가져오기 (get_queryset에서 client_id 필터 적용)
        queryset = viewset.get_queryset()

        # 프론트엔드에서 보낸 쿼리 파라미터로 필터 및 검색 적용
        queryset = viewset.filter_queryset(queryset)

        # 정렬
        queryset = queryset.annotate(
            client_name=F('client__client_name')
        ).order_by('client__client_name', 'auditorium')

        # 2. 엑셀 생성기 초기화
        excel = ExcelGenerator(sheet_name="극장관정보")

        # 3. 헤더 정의
        headers = ["극장명", "관(번호)", "좌석수", "관이름"]
        excel.add_header(headers)

        # 4. 데이터 행 구성
        data_rows = []
        for t in queryset:
            data_rows.append([
                t.client_name or "미등록 극장",
                t.auditorium,
                int(t.seat_count) if t.seat_count and t.seat_count.isdigit() else 0,
                t.auditorium_name
            ])

        excel.add_rows(data_rows)

        # 5. 응답 반환
        filename = f"Theater_List_{datetime.now().strftime('%Y%m%d')}"
        return excel.to_response(filename)


class TheaterMapExcelExportView(APIView):
    def get(self, request):
        """
        프론트엔드에서 보낸 쿼리 파라미터를 그대로 사용하여 필터링
        페이지네이션은 무시하고 필터된 전체 데이터를 엑셀로 출력
        """
        # 배급사 선택 필수 확인
        if not request.query_params.get("distributor"):
            return Response(
                {"detail": "배급사를 선택해주세요."}, 
                status=status.HTTP_400_BAD_REQUEST
            )

        # 리스트 뷰와 동일한 필터 조건 재사용
        viewset = TheaterMapViewSet()
        viewset.request = request
        viewset.format_kwarg = None

        # 기본 queryset 가져오기
        queryset = viewset.get_queryset()

        # 프론트엔드에서 보낸 쿼리 파라미터로 필터 및 검색 적용
        queryset = viewset.filter_queryset(queryset)

        # annotate 추가
        queryset = queryset.annotate(
            t_code=F('theater__client_code'),
            t_name=F('theater__client_name'),
        )
        
        # DB 정렬 대신 Python 정렬 사용 (특수문자 -> 한글 -> 영어)
        maps_list = list(queryset)
        import re
        def map_sort_key(m):
            name = m.t_name or ""
            if re.match(r'^[^0-9a-zA-Z가-힣]', name):
                return (0, name)
            elif re.match(r'^[가-힣]', name):
                return (1, name)
            else:
                return (2, name)
        maps_list.sort(key=map_sort_key)

        # 4. 엑셀 생성기 초기화
        excel = ExcelGenerator(sheet_name="극장명매핑현황")
        # 헤더 변경: 극장코드 / 시스템상 극장명 / 배급사측 지정명(현재) / 최종 적용일
        headers = ["극장코드", "시스템상 극장명", "배급사측 지정명(현재)", "최종 적용일"]
        excel.add_header(headers)

        # 5. 데이터 행 구성
        data_rows = []
        for m in maps_list:
            data_rows.append([
                m.t_code,
                m.t_name,
                m.distributor_theater_name,
                m.apply_date.strftime("%Y-%m-%d") if m.apply_date else "",
            ])

        excel.add_rows(data_rows)

        # 6. 응답 반환
        filename = f"TheaterMap_Export_{datetime.now().strftime('%Y%m%d')}"
        return excel.to_response(filename)