# -*- coding: utf-8 -*-
"""T001·T003(0827): 시간표 조회 — 화면 그대로 엑셀/PDF 내보내기.

- '상영일자 추이' 그래프는 제외하고, 화면의 표들을 그대로 담는다.
- 두 출력물 모두 우측 상단에 캐스팅라인 로고를 넣는다.
- 숫자는 timetable_agg 의 화면용 데이터만 사용하므로 화면과 완전히 일치한다.
"""
import io

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XlImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from crawler.report.common import find_korean_font, flat_logo_bytes

# ---------- 엑셀 공통 ----------
FONT = "맑은 고딕"
_side = Side(style="thin", color="BFBFBF")
BORDER = Border(left=_side, right=_side, top=_side, bottom=_side)
HEAD_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
TOTAL_FILL = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
TITLE_FONT = Font(name=FONT, size=14, bold=True)
BOLD = Font(name=FONT, size=10, bold=True)
NORM = Font(name=FONT, size=10)
RED_BOLD = Font(name=FONT, size=9, bold=True, color="D83A34")
BLUE_BOLD = Font(name=FONT, size=9, bold=True, color="1D4ED8")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _put(ws, r, c, v, font=NORM, fill=None, fmt=None, border=True):
    cell = ws.cell(row=r, column=c, value=v)
    cell.font = font
    cell.alignment = CENTER
    if border:
        cell.border = BORDER
    if fill:
        cell.fill = fill
    if fmt:
        cell.number_format = fmt
    return cell


def _row(ws, r, values, font=NORM, fill=None, fmts=None):
    for i, v in enumerate(values, 1):
        fmt = None
        if fmts:
            fmt = fmts.get(i)
        elif isinstance(v, int):
            fmt = "#,##0"
        _put(ws, r, i, v, font=font, fill=fill, fmt=fmt)
    return r + 1


def _logo(ws, last_col):
    """우측 상단 캐스팅라인 로고 (V001: 투명 가장자리 평탄화본 — 회색 바 방지)."""
    try:
        img = XlImage(io.BytesIO(flat_logo_bytes()))
        h = 40
        img.width = int(img.width * h / img.height)
        img.height = h
        ws.add_image(img, f"{get_column_letter(max(last_col - 1, 1))}1")
    except Exception:
        pass


def _autow(ws, min_width=8, max_width=30):
    for i, col in enumerate(ws.columns, 1):
        m = 0
        for c in col:
            if c.value is None:
                continue
            v = str(c.value)
            m = max(m, sum(2 if ord(ch) > 127 else 1 for ch in v.split("\n")[0]))
        ws.column_dimensions[get_column_letter(i)].width = min(max((m + 2) * 1.05, min_width), max_width)


def _cmp_txt(c, unit=""):
    if not c:
        return ""
    d = c.get("diff", 0)
    arrow = "▲" if d >= 0 else "▼"
    rate = c.get("rate")
    rate_s = f" ({rate:+.1f}%)" if rate is not None else ""
    return f"{arrow} {d:+,d}{unit}{rate_s}"


PCT = '0.0"%"'

# =====================================================================
# T001: 주요작 시간표 엑셀 (B002/0829 — 일자별 탭 = 시트 1개)
# =====================================================================

def _cmp_seat_txt(c):
    """전일比/전주比 좌석 증감 — 비교 대상이 없으면 '-'."""
    if not c:
        return "-"
    d = c["diff"]
    return f"{'▲' if d >= 0 else '▼'} {d:+,d}석"


def _cmp_pp_txt(c):
    """점유율 차이 — %p 표기."""
    if not c:
        return "-"
    d = c["diff"]
    return f"{'▲' if d >= 0 else '▼'} {abs(d):.1f}%p"


def _cmp_kpi_txt(c, unit):
    if not c:
        return "-"
    return _cmp_txt(c, unit)


def _move_txt(v):
    """순위 변동 — 양수면 상승(▲), 음수면 하락(▼), 0이면 '-'."""
    if v is None:
        return "-"
    if v == 0:
        return "-"
    return f"{'▲' if v > 0 else '▼'} {abs(v)}"


def timetable_excel_bytes(data):
    meta = data["meta"]
    wb = Workbook()
    wb.remove(wb.active)

    for tab in data["tabs"]:
        ks = tab["key_summary"]
        ws = wb.create_sheet(tab["label"].replace("/", ".")[:28])
        ws.sheet_view.showGridLines = False

        _put(ws, 1, 1, f"주요작 시간표 — {meta['movie_title']} · {tab['label']}",
             font=TITLE_FONT, border=False)
        _put(ws, 2, 1,
             f"조사기간 {meta['date_from']} ~ {meta['date_to']}"
             f"  |  전일 {ks['prev_day']}  |  전주 {ks['prev_week']}"
             f"  |  개봉일 {meta.get('release_date') or '-'}"
             f"  |  배급사 {meta.get('distributor_name') or '-'}"
             f"  |  수집 완료 {meta.get('last_crawled_at') or '-'}",
             font=NORM, border=False)
        _logo(ws, 7)

        r = 4
        # ---- ① KEY SUMMARY ----
        r = _row(ws, r, ["KEY SUMMARY"], font=BOLD)
        r = _row(ws, r, ["구분", "총 좌석수", "예매좌석수", "좌석점유율",
                         "총 회차수", "총 극장수", "총 스크린수"], font=BOLD, fill=HEAD_FILL)
        r = _row(ws, r, [ks["label"], ks["total_seats"], ks["sold_seats"], ks["occupancy"],
                         ks["shows"], ks["theaters"], ks["screens"]],
                 font=BOLD, fill=TOTAL_FILL,
                 fmts={2: "#,##0", 3: "#,##0", 4: PCT, 5: "#,##0", 6: "#,##0", 7: "#,##0"})
        for label, key in (("증감 (전일比)", "prev_day_cmp"), ("증감 (전주比)", "prev_week_cmp")):
            c = ks.get(key)
            vals = [label]
            if c:
                vals += [_cmp_kpi_txt(c["total_seats"], "석"), _cmp_kpi_txt(c["sold_seats"], "석"),
                         _cmp_pp_txt(c["occupancy"]), _cmp_kpi_txt(c["shows"], "회"),
                         _cmp_kpi_txt(c["theaters"], "개"), _cmp_kpi_txt(c["screens"], "개")]
            else:
                vals += ["-"] * 6
            for i, v in enumerate(vals, 1):
                _put(ws, r, i, v, font=RED_BOLD if i > 1 else BOLD)
            r += 1
        r += 1

        # ---- ②③⑤ 멀티사별 / 포맷별 / 지역별 ----
        def detail_block(r, title, rows, count_label):
            r = _row(ws, r, [title], font=BOLD)
            r = _row(ws, r, ["구분", "총 좌석수", "비율", "전일比 좌석 증감",
                             "전주比 좌석 증감", count_label, "회차수"],
                     font=BOLD, fill=HEAD_FILL)
            for row in rows:
                r = _row(ws, r, [row["label"], row["total_seats"], row["share"],
                                 _cmp_seat_txt(row["prev_day_cmp"]),
                                 _cmp_seat_txt(row["prev_week_cmp"]),
                                 row["count"], row["shows"]],
                         fmts={2: "#,##0", 3: PCT, 6: "#,##0", 7: "#,##0"})
            return r + 1

        r = detail_block(r, "멀티사별 상세 현황", tab["multi_detail"], "총 극장수")
        r = detail_block(r, "포맷별 상세 현황", tab["format_detail"], "스크린수")

        # ---- ④ 시간대별 ----
        r = _row(ws, r, ["시간대별 상세 현황"], font=BOLD)
        r = _row(ws, r, ["시간대 구분", "상영 회차수", "총 좌석수", "예매 좌석수",
                         "좌석 점유율", "전일比 점유율 차이", "전주比 점유율 차이"],
                 font=BOLD, fill=HEAD_FILL)
        for row in tab["time_detail"]:
            r = _row(ws, r, [row["label"], row["shows"], row["total_seats"], row["sold_seats"],
                             row["occupancy"], _cmp_pp_txt(row["prev_day_cmp"]),
                             _cmp_pp_txt(row["prev_week_cmp"])],
                     fmts={2: "#,##0", 3: "#,##0", 4: "#,##0", 5: PCT})
        r += 1

        r = detail_block(r, "지역별 상세 현황", tab["region_detail"], "총 극장수")

        # ---- ⑥ 주요작 vs 경쟁작 ----
        r = _row(ws, r, [f"주요작 vs 경쟁작 — 동시 상영 경쟁작 TOP 10 순위 ({tab['label']})"], font=BOLD)
        r = _row(ws, r, ["순위", "영화명", "총 좌석수", "좌석 점유율", "상영 회차수",
                         "전일比 순위변동", "전주比 순위변동"], font=BOLD, fill=HEAD_FILL)
        for row in tab["competitor_top"]:
            r = _row(ws, r,
                     [f"{row['rank']} (★당사)" if row["is_main"] else row["rank"],
                      row["title"], row["total_seats"], row["occupancy"], row["shows"],
                      _move_txt(row["prev_day_move"]), _move_txt(row["prev_week_move"])],
                     font=BOLD if row["is_main"] else NORM,
                     fill=TOTAL_FILL if row["is_main"] else None,
                     fmts={3: "#,##0", 4: PCT, 5: "#,##0"})

        _autow(ws)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# =====================================================================
# T003: 경쟁작 엑셀
# =====================================================================

def competitor_excel_bytes(data):
    wb = Workbook()
    wb.remove(wb.active)
    meta = data["meta"]

    for tab in data["tabs"]:
        title = tab["label"].replace("/", ".")
        ws = wb.create_sheet(title[:28])
        ws.sheet_view.showGridLines = False
        _put(ws, 1, 1, f"경쟁작 성과 종합 — {tab['label']}", font=TITLE_FONT, border=False)
        _put(ws, 2, 1, f"조사기간 {meta['date_from']} ~ {meta['date_to']}"
                       f"  |  경쟁작 {meta['movie_count']}편  |  수집 완료 {meta.get('last_crawled_at') or '-'}",
             font=NORM, border=False)
        _logo(ws, 8)
        r = 4

        # ① 종합 요약
        r = _row(ws, r, [f"{meta['movie_count']}개 경쟁작 성과 종합 요약"], font=BOLD)
        r = _row(ws, r, ["순위", "영화명", "총 좌석수", "평균 좌석점유율", "총 상영회차",
                         "스크린수", "총 극장수"], font=BOLD, fill=HEAD_FILL)
        for i, s in enumerate(tab["summary"]):
            fill = TOTAL_FILL if i == 0 else None
            r = _row(ws, r, [s["rank"], ("★ " if i == 0 else "") + s["title"], s["total_seats"],
                             s["occupancy"], s["shows"], s["screens"], s["theaters"]],
                     font=BOLD if i == 0 else NORM, fill=fill,
                     fmts={3: "#,##0", 4: PCT, 5: "#,##0", 6: "#,##0", 7: "#,##0"})
        r += 1

        # ② 권역별
        r = _row(ws, r, ["서울 및 주요 권역별 좌석 점유 현황"], font=BOLD)
        r = _row(ws, r, ["영화명", "서울 좌석수", "서울 좌점율", "서울 비중",
                         "수도권(경강) 좌석수", "수도권 좌점율",
                         "그 외 지방 좌석수", "지방 좌점율"], font=BOLD, fill=HEAD_FILL)
        for row in tab["regions"]:
            r = _row(ws, r, [row["title"],
                             row["seoul"]["seats"], row["seoul"]["occupancy"], row["seoul"]["share"],
                             row["metro"]["seats"], row["metro"]["occupancy"],
                             row["local"]["seats"], row["local"]["occupancy"]],
                     fmts={2: "#,##0", 3: PCT, 4: PCT, 5: "#,##0", 6: PCT, 7: "#,##0", 8: PCT})
        r += 1

        # ③ 골든타임
        r = _row(ws, r, ["골든타임(14~21시) 집중도"], font=BOLD)
        r = _row(ws, r, ["영화명", "골든타임 좌석수", "골든타임 점유율", "골든타임 비중"],
                 font=BOLD, fill=HEAD_FILL)
        for row in tab["golden"]:
            r = _row(ws, r, [row["title"], row["seats"], row["occupancy"], row["share"]],
                     fmts={2: "#,##0", 3: PCT, 4: PCT})
        r += 1

        # ④ 특별관 — V004(0831): 타입별(IMAX/4DX/SCREENX/Dolby) 개별 표
        if tab["special"]:
            for blk in tab["special"]:
                r = _row(ws, r, [f"특별관 ({blk['format']})"], font=BOLD)
                r = _row(ws, r, ["영화명", "특별관 회차", "특별관 좌석수"],
                         font=BOLD, fill=HEAD_FILL)
                for row in blk["rows"]:
                    r = _row(ws, r, [row["title"], row["shows"], row["seats"]],
                             fmts={2: "#,##0", 3: "#,##0"})
                r += 1
        else:
            r = _row(ws, r, ["특별관 (IMAX/4DX/SCREENX/Dolby)"], font=BOLD)
            r = _row(ws, r, ["영화명", "특별관 회차", "특별관 좌석수"], font=BOLD, fill=HEAD_FILL)
            r = _row(ws, r, ["특별관 상영 데이터 없음", "", ""])
            r += 1

        # ⑤ 계열사별 세부 현황
        bb = tab["by_brand"]
        if bb["movies"]:
            r = _row(ws, r, ["계열사별 세부 현황"], font=BOLD)
            r = _row(ws, r, ["구분"] + bb["movies"], font=BOLD, fill=HEAD_FILL)
            for row in bb["rows"]:
                vals = [row["brand"]] + [
                    f"{c['seats']:,}석 ({c['share']:.1f}%)" if c["seats"] else "0석 (0.0%)"
                    for c in row["cells"]]
                r = _row(ws, r, vals)

        _autow(ws)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# =====================================================================
# PDF 공통 (reportlab)
# =====================================================================

def _pdf_doc(buf, title):
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.units import mm
    from reportlab.platypus import BaseDocTemplate, Frame, PageTemplate
    from reportlab.lib.utils import ImageReader

    page_size = landscape(A4)
    margins = dict(leftMargin=28, rightMargin=28, topMargin=34, bottomMargin=24)

    try:
        # V001: 투명 가장자리 평탄화본 — 로고 위 회색 바 방지
        logo = ImageReader(io.BytesIO(flat_logo_bytes()))
    except Exception:
        logo = None

    def on_page(canvas, doc):
        # 우측 상단 캐스팅라인 로고
        try:
            if logo is not None:
                iw, ih = logo.getSize()
                h = 11 * mm
                w = iw * h / ih
                canvas.drawImage(logo, page_size[0] - 28 - w, page_size[1] - 30 - 2,
                                 width=w, height=h)
        except Exception:
            pass

    doc = BaseDocTemplate(buf, pagesize=page_size, title=title, **margins)
    frame = Frame(margins['leftMargin'], margins['bottomMargin'],
                  page_size[0] - margins['leftMargin'] - margins['rightMargin'],
                  page_size[1] - margins['topMargin'] - margins['bottomMargin'])
    doc.addPageTemplates([PageTemplate(id='page', frames=[frame], onPage=on_page)])
    content_w = page_size[0] - margins['leftMargin'] - margins['rightMargin']
    return doc, content_w


_FONT_REGISTERED = False
PDF_FONT = "KR"
PDF_FONT_B = "KR-B"


def _ensure_pdf_fonts():
    global _FONT_REGISTERED
    if _FONT_REGISTERED:
        return
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    regular, bold = find_korean_font()
    pdfmetrics.registerFont(TTFont(PDF_FONT, regular))
    pdfmetrics.registerFont(TTFont(PDF_FONT_B, bold))
    _FONT_REGISTERED = True


def _pp(txt, size=8, bold=False, color=None, align="CENTER"):
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.platypus import Paragraph
    from reportlab.lib import colors
    st = ParagraphStyle(
        "c", fontName=PDF_FONT_B if bold else PDF_FONT, fontSize=size,
        leading=size + 3, textColor=color or colors.black,
        alignment={"LEFT": 0, "CENTER": 1, "RIGHT": 2}[align])
    return Paragraph(str(txt), st)


def _pdf_table(rows, col_widths, header_rows=1, total_rows=()):
    from reportlab.platypus import Table, TableStyle
    from reportlab.lib import colors
    HEAD = colors.HexColor("#D9E1F2")
    TOTAL = colors.HexColor("#FFE699")
    GRID = colors.HexColor("#BFBFBF")
    st = [
        ("GRID", (0, 0), (-1, -1), 0.5, GRID),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ("BACKGROUND", (0, 0), (-1, header_rows - 1), HEAD),
    ]
    for tr in total_rows:
        st.append(("BACKGROUND", (0, tr), (-1, tr), TOTAL))
    t = Table(rows, colWidths=col_widths, repeatRows=header_rows)
    t.setStyle(TableStyle(st))
    return t


def _fnum(n):
    return f"{int(round(n)):,}"


def _fpct(v):
    return f"{v:.1f}%"


# =====================================================================
# T001: 주요작 시간표 PDF (B002/0829 — 일자별 탭 = 페이지 1장)
# =====================================================================

def timetable_pdf_bytes(data):
    from reportlab.lib import colors
    from reportlab.platypus import PageBreak, Spacer

    _ensure_pdf_fonts()
    meta = data["meta"]
    RED = colors.HexColor("#D83A34")
    BLUE = colors.HexColor("#1D4ED8")

    def cmp_cell(txt, size=8):
        """증감 셀 — 증가는 빨강, 감소는 파랑 (화면과 같은 규칙)."""
        if txt == "-":
            return _pp("-", size=size)
        color = RED if txt.startswith("▲") else BLUE
        return _pp(txt, size=size, bold=True, color=color)

    buf = io.BytesIO()
    doc, content_w = _pdf_doc(buf, f"주요작 시간표 — {meta['movie_title']}")
    story = []

    for ti, tab in enumerate(data["tabs"]):
        if ti > 0:
            story.append(PageBreak())
        ks = tab["key_summary"]

        story.append(_pp(f"주요작 시간표 — {meta['movie_title']} · {tab['label']}",
                         size=15, bold=True, align="LEFT"))
        story.append(Spacer(1, 3))
        story.append(_pp(
            f"조사기간 {meta['date_from']} ~ {meta['date_to']}"
            f"  |  전일 {ks['prev_day']}  |  전주 {ks['prev_week']}"
            f"  |  개봉일 {meta.get('release_date') or '-'}"
            f"  |  배급사 {meta.get('distributor_name') or '-'}"
            f"  |  수집 완료 {meta.get('last_crawled_at') or '-'}",
            size=8.5, align="LEFT"))
        story.append(Spacer(1, 9))

        # ---- ① KEY SUMMARY ----
        story.append(_pp("KEY SUMMARY", size=11, bold=True, align="LEFT"))
        story.append(Spacer(1, 4))
        head = [_pp(h, bold=True) for h in
                ["구분", "총 좌석수", "예매좌석수", "좌석점유율",
                 "총 회차수", "총 극장수", "총 스크린수"]]
        rows = [head, [
            _pp(ks["label"], bold=True),
            _pp(_fnum(ks["total_seats"]) + "석", bold=True),
            _pp(_fnum(ks["sold_seats"]) + "석", bold=True),
            _pp(_fpct(ks["occupancy"]), bold=True),
            _pp(_fnum(ks["shows"]) + "회", bold=True),
            _pp(_fnum(ks["theaters"]) + "개", bold=True),
            _pp(_fnum(ks["screens"]) + "개", bold=True),
        ]]
        for label, key in (("증감 (전일比)", "prev_day_cmp"), ("증감 (전주比)", "prev_week_cmp")):
            c = tab["key_summary"].get(key)
            if c:
                cells = [cmp_cell(_cmp_kpi_txt(c["total_seats"], "석")),
                         cmp_cell(_cmp_kpi_txt(c["sold_seats"], "석")),
                         cmp_cell(_cmp_pp_txt(c["occupancy"])),
                         cmp_cell(_cmp_kpi_txt(c["shows"], "회")),
                         cmp_cell(_cmp_kpi_txt(c["theaters"], "개")),
                         cmp_cell(_cmp_kpi_txt(c["screens"], "개"))]
            else:
                cells = [_pp("-") for _ in range(6)]
            rows.append([_pp(label, bold=True)] + cells)
        w0 = content_w * 0.14
        wn = (content_w - w0) / 6
        story.append(_pdf_table(rows, [w0] + [wn] * 6, total_rows=[1]))
        story.append(Spacer(1, 11))

        # ---- ②③⑤ 멀티사별 / 포맷별 / 지역별 ----
        def detail_table(title, detail_rows, count_label):
            story.append(_pp(title, size=11, bold=True, align="LEFT"))
            story.append(Spacer(1, 4))
            rows = [[_pp(h, bold=True) for h in
                     ["구분", "총 좌석수", "비율", "전일比 좌석 증감",
                      "전주比 좌석 증감", count_label, "회차수"]]]
            for row in detail_rows:
                rows.append([
                    _pp(row["label"], bold=True),
                    _pp(_fnum(row["total_seats"]) + "석"),
                    _pp(_fpct(row["share"])),
                    cmp_cell(_cmp_seat_txt(row["prev_day_cmp"])),
                    cmp_cell(_cmp_seat_txt(row["prev_week_cmp"])),
                    _pp(_fnum(row["count"])),
                    _pp(_fnum(row["shows"])),
                ])
            w = content_w / 7
            story.append(_pdf_table(rows, [w * 1.3] + [w * 0.95] * 6))
            story.append(Spacer(1, 11))

        detail_table("멀티사별 상세 현황", tab["multi_detail"], "총 극장수")
        detail_table("포맷별 상세 현황", tab["format_detail"], "스크린수")

        # ---- ④ 시간대별 ----
        story.append(_pp("시간대별 상세 현황", size=11, bold=True, align="LEFT"))
        story.append(Spacer(1, 4))
        rows = [[_pp(h, bold=True) for h in
                 ["시간대 구분", "상영 회차수", "총 좌석수", "예매 좌석수",
                  "좌석 점유율", "전일比 점유율 차이", "전주比 점유율 차이"]]]
        for row in tab["time_detail"]:
            rows.append([
                _pp(row["label"], bold=True),
                _pp(_fnum(row["shows"]) + "회"),
                _pp(_fnum(row["total_seats"]) + "석"),
                _pp(_fnum(row["sold_seats"]) + "석"),
                _pp(_fpct(row["occupancy"])),
                cmp_cell(_cmp_pp_txt(row["prev_day_cmp"])),
                cmp_cell(_cmp_pp_txt(row["prev_week_cmp"])),
            ])
        w = content_w / 7
        story.append(_pdf_table(rows, [w * 1.6] + [w * 0.9] * 6))
        story.append(Spacer(1, 11))

        detail_table("지역별 상세 현황", tab["region_detail"], "총 극장수")

        # ---- ⑥ 주요작 vs 경쟁작 ----
        story.append(_pp(f"주요작 vs 경쟁작 — 동시 상영 경쟁작 TOP 10 순위 ({tab['label']})",
                         size=11, bold=True, align="LEFT"))
        story.append(Spacer(1, 4))
        rows = [[_pp(h, bold=True) for h in
                 ["순위", "영화명", "총 좌석수", "좌석 점유율", "상영 회차수",
                  "전일比 순위변동", "전주比 순위변동"]]]
        main_rows = []
        for i, row in enumerate(tab["competitor_top"], 1):
            bold = row["is_main"]
            rows.append([
                _pp(f"{row['rank']} (★당사)" if bold else row["rank"], bold=bold),
                _pp(row["title"], bold=bold, align="LEFT"),
                _pp(_fnum(row["total_seats"]) + "석", bold=bold),
                _pp(_fpct(row["occupancy"]), bold=bold),
                _pp(_fnum(row["shows"]) + "회", bold=bold),
                cmp_cell(_move_txt(row["prev_day_move"])),
                cmp_cell(_move_txt(row["prev_week_move"])),
            ])
            if bold:
                main_rows.append(i)
        w = content_w / 7
        story.append(_pdf_table(rows, [w * 0.75, w * 1.8] + [w * 0.89] * 5,
                                total_rows=main_rows))

    doc.build(story)
    return buf.getvalue()


# =====================================================================
# T003: 경쟁작 PDF
# =====================================================================

def competitor_pdf_bytes(data):
    from reportlab.platypus import PageBreak, Spacer

    _ensure_pdf_fonts()
    meta = data["meta"]

    buf = io.BytesIO()
    doc, content_w = _pdf_doc(buf, "경쟁작 성과 종합")
    story = []

    for ti, tab in enumerate(data["tabs"]):
        if ti > 0:
            story.append(PageBreak())
        story.append(_pp(f"경쟁작 성과 종합 — {tab['label']}", size=15, bold=True, align="LEFT"))
        story.append(Spacer(1, 3))
        story.append(_pp(
            f"조사기간 {meta['date_from']} ~ {meta['date_to']}  |  경쟁작 {meta['movie_count']}편"
            f"  |  수집 완료 {meta.get('last_crawled_at') or '-'}", size=8.5, align="LEFT"))
        story.append(Spacer(1, 10))

        # ① 종합 요약
        story.append(_pp("경쟁작 성과 종합 요약", size=11, bold=True, align="LEFT"))
        story.append(Spacer(1, 4))
        rows = [[_pp(h, bold=True) for h in
                 ["순위", "영화명", "총 좌석수", "평균 좌석점유율", "총 상영회차", "스크린수", "총 극장수"]]]
        total_rows = []
        for i, s in enumerate(tab["summary"]):
            bold = i == 0
            rows.append([_pp(s["rank"], bold=bold),
                         _pp(("★ " if bold else "") + s["title"], bold=bold, align="LEFT"),
                         _pp(_fnum(s["total_seats"]) + "석", bold=bold),
                         _pp(_fpct(s["occupancy"]), bold=bold),
                         _pp(_fnum(s["shows"]) + "회", bold=bold),
                         _pp(_fnum(s["screens"]) + "개", bold=bold),
                         _pp(_fnum(s["theaters"]) + "개", bold=bold)])
            if bold:
                total_rows.append(1)
        w = content_w
        story.append(_pdf_table(rows, [w * 0.06, w * 0.28, w * 0.14, w * 0.14, w * 0.13, w * 0.12, w * 0.13],
                                total_rows=total_rows))
        story.append(Spacer(1, 10))

        # ② 권역별
        story.append(_pp("서울 및 주요 권역별 좌석 점유 현황", size=11, bold=True, align="LEFT"))
        story.append(Spacer(1, 4))
        rows = [[_pp(h, bold=True) for h in
                 ["영화명", "서울 좌석수", "서울 좌점율", "서울 비중",
                  "수도권(경강) 좌석수", "수도권 좌점율", "그 외 지방 좌석수", "지방 좌점율"]]]
        for row in tab["regions"]:
            rows.append([_pp(row["title"], align="LEFT"),
                         _pp(_fnum(row["seoul"]["seats"]) + "석"), _pp(_fpct(row["seoul"]["occupancy"])),
                         _pp(_fpct(row["seoul"]["share"])),
                         _pp(_fnum(row["metro"]["seats"]) + "석"), _pp(_fpct(row["metro"]["occupancy"])),
                         _pp(_fnum(row["local"]["seats"]) + "석"), _pp(_fpct(row["local"]["occupancy"]))])
        wn = (content_w - content_w * 0.2) / 7
        story.append(_pdf_table(rows, [content_w * 0.2] + [wn] * 7))
        story.append(Spacer(1, 10))

        # ③ 골든타임 + ④ 특별관 (나란히 대신 순차)
        story.append(_pp("골든타임(14~21시) 집중도", size=11, bold=True, align="LEFT"))
        story.append(Spacer(1, 4))
        rows = [[_pp(h, bold=True) for h in ["영화명", "골든타임 좌석수", "골든타임 점유율", "골든타임 비중"]]]
        for row in tab["golden"]:
            rows.append([_pp(row["title"], align="LEFT"), _pp(_fnum(row["seats"]) + "석"),
                         _pp(_fpct(row["occupancy"])), _pp(_fpct(row["share"]))])
        wn = (content_w - content_w * 0.3) / 3
        story.append(_pdf_table(rows, [content_w * 0.3] + [wn] * 3))
        story.append(Spacer(1, 10))

        # ④ 특별관 — V004(0831): 타입별(IMAX/4DX/SCREENX/Dolby) 개별 표
        wn = (content_w - content_w * 0.3) / 2
        if tab["special"]:
            for blk in tab["special"]:
                story.append(_pp(f"특별관 ({blk['format']})", size=11, bold=True, align="LEFT"))
                story.append(Spacer(1, 4))
                rows = [[_pp(h, bold=True) for h in ["영화명", "특별관 회차", "특별관 좌석수"]]]
                for row in blk["rows"]:
                    rows.append([_pp(row["title"], align="LEFT"),
                                 _pp(_fnum(row["shows"]) + "회"), _pp(_fnum(row["seats"]) + "석")])
                story.append(_pdf_table(rows, [content_w * 0.3] + [wn] * 2))
                story.append(Spacer(1, 10))
        else:
            story.append(_pp("특별관 (IMAX/4DX/SCREENX/Dolby)", size=11, bold=True, align="LEFT"))
            story.append(Spacer(1, 4))
            rows = [[_pp(h, bold=True) for h in ["영화명", "특별관 회차", "특별관 좌석수"]],
                    [_pp("특별관 상영 데이터 없음", align="LEFT"), _pp("-"), _pp("-")]]
            story.append(_pdf_table(rows, [content_w * 0.3] + [wn] * 2))
            story.append(Spacer(1, 10))

        # ⑤ 계열사별 세부 현황 — 작품이 많으면 8개씩 끊어 표를 나눈다
        bb = tab["by_brand"]
        if bb["movies"]:
            story.append(_pp("계열사별 세부 현황", size=11, bold=True, align="LEFT"))
            story.append(Spacer(1, 4))
            CHUNK = 8
            for start in range(0, len(bb["movies"]), CHUNK):
                movies = bb["movies"][start:start + CHUNK]
                rows = [[_pp("구분", bold=True)] + [_pp(m, bold=True) for m in movies]]
                for row in bb["rows"]:
                    cells = [_pp(row["brand"], bold=True)]
                    for c in row["cells"][start:start + CHUNK]:
                        cells.append(_pp(f"{_fnum(c['seats'])}석 ({c['share']:.1f}%)"))
                    rows.append(cells)
                w0 = content_w * 0.1
                wn = (content_w - w0) / max(len(movies), 1)
                story.append(_pdf_table(rows, [w0] + [wn] * len(movies)))
                story.append(Spacer(1, 6))

    doc.build(story)
    return buf.getvalue()
