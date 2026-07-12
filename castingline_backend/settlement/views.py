import calendar
import os
import openpyxl
from django.conf import settings
from decimal import Decimal, ROUND_HALF_UP
from datetime import datetime
from django.db.models import F, Q, Value
from django.db.models.functions import Replace
from django.http import HttpResponse
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from openpyxl.styles import Font, PatternFill, Alignment  # ✅ 스타일 관련 임포트

# 모델 및 유틸 임포트
from score.models import Score
from fund.models import DailyFund, MonthlyFund, Fund
from rate.models import Rate, TheaterRate, DefaultRate
from client.models import Client, Theater
from movie.models import Movie
from castingline_backend.utils.excel_helper import ExcelGenerator

# 1. 정산용 영화 목록 조회


class SettlementMovieListView(APIView):
    """선택된 연월에 상영 실적이 있는 영화의 '대표 영화' 목록 (가나다->ABC->숫자 순)"""

    def get(self, request):
        yyyy_mm = request.query_params.get("yyyyMm")
        if not yyyy_mm:
            return Response([])

        try:
            yyyy, mm = yyyy_mm.split("-")
            # 1. 해당 월에 실적이 있는 모든 영화 ID 조회
            active_movie_ids = (
                Score.objects.filter(entry_date__year=yyyy,
                                     entry_date__month=mm)
                .values_list("movie", flat=True)
                .distinct()
            )

            # 2. 해당 영화들 조회
            active_movies = Movie.objects.filter(id__in=active_movie_ids)
            # 3. 대표 영화들 식별
            primary_movie_codes = set()
            for m in active_movies:
                raw_code = ""
                if m.is_primary_movie:
                    raw_code = m.movie_code
                elif m.primary_movie_code:
                    raw_code = m.primary_movie_code
                else:
                    raw_code = m.movie_code
                
                if raw_code:
                    primary_movie_codes.add(raw_code.replace(" ", ""))
            
            # 4. 최종 대표 영화 목록 (DB에서 정보 가져오기)
            # DB의 movie_code에서도 공백을 제거한 뒤 비교 (Replace 사용)
            primary_movies = list(Movie.objects.annotate(
                clean_code=Replace(F('movie_code'), Value(' '), Value(''))
            ).filter(
                clean_code__in=primary_movie_codes, is_primary_movie=True
            ).values("id", "title_ko"))
            # 만약 primary_movie 정보가 부족하면(데이터 불일치 등), active_movies 중 is_primary_movie인 것들이라도 포함
            if not primary_movies:
                primary_movies = list(active_movies.filter(
                    is_primary_movie=True).values("id", "title_ko"))

            def sort_key(movie):
                title = movie["title_ko"] or ""
                if not title:
                    return (3, "")
                first_char = title[0]
                if "가" <= first_char <= "힣":
                    return (0, title)
                if "a" <= first_char.lower() <= "z":
                    return (1, title)
                if "0" <= first_char <= "9":
                    return (2, title)
                return (3, title)

            primary_movies.sort(key=sort_key)
            return Response([{"id": m["id"], "title": m["title_ko"]} for m in primary_movies])
        except Exception:
            return Response([])

# 2. 월간 부금 정산 관리 (메인 로직)


class SettlementListView(APIView):
    def get_processed_data(self, yyyy_mm, movie_id, target_filter, client_id=None,
                           include_adjustments=True, show_adjustment_info=False):
        """API와 엑셀에서 공통으로 사용하는 핵심 계산 및 집계 로직.

        include_adjustments=False 는 부금 대사(compare) 전용 — 대사는 조정을
        자체 적용하므로 이중 반영을 막는다.
        show_adjustment_info: 조정액/'(수동조정)' 표기 노출 여부 — 관리자 계정만 True.
        (조정이 반영된 최종 금액은 노출 여부와 무관하게 동일)
        """
        yyyy, mm = map(int, yyyy_mm.split("-"))

        # [대표영화 합산 로직 추가]
        try:
            primary_movie = Movie.objects.get(id=movie_id)
            clean_parent_code = (primary_movie.movie_code or "").replace(" ", "")

            # 본인(대표영화) + 본인을 부모로 둔 모든 하위(프린트) 영화 ID들 추출
            # DB의 primary_movie_code에서도 공백을 제거한 뒤 비교
            all_related_movie_ids = list(Movie.objects.annotate(
                clean_primary_code=Replace(F('primary_movie_code'), Value(' '), Value(''))
            ).filter(
                Q(id=movie_id) | Q(clean_primary_code=clean_parent_code)
            ).values_list("id", flat=True))
        except Movie.DoesNotExist:
            return []

        # 스코어 데이터 조회 (합산된 영화 ID들 사용)
        scores = Score.objects.filter(
            entry_date__year=yyyy, entry_date__month=mm, movie_id__in=all_related_movie_ids
        ).select_related("client", "movie").order_by("entry_date")

        if client_id:
            scores = scores.filter(client_id=client_id)

        if not scores.exists():
            return []

        # 캐싱 데이터 준비
        client_ids = list(scores.values_list(
            "client_id", flat=True).distinct())
        daily_fund_map = {(f.client_id, f.dd): f.fund_yn for f in DailyFund.objects.filter(
            client_id__in=client_ids, yyyy=yyyy, mm=mm)}
        monthly_fund_map = {(f.client_id, f.mm): f.fund_yn for f in MonthlyFund.objects.filter(
            client_id__in=client_ids, yyyy=yyyy)}
        yearly_fund_map = {(f.client_id, f.yyyy): f.fund_yn for f in Fund.objects.filter(
            client_id__in=client_ids, yyyy=yyyy)}

        # 부율(Rate)은 실제 상영 포맷(하위영화)별로 설정된 것을 따른다.
        # (대표영화 폴백 없음 — 행의 상영타입에 해당하는 포맷 부율을 그대로 사용)
        rates = Rate.objects.filter(movie_id__in=all_related_movie_ids, client_id__in=client_ids).filter(
            Q(start_date__year=yyyy, start_date__month=mm) | Q(end_date__year=yyyy, end_date__month=mm) |
            Q(start_date__lte=datetime(yyyy, mm, 1), end_date__gte=datetime(
                yyyy, mm, calendar.monthrange(yyyy, mm)[1]))
        )
        rate_map = {}
        for r in rates:
            rate_map.setdefault((r.client_id, r.movie_id), []).append(r)

        theater_rate_map = {(tr.rate_id, tr.theater.auditorium_name): tr.share_rate for tr in TheaterRate.objects.filter(
            rate__in=rates).select_related("theater")}
        default_rate_map = {(dr.region_code, dr.theater_kind): dr.share_rate for dr in DefaultRate.objects.all()}

        # 데이터 집계
        aggregated_data = {}
        for score in scores:
            client = score.client
            c_id = client.id
            entry_date = score.entry_date

            # 부율 조회는 해당 스코어의 실제 하위영화(상영 포맷) 기준
            share_rate = self._get_cached_rate(
                c_id, score.movie_id, entry_date, score.auditorium, rate_map, theater_rate_map, default_rate_map, client)
            is_fund_exempt = self._get_cached_fund(
                c_id, entry_date, daily_fund_map, monthly_fund_map, yearly_fund_map)

            # 하위영화별 상영타입을 구분하여 그룹핑 (필름/디지털, 자막/더빙, 2D/3D 등 모든 타입 조합)
            screening_type = self._get_screening_type(score.movie)
            group_key = (c_id, share_rate, is_fund_exempt, screening_type)
            if group_key not in aggregated_data:
                aggregated_data[group_key] = self.init_data_struct(
                    client, score.movie, share_rate, is_fund_exempt, entry_date)

            target = aggregated_data[group_key]
            visitor_count = int(score.visitor or 0)
            fare = Decimal(str(score.fare or 0))

            unit_excl_fund = fare if is_fund_exempt else (
                fare / Decimal("1.03")).quantize(Decimal("1"), rounding=ROUND_HALF_UP)

            target["인원"] += visitor_count
            target["_total_raw_amt"] += int(fare) * visitor_count
            target["_total_excl_fund_sum"] += unit_excl_fund * visitor_count
            fare_key = int(fare)
            target["_fare_excl_fund"][fare_key] = (
                target["_fare_excl_fund"].get(fare_key, Decimal("0")) + unit_excl_fund * visitor_count)

            if entry_date < target["_min_date"]:
                target["_min_date"] = entry_date
            if entry_date > target["_max_date"]:
                target["_max_date"] = entry_date

        # 결과 리스트 생성 및 필터링
        calculated_list = []
        for data in aggregated_data.values():
            if target_filter == "일반극장" and data["is_fund_exempt"]:
                continue
            if target_filter == "기금면제극장" and not data["is_fund_exempt"]:
                continue

            item = self.calculate_amounts(data)
            item["날짜(From)"] = data["_min_date"].strftime("%Y-%m-%d")
            item["날짜(To)"] = data["_max_date"].strftime("%Y-%m-%d")
            calculated_list.append(item)

        # 수동 조정 반영 — 별도 행 없이 해당 극장 행(지급금이 가장 큰 행)에 차액을
        # 합산해 한 줄로 표시한다. 상영타입에 '(수동조정)' 표기로 구분.
        # (부금 대사 원 단위 잔차 확정용. 원래값은 대사 화면 툴팁에서 확인 가능)
        if include_adjustments:
            from settlement.models import SettlementAdjustment
            adjustments = SettlementAdjustment.objects.filter(
                yyyymm=yyyy_mm, movie_id=movie_id).select_related("client")
            if client_id:
                adjustments = adjustments.filter(client_id=client_id)
            for adj in adjustments:
                c = adj.client
                candidates = [it for it in calculated_list
                              if it.get("거래처코드") == c.client_code
                              and it.get("영화사 지급금") is not None]
                if candidates:
                    tgt = max(candidates, key=lambda it: it["영화사 지급금"])
                    tgt["공급가액"] += adj.supply_delta
                    tgt["부가세"] += adj.vat_delta
                    tgt["영화사 지급금"] += adj.payout_delta
                    if show_adjustment_info:  # 조정 사실/금액은 관리자에게만 표시
                        tgt["상영타입"] = f"{tgt['상영타입']} (수동조정)".strip()
                        tgt["is_adjusted"] = True
                        # 프론트에서 조정액 표시(다른 색) + 테이블에서 바로 해제용 ID
                        tgt["조정액"] = {"공급가액": adj.supply_delta, "부가세": adj.vat_delta,
                                      "영화사 지급금": adj.payout_delta}
                        tgt["조정ID"] = adj.id
                else:
                    # 해당 극장 계산 행이 없으면(스코어 삭제 등) 조정만 별도 행으로 표시
                    calculated_list.append({
                        "지역": c.region_code, "멀티구분": c.theater_kind,
                        "거래처코드": c.client_code,
                        "거래처코드(바이포엠만 해당)": c.by4m_theater_code or "-",
                        "극장명": c.client_name or "-",
                        "사업자 등록번호": c.business_registration_number,
                        "종사업장번호": c.business_operator,
                        "공급받는자 상호": c.business_name,
                        "공급받는자 성명": c.representative_name,
                        "사업장 소재": c.business_address, "업태": c.business_category,
                        "업종": c.business_industry,
                        "수신자이메일": c.invoice_email_address,
                        "수신자이메일2": c.invoice_email_address2,
                        "수신자 전화번호": c.settlement_phone_number,
                        "상영타입": "수동조정" if show_adjustment_info else "-",
                        "인원": 0, "부율": None,
                        "is_fund_exempt": False,
                        "classification": c.classification,
                        "금액(입장료)": 0, "기금제외금액": 0, "부가세제외금액": 0,
                        "공급가액": adj.supply_delta, "부가세": adj.vat_delta,
                        "영화사 지급금": adj.payout_delta,
                        "날짜(From)": "", "날짜(To)": "",
                        "is_adjustment": True,
                        "조정ID": adj.id if show_adjustment_info else None,
                    })

        return self.sort_and_add_subtotals(calculated_list)

    def get(self, request):
        yyyy_mm = request.query_params.get("yyyyMm")
        movie_id = request.query_params.get("movie_id")
        target_filter = request.query_params.get("target", "전체극장")
        client_id = request.query_params.get("client_id") or None

        if not yyyy_mm or not movie_id:
            return Response({"error": "년월과 영화를 선택해주세요."}, status=400)

        results = self.get_processed_data(
            yyyy_mm, movie_id, target_filter, client_id=client_id,
            show_adjustment_info=bool(getattr(request.user, "is_superuser", False)))
        return Response(results)

    # 헬퍼 함수들
    def _get_cached_fund(self, c_id, date, d_map, m_map, y_map):
        val = d_map.get((c_id, date.day))
        if val is not None:
            return val
        val = m_map.get((c_id, date.month))
        if val is not None:
            return val
        val = y_map.get((c_id, date.year))
        return val if val is not None else False

    def _get_cached_rate(self, c_id, m_id, date, aud_name, r_map, tr_map, dr_map, client):
        # 해당 포맷(하위영화)에 설정된 부율만 사용 (상영관 예외 우선)
        for r in r_map.get((c_id, m_id), []):
            if r.start_date <= date <= r.end_date:
                tr_val = tr_map.get((r.id, aud_name))
                return tr_val if tr_val is not None else r.share_rate
        # 포맷 부율이 없으면 기본부율(DefaultRate)만, 그것도 없으면 None → 화면 빈칸
        # (대표영화 폴백·하드코딩 50 제거)
        return dr_map.get((client.region_code, client.theater_kind))

    def _get_screening_type(self, movie):
        """영화의 모든 상영 타입 정보를 조합하여 반환 (필름/디지털, 자막/더빙, 2D/3D, 4DX 등)"""
        parts = [
            movie.media_type,           # 필름/디지털
            movie.audio_mode,           # 자막/더빙
            movie.viewing_dimension,    # 2D/3D
            movie.screening_type,       # 일반/특별
            movie.dx4_viewing_dimension,# 4DX 상영 차원
            movie.imax_l,               # IMAX-L
            movie.screen_x,             # Screen X
        ]
        return " ".join([p for p in parts if p]).strip()

    def init_data_struct(self, client, movie, rate, exempt, entry_date):
        return {
            "지역": client.region_code, "멀티구분": client.theater_kind, "거래처코드": client.client_code,
            "거래처코드(바이포엠만 해당)": client.by4m_theater_code or "-",
            "극장명": client.client_name or "-",
            "사업자 등록번호": client.business_registration_number, "종사업장번호": client.business_operator,
            "공급받는자 상호": client.business_name, "공급받는자 성명": client.representative_name,
            "사업장 소재": client.business_address, "업태": client.business_category, "업종": client.business_industry,
            "수신자이메일": client.invoice_email_address,
            "수신자이메일2": client.invoice_email_address2,
            "수신자 전화번호": client.settlement_phone_number,
            "상영타입": self._get_screening_type(movie), "인원": 0,
            "부율": float(rate) if rate is not None else None, "is_fund_exempt": exempt,
            "_total_raw_amt": 0, "_total_excl_fund_sum": Decimal("0"), "_min_date": entry_date, "_max_date": entry_date,
            "_fare_excl_fund": {},  # 가격대(요금)별 기금제외금액 합 — 메가박스 부가세 계산용
            "classification": client.classification,
        }

    def calculate_amounts(self, data):
        fare_excl_fund = data.pop("_fare_excl_fund", {})  # 가격대별 기금제외금액 합

        # [Special Logic] Indie Plus (인디플러스포항/천안)
        theater_name = data.get("극장명", "").replace(" ", "")
        fixed_rate_per_person = 0

        if "인디플러스포항" in theater_name:
            fixed_rate_per_person = 2500
        elif "인디플러스천안" in theater_name:
            fixed_rate_per_person = 3500

        if fixed_rate_per_person > 0:
            visitors = data["인원"]
            total_payment = int(visitors * fixed_rate_per_person)

            # 역산: 지급금 = 공급가액 * 1.1
            supply_val = (Decimal(total_payment) / Decimal("1.1")
                          ).quantize(Decimal("1"), rounding=ROUND_HALF_UP)
            vat = total_payment - int(supply_val)

            data.update({
                "금액(입장료)": data["_total_raw_amt"],
                "기금제외금액": data["_total_excl_fund_sum"],
                "부가세제외금액": (data["_total_excl_fund_sum"] / Decimal("1.1")).quantize(Decimal("1"), rounding=ROUND_HALF_UP),
                "공급가액": int(supply_val),
                "부가세": int(vat),
                "영화사 지급금": total_payment,
            })
            return data

        total_amt = data["_total_raw_amt"]
        rounded_excl_fund = data["_total_excl_fund_sum"]

        # 부율이 없으면(해당 포맷 부율 미설정) 부율에 의존하는 금액은 비움
        if data["부율"] is None:
            data.update({
                "금액(입장료)": total_amt, "기금제외금액": int(rounded_excl_fund),
                "부가세제외금액": None, "공급가액": None, "부가세": None, "영화사 지급금": None,
            })
            return data

        rate = Decimal(str(data["부율"]))

        rounded_excl_vat_total = (
            rounded_excl_fund / Decimal("1.1")).quantize(Decimal("1"), rounding=ROUND_HALF_UP)
        if (data.get("멀티구분") or "") in ("메가박스", "롯데") and fare_excl_fund:
            # 메가박스/롯데 정산서 방식(메가 1,758행 전수 + 롯데 극장×영화 85% 검증):
            # 가격대(티켓가)별로
            #   부금총액 = round(입장액 × 부율) → 공급가 = round(부금총액/1.1)
            #   → 부가세 = 부금총액 - 공급가
            # 로 역산한 뒤 합산한다. 극장 합계에서 한 번 반올림하는 기존 방식과
            # 1~3원 차이가 나므로 두 체인만 정산서와 동일한 방식을 적용.
            # (CGV는 회차보다 세분화된 내부 단위로 반올림해 재현 불가 → 기존 방식 유지)
            rounded_supply_val = Decimal("0")
            rounded_vat = Decimal("0")
            for f_sum in fare_excl_fund.values():
                payout_f = (f_sum * (rate / Decimal("100"))).quantize(Decimal("1"), rounding=ROUND_HALF_UP)
                supply_f = (payout_f / Decimal("1.1")).quantize(Decimal("1"), rounding=ROUND_HALF_UP)
                rounded_supply_val += supply_f
                rounded_vat += payout_f - supply_f
        else:
            rounded_supply_val = (rounded_excl_vat_total * (rate / Decimal("100"))
                                  ).quantize(Decimal("1"), rounding=ROUND_HALF_UP)
            rounded_vat = (rounded_supply_val * Decimal("0.1")
                           ).quantize(Decimal("1"), rounding=ROUND_HALF_UP)

        data.update({
            "금액(입장료)": total_amt, "기금제외금액": int(rounded_excl_fund), "부가세제외금액": int(rounded_excl_vat_total),
            "공급가액": int(rounded_supply_val), "부가세": int(rounded_vat), "영화사 지급금": int(rounded_supply_val + rounded_vat),
        })
        return data

    def sort_and_add_subtotals(self, items):
        brand_order = ["CGV", "롯데", "메가", "씨네큐"]
        region_order = ["서울", "경강", "충청", "호남", "경북", "경남"]

        def get_sort_key(x):
            brand_idx = 99
            for i, b in enumerate(brand_order):
                if b in (x["멀티구분"] or ""):
                    brand_idx = i
                    break
            class_idx = 0 if x["classification"] == "직영" else 1
            region_idx = region_order.index(
                x["지역"]) if x["지역"] in region_order else 99
            return (brand_idx, class_idx, region_idx, x["극장명"])

        items.sort(key=get_sort_key)
        if not items:
            return []

        processed = []

        def get_section(item):
            brand = "기타"
            for b in brand_order:
                if b in (item["멀티구분"] or ""):
                    brand = b
                    break
            return (brand, item["classification"])

        current_section = get_section(items[0])
        section_sum = {
            "인원": 0, "금액(입장료)": 0, "기금제외금액": 0,
            "부가세제외금액": 0, "공급가액": 0, "부가세": 0, "영화사 지급금": 0
        }

        for item in items:
            section_key = get_section(item)
            if section_key != current_section:
                processed.append(self.make_subtotal_row(
                    current_section, section_sum))
                current_section = section_key
                section_sum = {
                    "인원": 0, "금액(입장료)": 0, "기금제외금액": 0,
                    "부가세제외금액": 0, "공급가액": 0, "부가세": 0, "영화사 지급금": 0
                }

            processed.append(item)
            section_sum["인원"] += item.get("인원") or 0
            section_sum["금액(입장료)"] += item.get("금액(입장료)") or 0
            section_sum["기금제외금액"] += item.get("기금제외금액") or 0
            section_sum["부가세제외금액"] += item.get("부가세제외금액") or 0
            section_sum["공급가액"] += item.get("공급가액") or 0
            section_sum["부가세"] += item.get("부가세") or 0
            section_sum["영화사 지급금"] += item.get("영화사 지급금") or 0

        processed.append(self.make_subtotal_row(current_section, section_sum))
        return processed

    def make_subtotal_row(self, section_key, sums):
        brand, clss = section_key
        return {
            "is_subtotal": True, "극장명": f"[{brand} {clss}] 합계",
            "인원": sums["인원"], "금액(입장료)": sums["금액(입장료)"],
            "기금제외금액": sums["기금제외금액"], "부가세제외금액": sums["부가세제외금액"],
            "공급가액": sums["공급가액"], "부가세": sums["부가세"], "영화사 지급금": sums["영화사 지급금"],
            "지역": "", "멀티구분": "", "거래처코드": "", "거래처코드(바이포엠만 해당)": "", "사업자 등록번호": "", "종사업장번호": "", "공급받는자 상호": "",
            "공급받는자 성명": "", "사업장 소재": "", "업태": "", "업종": "", "수신자이메일": "", "수신자이메일2": "", "수신자 전화번호": "",
            "날짜(From)": "", "날짜(To)": "", "상영타입": "", "부율": "",
            "classification": "",
        }

# 2-1. 직영 부금정산서 엑셀 대사(비교)


class SettlementCompareView(SettlementListView):
    """직영 체인(CGV/롯데/메가박스) 부금정산서 엑셀을 업로드받아
    파일 안의 '모든 영화'를 자동 매칭해 화면 정산 데이터
    (인원/공급가액/부가세/영화사지급금)와 영화×극장별로 한 번에 비교한다.

    POST (multipart):
      files  : 체인 부금정산서 엑셀 (.xlsx) — 여러 개 동시 업로드 가능,
               파일별로 양식/체인 자동 감지 후 전체 합산 비교 (구버전 'file' 단일도 지원)
      yyyyMm : 부금년월 (YYYY-MM, 선택) — 파일의 상영일에서 자동 인식하며,
               인식 실패 시에만 이 값을 폴백으로 사용
    """

    METRICS = ["인원", "공급가액", "부가세", "영화사 지급금"]

    def post(self, request):
        import re as _re
        from collections import Counter

        from settlement.compare import parse_settlement_excel, norm_theater, norm_title

        files = request.FILES.getlist("files")
        if not files and request.FILES.get("file"):
            files = [request.FILES["file"]]
        req_yyyy_mm = (request.data.get("yyyyMm") or "").strip()
        if not files:
            return Response({"error": "파일을 지정해주세요."}, status=400)

        # 1. 파일별 엑셀 파싱 (체인 자동 감지) 후 행 합치기 (행마다 체인 태깅)
        file_infos = []
        all_rows = []
        for idx, f in enumerate(files):
            try:
                parsed = parse_settlement_excel(f)
            except Exception as e:
                return Response({"error": f"'{f.name}' 파싱 실패: {e}"}, status=400)
            file_infos.append({"filename": f.name, "chain": parsed["chain"],
                               "row_count": len(parsed["rows"])})
            for r in parsed["rows"]:
                r["chain"] = parsed["chain"]
                r["file_idx"] = idx  # 정산 회차(파일) 구분 — 메가박스 기간 분할 추정용
                all_rows.append(r)
        chains = sorted({fi["chain"] for fi in file_infos})

        # 1-1. 정산월 자동 인식: 파일 각 행의 상영일에서 최빈 연-월 추출
        #      (페이지에서 보고 있는 연월과 무관하게 파일 기준으로 대사)
        month_counter = Counter()
        for r in all_rows:
            m = _re.search(r"(20\d{2})[-./]?(\d{1,2})", str(r.get("date") or ""))
            if m:
                month_counter[f"{m.group(1)}-{int(m.group(2)):02d}"] += 1
        if month_counter:
            yyyy_mm = month_counter.most_common(1)[0][0]
            yyyy_mm_source = "file"
        elif req_yyyy_mm:
            yyyy_mm = req_yyyy_mm
            yyyy_mm_source = "request"
        else:
            return Response({"error": "파일에서 상영월을 인식하지 못했습니다. 부금년월을 지정해주세요."},
                            status=400)

        # 2. 해당 월에 실적이 있는 대표영화 목록 (정산 화면 영화 드롭다운과 동일 기준)
        yyyy, mm = yyyy_mm.split("-")
        active_movie_ids = (
            Score.objects.filter(entry_date__year=yyyy, entry_date__month=mm)
            .values_list("movie", flat=True).distinct()
        )
        active_movies = Movie.objects.filter(id__in=active_movie_ids)
        primary_codes = set()
        for m in active_movies:
            raw = m.movie_code if m.is_primary_movie else (m.primary_movie_code or m.movie_code)
            if raw:
                primary_codes.add(raw.replace(" ", ""))
        primary_movies = list(Movie.objects.annotate(
            clean_code=Replace(F('movie_code'), Value(' '), Value(''))
        ).filter(clean_code__in=primary_codes, is_primary_movie=True))
        primary_norms = [(p, norm_title(p.title_ko)) for p in primary_movies if p.title_ko]

        # 3. 파일의 영화명 → 대표영화 자동 매칭
        #    (양방향 포함 매칭, 겹치는 후보는 제목이 긴 쪽 = 더 구체적인 쪽 채택)
        def match_primary(file_movie_name):
            f_norm = norm_title(file_movie_name)
            best = None
            for p, p_norm in primary_norms:
                if not p_norm:
                    continue
                if p_norm in f_norm or f_norm in p_norm:
                    if best is None or len(p_norm) > len(best[1]):
                        best = (p, p_norm)
            return best[0] if best else None

        file_movie_names = {r["movie"] for r in all_rows}
        movie_match = {name: match_primary(name) for name in file_movie_names}

        # 대표영화별 파일 행 그룹 + 미매칭 영화 집계
        rows_by_primary = {}
        unmatched = {}
        for row in all_rows:
            primary = movie_match.get(row["movie"])
            if primary is None:
                agg = unmatched.setdefault(row["movie"], {
                    "movie": row["movie"], "인원": 0, "공급가액": 0, "부가세": 0, "영화사 지급금": 0,
                })
                agg["인원"] += row["visitors"]
                agg["공급가액"] += row["supply"]
                agg["부가세"] += row["vat"]
                agg["영화사 지급금"] += row["payout"]
                continue
            rows_by_primary.setdefault(primary.id, {"movie": primary, "rows": []})["rows"].append(row)

        # 4. 영화별 비교 실행
        movie_sections = []
        grand_totals = {m: {"system": 0, "file": 0} for m in self.METRICS}
        grand_summary = {"equal": 0, "diff": 0, "file_only": 0, "system_only": 0, "theater_count": 0}
        for primary_id, group in sorted(rows_by_primary.items(),
                                        key=lambda kv: kv[1]["movie"].title_ko or ""):
            section = self._compare_one_movie(
                yyyy_mm, group["movie"], group["rows"], chains, norm_theater,
                show_adjustment_info=bool(getattr(request.user, "is_superuser", False)))
            movie_sections.append(section)
            for m in self.METRICS:
                grand_totals[m]["system"] += section["totals"][m]["system"]
                grand_totals[m]["file"] += section["totals"][m]["file"]
            for k in ("equal", "diff", "file_only", "system_only", "theater_count"):
                grand_summary[k] += section["summary"][k]
        for m in self.METRICS:
            grand_totals[m]["diff"] = grand_totals[m]["file"] - grand_totals[m]["system"]
        grand_summary["movie_count"] = len(movie_sections)

        return Response({
            "chains": chains,
            "files": file_infos,
            "yyyyMm": yyyy_mm,
            "yyyyMm_source": yyyy_mm_source,
            "file_row_count": len(all_rows),
            "movies": movie_sections,
            "unmatched_file_movies": sorted(unmatched.values(), key=lambda x: x["movie"]),
            "grand_totals": grand_totals,
            "grand_summary": grand_summary,
        })

    def _compare_one_movie(self, yyyy_mm, primary_movie, file_rows, chains, norm_theater,
                           show_adjustment_info=False):
        """대표영화 하나에 대한 극장별 대사(업로드된 체인 전체 합산). 반환: 영화 섹션 dict.

        멀티 체인 동시 비교이므로 극장 키는 (체인, 정규화극장명) — 브랜드 접두사
        제거 후 같은 이름(예: CGV/롯데 '센텀시티')이 충돌하지 않게 한다.
        """
        import re as _re

        # 시스템측: 화면과 동일 계산(get_processed_data, 조정 미포함 — 아래에서 자체 적용)
        # → 업로드 체인+직영 필터 → 극장별 합산
        # (직영 부금정산서엔 직영 지점만 있으므로 위탁/기타 극장은 목록에서 제외 — 사용자 확정)
        system_rows = self.get_processed_data(yyyy_mm, primary_movie.id, "전체극장",
                                              include_adjustments=False)
        sys_by_theater = {}
        for row in system_rows:
            if row.get("is_subtotal") or row.get("is_adjustment"):
                continue  # 수동조정 행은 아래에서 별도 적용 (메가박스 재계산이 덮지 않도록)
            row_chain = row.get("멀티구분") or ""
            if row_chain not in chains:
                continue
            if row.get("classification") != "직영":
                continue
            key = (row_chain, norm_theater(row["극장명"]))
            agg = sys_by_theater.setdefault(key, {
                # 발전기금면제관 별도 거래처는 본 극장에 합산되므로 표시명에서 접미사 제거
                "name": row["극장명"].replace("(발전기금면제관)", ""),
                "chain": row_chain, "client_code": row.get("거래처코드"),
                "인원": 0, "공급가액": 0, "부가세": 0,
                "영화사 지급금": 0, "missing_rate": False,
            })
            agg["인원"] += row.get("인원") or 0
            for m in ("공급가액", "부가세", "영화사 지급금"):
                if row.get(m) is None:
                    agg["missing_rate"] = True  # 부율 미설정 포맷 존재
                else:
                    agg[m] += row[m]

        # 파일측: 극장별 합산 (씨네드쉐프 → 해당 CGV 지점 병합 규칙 적용)
        from settlement.compare import FILE_THEATER_MERGE

        file_by_theater = {}
        file_movie_names = set()
        for row in file_rows:
            file_movie_names.add(row["movie"])
            norm = norm_theater(row["theater"])
            norm = FILE_THEATER_MERGE.get(norm, norm)
            key = (row["chain"], norm)
            agg = file_by_theater.setdefault(key, {
                "name": row["theater"], "chain": row["chain"],
                "인원": 0, "공급가액": 0, "부가세": 0, "영화사 지급금": 0,
            })
            agg["인원"] += row["visitors"]
            agg["공급가액"] += row["supply"]
            agg["부가세"] += row["vat"]
            agg["영화사 지급금"] += row["payout"]

        # 미매칭 파일 극장 폴백: 같은 체인 안에서 괄호 부가정보 제거 코어로
        # 유일 매칭 시 병합 (예: 파일 '아산터미널(사용불가)' ↔ 시스템 '롯데아산터미널')
        def _core(k):
            return (k[0], _re.sub(r"\([^)]*\)", "", k[1]))

        sys_core_map = {}
        for k in sys_by_theater:
            sys_core_map.setdefault(_core(k), []).append(k)
        for fk in list(file_by_theater):
            if fk in sys_by_theater:
                continue
            candidates = [c for c in sys_core_map.get(_core(fk), [])
                          if c not in file_by_theater]
            if len(candidates) == 1:
                file_by_theater[candidates[0]] = file_by_theater.pop(fk)

        # 메가박스: 시스템 금액을 정산 회차(파일) 구간 단위로 재계산
        if "메가박스" in chains:
            mega_rows = [r for r in file_rows if r["chain"] == "메가박스"]
            self._apply_megabox_period_amounts(
                yyyy_mm, primary_movie, mega_rows, sys_by_theater, norm_theater)

        # 수동 조정 적용 (재계산 이후) — 조정값을 더하고 원래값을 함께 보존
        from settlement.models import SettlementAdjustment
        for adj in SettlementAdjustment.objects.filter(
                yyyymm=yyyy_mm, movie=primary_movie).select_related("client"):
            key = ((adj.client.theater_kind or ""), norm_theater(adj.client.client_name))
            agg = sys_by_theater.get(key)
            if not agg:
                continue
            if show_adjustment_info:  # 조정 내역(차액/원래값)은 관리자에게만 노출
                agg["adjustment"] = {
                    "id": adj.id,
                    "supply_delta": adj.supply_delta,
                    "vat_delta": adj.vat_delta,
                    "payout_delta": adj.payout_delta,
                    "note": adj.note,
                    "original": {"공급가액": agg["공급가액"], "부가세": agg["부가세"],
                                 "영화사 지급금": agg["영화사 지급금"]},
                }
            agg["공급가액"] += adj.supply_delta
            agg["부가세"] += adj.vat_delta
            agg["영화사 지급금"] += adj.payout_delta

        # 극장별 비교
        rows = []
        totals = {m: {"system": 0, "file": 0} for m in self.METRICS}
        for key in sorted(set(sys_by_theater) | set(file_by_theater),
                          key=lambda k: (k[0], (sys_by_theater.get(k) or file_by_theater.get(k))["name"])):
            sys_agg = sys_by_theater.get(key)
            file_agg = file_by_theater.get(key)
            status = "both" if sys_agg and file_agg else ("system_only" if sys_agg else "file_only")
            metrics = {}
            equal = status == "both"
            for m in self.METRICS:
                s_val = sys_agg[m] if sys_agg else None
                f_val = file_agg[m] if file_agg else None
                diff = (f_val or 0) - (s_val or 0)
                if status == "both" and diff != 0:
                    equal = False
                metrics[m] = {"system": s_val, "file": f_val, "diff": diff}
                totals[m]["system"] += s_val or 0
                totals[m]["file"] += f_val or 0
            rows.append({
                "체인": (sys_agg or file_agg)["chain"],
                "극장명": (sys_agg or file_agg)["name"],
                "파일극장명": file_agg["name"] if file_agg else None,
                "client_code": sys_agg.get("client_code") if sys_agg else None,
                "status": status,
                "equal": equal,
                "missing_rate": bool(sys_agg and sys_agg["missing_rate"]),
                "adjustment": sys_agg.get("adjustment") if sys_agg else None,
                "metrics": metrics,
            })

        # 불일치 → 파일에만 → 시스템에만 → 일치 순, 같은 상태끼리는 체인·극장명 순 정렬
        order = {"both": 0, "file_only": 1, "system_only": 2}
        chain_order = {"CGV": 0, "롯데": 1, "메가박스": 2}
        rows.sort(key=lambda r: (r["equal"], order.get(r["status"], 9),
                                 chain_order.get(r["체인"], 9), r["극장명"]))

        for m in self.METRICS:
            totals[m]["diff"] = totals[m]["file"] - totals[m]["system"]

        return {
            "movie_id": primary_movie.id,
            "movie_title": primary_movie.title_ko,
            "file_movie_names": sorted(file_movie_names),
            "rows": rows,
            "totals": totals,
            "summary": {
                "theater_count": len(rows),
                "equal": sum(1 for r in rows if r["equal"]),
                "diff": sum(1 for r in rows if not r["equal"] and r["status"] == "both"),
                "file_only": sum(1 for r in rows if r["status"] == "file_only"),
                "system_only": sum(1 for r in rows if r["status"] == "system_only"),
            },
        }

    def _apply_megabox_period_amounts(self, yyyy_mm, primary_movie, mega_file_rows,
                                      sys_by_theater, norm_theater):
        """메가박스 시스템 금액(공급가액/부가세/지급금)을 정산 회차(파일) 구간으로 재계산.

        메가박스는 정산 회차가 월 중간에 나뉠 수 있어(예: 6/1~14, 6/15~30) 같은 요금이
        회차(파일)별로 각각 반올림된다. 파일의 기간 컬럼은 실제 내용과 다르게 표기되므로
        (전량 20260601~말일로 찍힘) 신뢰하지 않고, **파일별 관객수 합 = 시스템 일별
        관객수의 연속 구간 합**이 되는 분할 경계를 자동 추정해 그 구간별로 반올림한다.
        경계를 못 찾으면 월 합산(기존 방식)으로 계산한다.
        (대사 전용 — 정산 화면 값은 월 단위 계산 그대로)
        """
        yyyy, mm = map(int, yyyy_mm.split("-"))
        clean_parent_code = (primary_movie.movie_code or "").replace(" ", "")
        movie_ids = list(Movie.objects.annotate(
            clean_primary_code=Replace(F('primary_movie_code'), Value(' '), Value(''))
        ).filter(
            Q(id=primary_movie.id) | Q(clean_primary_code=clean_parent_code)
        ).values_list("id", flat=True))

        scores = Score.objects.filter(
            entry_date__year=yyyy, entry_date__month=mm, movie_id__in=movie_ids,
            client__theater_kind="메가박스", client__classification="직영",
        ).select_related("client")
        if not scores.exists():
            return

        # 부율/기금 캐시 (get_processed_data 와 동일 기준)
        client_ids = list(scores.values_list("client_id", flat=True).distinct())
        daily_fund_map = {(f.client_id, f.dd): f.fund_yn for f in DailyFund.objects.filter(
            client_id__in=client_ids, yyyy=yyyy, mm=mm)}
        monthly_fund_map = {(f.client_id, f.mm): f.fund_yn for f in MonthlyFund.objects.filter(
            client_id__in=client_ids, yyyy=yyyy)}
        yearly_fund_map = {(f.client_id, f.yyyy): f.fund_yn for f in Fund.objects.filter(
            client_id__in=client_ids, yyyy=yyyy)}
        rates = Rate.objects.filter(movie_id__in=movie_ids, client_id__in=client_ids).filter(
            Q(start_date__year=yyyy, start_date__month=mm) | Q(end_date__year=yyyy, end_date__month=mm) |
            Q(start_date__lte=datetime(yyyy, mm, 1), end_date__gte=datetime(
                yyyy, mm, calendar.monthrange(yyyy, mm)[1]))
        )
        rate_map = {}
        for r in rates:
            rate_map.setdefault((r.client_id, r.movie_id), []).append(r)
        theater_rate_map = {(tr.rate_id, tr.theater.auditorium_name): tr.share_rate
                            for tr in TheaterRate.objects.filter(rate__in=rates).select_related("theater")}
        default_rate_map = {(dr.region_code, dr.theater_kind): dr.share_rate
                            for dr in DefaultRate.objects.all()}

        # (체인,극장) → {(요금, 부율): {날짜: 기금제외금액}} + 요금별 인원 + 시스템 일별 관객수
        per_theater = {}
        per_theater_n = {}
        daily_visitors = {}
        skip_keys = set()  # 부율 미설정 스코어가 있는 극장은 재계산하지 않음(기존 값 유지)
        for s in scores:
            key = ("메가박스", norm_theater(s.client.client_name))
            n = int(s.visitor or 0)
            daily_visitors[s.entry_date] = daily_visitors.get(s.entry_date, 0) + n
            rate = self._get_cached_rate(s.client_id, s.movie_id, s.entry_date, s.auditorium,
                                         rate_map, theater_rate_map, default_rate_map, s.client)
            if rate is None:
                skip_keys.add(key)
                continue
            exempt = self._get_cached_fund(s.client_id, s.entry_date,
                                           daily_fund_map, monthly_fund_map, yearly_fund_map)
            fare = Decimal(str(s.fare or 0))
            unit = fare if exempt else (fare / Decimal("1.03")).quantize(
                Decimal("1"), rounding=ROUND_HALF_UP)
            # 단가(부가단가)까지 키에 포함 — 발전기금면제관(단가=티켓가)과 일반관이
            # 한 극장으로 합산돼도 파일의 부가단가 기준으로 정확히 구분된다.
            fr_key = (int(fare), Decimal(str(rate)), int(unit))
            fmap = per_theater.setdefault(key, {}).setdefault(fr_key, {})
            fmap[s.entry_date] = fmap.get(s.entry_date, Decimal("0")) + unit * n
            nmap = per_theater_n.setdefault(key, {})
            nmap[fr_key] = nmap.get(fr_key, 0) + n

        # 파일 행별 (극장, 요금, 부가단가) → 인원 리스트 — 정산서 행 단위 반올림 재현용.
        # 회차 분할·무대인사·굿즈증정 등으로 같은 요금이 여러 행으로 쪼개져도
        # 단가가 같으면 입장액은 인원에 비례하므로, 행별 인원으로 정확히 재구성된다.
        file_fare_ns = {}
        for r in mega_file_rows:
            if not r.get("fare"):
                continue
            k = ("메가박스", norm_theater(r["theater"]))
            file_fare_ns.setdefault((k, int(r["fare"]), int(r.get("danga") or 0)),
                                    []).append(int(r["visitors"]))

        # 정산 회차(파일) 구간 추정: 파일별 관객수 합이 시스템 일별 관객수의
        # 연속 구간 합과 일치하는 분할을 찾는다. (파일 1개면 분할 없음)
        from itertools import permutations

        file_visitors = {}
        for r in mega_file_rows:
            file_visitors[r["file_idx"]] = file_visitors.get(r["file_idx"], 0) + r["visitors"]
        file_visitors = {k: v for k, v in file_visitors.items() if v > 0}

        segments = None  # [set(dates), ...] 회차별 날짜 집합
        if len(file_visitors) >= 2:
            days = sorted(daily_visitors)
            for perm in permutations(file_visitors):
                i, segs, ok = 0, [], True
                for fid in perm:
                    target, acc, cur = file_visitors[fid], 0, set()
                    while i < len(days) and acc < target:
                        acc += daily_visitors[days[i]]
                        cur.add(days[i])
                        i += 1
                    if acc != target:
                        ok = False
                        break
                    segs.append(cur)
                if ok and i == len(days):
                    segments = segs
                    break

        for key, fare_rate_map in per_theater.items():
            if key in skip_keys or key not in sys_by_theater:
                continue
            supply_tot = Decimal("0")
            vat_tot = Decimal("0")
            for (fare, rate, unit_key), datemap in fare_rate_map.items():
                # 1순위: 파일 행 단위 재구성 — (요금,부가단가)별 인원 합이 시스템과
                #        일치할 때만 적용
                sys_n = per_theater_n.get(key, {}).get((fare, rate, unit_key), 0)
                base_total = sum(datemap.values())
                file_ns = file_fare_ns.get((key, fare, unit_key))
                bases = None
                if (file_ns and sys_n and sum(file_ns) == sys_n
                        and int(base_total) % sys_n == 0):
                    unit = base_total / sys_n
                    bases = [unit * n for n in file_ns if n]
                if bases is None:
                    # 2순위: 정산 회차(파일) 구간 분할 / 3순위: 월 합산
                    if segments:
                        chunks = [[d for d in datemap if d in seg] for seg in segments]
                    else:
                        chunks = [list(datemap)]
                    bases = [sum((datemap[d] for d in chunk), Decimal("0"))
                             for chunk in chunks]
                for base in bases:
                    if not base:
                        continue
                    payout = (base * rate / Decimal("100")).quantize(
                        Decimal("1"), rounding=ROUND_HALF_UP)
                    supply = (payout / Decimal("1.1")).quantize(
                        Decimal("1"), rounding=ROUND_HALF_UP)
                    supply_tot += supply
                    vat_tot += payout - supply

            agg = sys_by_theater[key]
            agg["공급가액"] = int(supply_tot)
            agg["부가세"] = int(vat_tot)
            agg["영화사 지급금"] = int(supply_tot + vat_tot)


# 2-2. 부금 정산 수동 조정 (대사 잔차 확정)


def _serialize_adjustment(a):
    return {
        "id": a.id,
        "yyyyMm": a.yyyymm,
        "movie_id": a.movie_id,
        "client_id": a.client_id,
        "client_name": a.client.client_name,
        "supply_delta": a.supply_delta,
        "vat_delta": a.vat_delta,
        "payout_delta": a.payout_delta,
        "supply_original": a.supply_original,
        "vat_original": a.vat_original,
        "payout_original": a.payout_original,
        "note": a.note,
        "updated_at": a.updated_at,
    }


def _require_admin(request):
    """수동조정 조회/저장/삭제는 관리자(superuser) 전용."""
    if not bool(getattr(request.user, "is_superuser", False)):
        return Response({"error": "관리자만 사용할 수 있습니다."},
                        status=status.HTTP_403_FORBIDDEN)
    return None


class SettlementAdjustmentView(APIView):
    """부금 정산 수동 조정 목록/저장. (관리자 전용)
    GET  /Api/settlement-adjustments/?yyyyMm=&movie_id=   - 목록
    POST /Api/settlement-adjustments/                     - 저장(업서트)
      body: yyyyMm, movie_id, client_code, supply_delta, vat_delta, payout_delta,
            supply_original?, vat_original?, payout_original?, note?
    """

    def get(self, request):
        denied = _require_admin(request)
        if denied:
            return denied
        from settlement.models import SettlementAdjustment
        qs = SettlementAdjustment.objects.select_related("client")
        yyyy_mm = request.query_params.get("yyyyMm")
        movie_id = request.query_params.get("movie_id")
        if yyyy_mm:
            qs = qs.filter(yyyymm=yyyy_mm)
        if movie_id:
            qs = qs.filter(movie_id=movie_id)
        return Response([_serialize_adjustment(a) for a in qs])

    def post(self, request):
        denied = _require_admin(request)
        if denied:
            return denied
        from settlement.models import SettlementAdjustment
        data = request.data
        yyyy_mm = (data.get("yyyyMm") or "").strip()
        movie_id = data.get("movie_id")
        client_code = (str(data.get("client_code") or "")).strip()
        if not yyyy_mm or not movie_id or not client_code:
            return Response({"error": "yyyyMm/movie_id/client_code는 필수입니다."}, status=400)
        try:
            client = Client.objects.get(client_code=client_code)
        except Client.DoesNotExist:
            return Response({"error": f"거래처코드 {client_code} 극장을 찾을 수 없습니다."}, status=400)
        except Client.MultipleObjectsReturned:
            return Response({"error": f"거래처코드 {client_code}가 중복입니다."}, status=400)

        def _i(v):
            try:
                return int(v)
            except (TypeError, ValueError):
                return None

        obj, _created = SettlementAdjustment.objects.update_or_create(
            yyyymm=yyyy_mm, movie_id=movie_id, client=client,
            defaults={
                "supply_delta": _i(data.get("supply_delta")) or 0,
                "vat_delta": _i(data.get("vat_delta")) or 0,
                "payout_delta": _i(data.get("payout_delta")) or 0,
                "supply_original": _i(data.get("supply_original")),
                "vat_original": _i(data.get("vat_original")),
                "payout_original": _i(data.get("payout_original")),
                "note": (data.get("note") or "")[:200],
            },
        )
        return Response(_serialize_adjustment(obj), status=status.HTTP_201_CREATED)


class SettlementAdjustmentDetailView(APIView):
    """DELETE /Api/settlement-adjustments/<pk>/ - 조정 해제 (관리자 전용)"""

    def delete(self, request, pk):
        denied = _require_admin(request)
        if denied:
            return denied
        from settlement.models import SettlementAdjustment
        try:
            SettlementAdjustment.objects.get(pk=pk).delete()
        except SettlementAdjustment.DoesNotExist:
            return Response({"error": "Not found"}, status=status.HTTP_404_NOT_FOUND)
        return Response(status=status.HTTP_204_NO_CONTENT)


# 3. 월간 부금 정산 엑셀 출력 (SettlementListView 상속)


class SettlementExcelExportView(SettlementListView):
    def get(self, request):
        yyyy_mm = request.query_params.get("yyyyMm")
        movie_id = request.query_params.get("movie_id")
        target_filter = request.query_params.get("target", "전체극장")

        if not yyyy_mm or not movie_id:
            return HttpResponse("년월과 영화를 선택해주세요.", status=400)

        # 영화 제목 조회 (시트명 및 파일명 사용)
        try:
            movie = Movie.objects.get(id=movie_id)
            movie_title = movie.title_ko or "정산"
        except Movie.DoesNotExist:
            movie_title = "정산"

        # 상속받은 핵심 로직 호출 (엑셀의 '(수동조정)' 표기도 관리자만)
        items = self.get_processed_data(
            yyyy_mm, movie_id, target_filter,
            show_adjustment_info=bool(getattr(request.user, "is_superuser", False)))

        if not items:
            return HttpResponse("조회된 데이터가 없습니다.", status=404)

        # 시트명을 영화 제목으로 설정 (엑셀 시트명 최대 31자)
        sheet_name = movie_title[:31]
        excel = ExcelGenerator(sheet_name=sheet_name)
        header_labels = [
            "지역", "멀티", "구분", "거래처코드(바이포엠만 해당)", "극장명",
            "사업자 등록번호", "종사업장번호", "공급받는자 상호", "공급받는자 성명",
            "사업장 소재", "업태", "업종", "수신자이메일", "수신자 전화번호",
            "날짜(From)", "날짜(To)", "상영타입", "인원", "금액(입장료)",
            "기금제외금액", "부가세제외금액", "부율", "공급가액", "부가세", "영화사 지급금"
        ]
        excel.add_header(header_labels)

        # 틀 고정: 극장명(5번째 열) 이후 고정 → F2
        excel.ws.freeze_panes = "F2"

        data_rows = []
        subtotal_row_indices = []  # 합계 행 위치 추적 (1-based, 헤더 제외)

        for item in items:
            row = [
                item.get("지역", ""), item.get("멀티구분", ""), item.get("classification", ""),
                item.get("거래처코드(바이포엠만 해당)", ""), item.get("극장명", ""),
                item.get("사업자 등록번호", ""), item.get("종사업장번호", ""),
                item.get("공급받는자 상호", ""), item.get("공급받는자 성명", ""),
                item.get("사업장 소재", ""), item.get("업태", ""), item.get("업종", ""),
                item.get("수신자이메일", ""), item.get("수신자 전화번호", ""),
                item.get("날짜(From)", ""), item.get("날짜(To)", ""),
                item.get("상영타입", "") or "-",
                item.get("인원", 0), item.get("금액(입장료)", 0),
                item.get("기금제외금액", 0), item.get("부가세제외금액", 0), item.get("부율", 0),
                item.get("공급가액", 0), item.get("부가세", 0), item.get("영화사 지급금", 0),
            ]
            if item.get("is_subtotal"):
                subtotal_row_indices.append(len(data_rows))
            data_rows.append(row)

        excel.add_rows(data_rows)

        # 합계 행 스타일 추가 적용 (굵게 + 배경색), 헤더 행이 1행이므로 데이터는 2행~
        for idx in subtotal_row_indices:
            excel_row_num = idx + 2  # 헤더(1행) + 1-based index
            for cell in excel.ws[excel_row_num]:
                cell.font = Font(bold=True, size=10)
                cell.fill = PatternFill(
                    start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")

        filename = f"Settlement_{movie_title}_{yyyy_mm}_{datetime.now().strftime('%Y%m%d')}"
        return excel.to_response(filename)

# 4. 지정 부금 관리 (리스트 및 엑셀)


class SpecialSettlementListView(APIView):
    """
    지정 부금 관리: 특정 기간, 특정 영화에 대한 극장/요금/일자별 관객수 집계
    """

    def get(self, request):
        movie_id = request.query_params.get("movie_id")
        start_date = request.query_params.get("start_date")
        end_date = request.query_params.get("end_date")
        client_id = request.query_params.get("client_id")  # 선택 사항

        # 필수 파라미터 체크
        if not all([movie_id, start_date, end_date]):
            return Response(
                {"error": "영화 및 조회 기간(시작일, 종료일)은 필수입니다."},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            # 1. 대표영화 및 하위영화 ID 집계
            primary_movie = Movie.objects.get(id=movie_id)
            clean_parent_code = (primary_movie.movie_code or "").replace(" ", "")

            all_related_movie_ids = list(Movie.objects.annotate(
                clean_primary_code=Replace(F('primary_movie_code'), Value(' '), Value(''))
            ).filter(
                Q(id=movie_id) | Q(clean_primary_code=clean_parent_code)
            ).values_list("id", flat=True))

            # 2. 기본 필터링 (영화 목록 및 기간)
            filters = {
                "movie_id__in": all_related_movie_ids,
                "entry_date__range": [start_date, end_date]
            }

            # 3. 특정 극장 선택 시 필터 추가
            if client_id:
                filters["client_id"] = client_id

            # 3. 데이터 조회
            # values를 사용하여 필요한 필드만 추출하고, 성능을 위해 select_related와 유사하게 관련 필드 참조
            scores = Score.objects.filter(**filters).annotate(
                client_name=F('client__client_name')
            ).values(
                'client_name',
                'fare',
                'entry_date',
                'visitor'
            ).order_by('client_name', 'entry_date')

            # 프론트엔드 JS 로직에서 처리하기 편하도록
            # visitor(CharField)를 int로 변환하여 전달 (데이터가 비어있을 경우 0)
            result_data = []
            for s in scores:
                try:
                    visitor_count = int(
                        s['visitor']) if s['visitor'] and s['visitor'].isdigit() else 0
                except (ValueError, TypeError):
                    visitor_count = 0

                result_data.append({
                    "client_name": s['client_name'] or "미등록 극장",
                    "fare": s['fare'],
                    "entry_date": s['entry_date'].strftime("%Y-%m-%d") if s['entry_date'] else None,
                    "visitor": visitor_count
                })

            return Response(result_data, status=status.HTTP_200_OK)

        except Exception as e:
            return Response(
                {"error": f"데이터 조회 중 오류가 발생했습니다: {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


class SpecialSettlementExcelView(APIView):
    def get(self, request):
        # 1. 파라미터 추출
        start_date = request.query_params.get("start_date")
        end_date = request.query_params.get("end_date")
        movie_id = request.query_params.get("movie_id")
        client_id = request.query_params.get("client_id")

        # 2. 데이터 조회
        try:
            primary_movie = Movie.objects.get(id=movie_id)
            clean_parent_code = (primary_movie.movie_code or "").replace(" ", "")

            all_related_movie_ids = list(Movie.objects.annotate(
                clean_primary_code=Replace(F('primary_movie_code'), Value(' '), Value(''))
            ).filter(
                Q(id=movie_id) | Q(clean_primary_code=clean_parent_code)
            ).values_list("id", flat=True))
        except Movie.DoesNotExist:
            all_related_movie_ids = [movie_id]

        filters = {
            "movie_id__in": all_related_movie_ids,
            "entry_date__range": [start_date, end_date]
        }
        if client_id:
            filters["client_id"] = client_id

        # 성능을 위해 필요한 필드만 select
        scores = Score.objects.filter(**filters).annotate(
            c_name=F('client__client_name'),
            m_title=F('movie__title_ko')
        ).values('c_name', 'm_title', 'fare', 'entry_date', 'visitor').order_by('c_name', 'entry_date')

        if not scores:
            return HttpResponse("조회된 데이터가 없습니다.", status=404)

        movie_title = scores[0]['m_title'] or "영화"

        # 3. 데이터 가공 (날짜 피벗 로직)
        # 실제 실적이 있는 날짜만 추출 (중복 제거 및 정렬)
        active_dates = sorted(list(set(s['entry_date'].strftime(
            "%Y-%m-%d") for s in scores if s['entry_date'])))

        # (극장, 요금)을 키로 하여 데이터를 그룹화
        grouped = {}
        for s in scores:
            c_name = s['c_name'] or "미등록 극장"
            fare_val = int(
                s['fare']) if s['fare'] and s['fare'].isdigit() else 0
            key = (c_name, fare_val)

            if key not in grouped:
                # 해당 행의 모든 날짜별 관객수를 0으로 초기화
                grouped[key] = {d: 0 for d in active_dates}
                grouped[key]['total_visitor'] = 0

            date_str = s['entry_date'].strftime("%Y-%m-%d")
            visitor_cnt = int(
                s['visitor']) if s['visitor'] and s['visitor'].isdigit() else 0

            grouped[key][date_str] += visitor_cnt
            grouped[key]['total_visitor'] += visitor_cnt

        # 4. ExcelGenerator를 이용한 파일 생성
        excel = ExcelGenerator(sheet_name="지정부금집계")

        # 헤더 구성
        headers = ["극장명", "요금"] + active_dates + ["관객합계", "매출합계"]
        excel.add_header(headers)

        # 틀 고정 (C2: A,B열 및 1행 고정)
        excel.ws.freeze_panes = 'C2'

        # 데이터 행 구성
        data_rows = []
        for (c_name, fare), values in grouped.items():
            row = [c_name, fare]
            # 각 날짜별 관객수 배치
            for d in active_dates:
                row.append(values[d])

            # 합계 데이터 추가
            row.append(values['total_visitor'])  # 관객합계
            row.append(values['total_visitor'] * fare)  # 매출합계
            data_rows.append(row)

        excel.add_rows(data_rows)

        # 5. 응답 반환
        filename = f"SpecialSettlement_{movie_title}_{datetime.now().strftime('%Y%m%d')}"
        return excel.to_response(filename)


class SettlementEseroExportView(SettlementListView):
    def get(self, request):
        yyyy_mm = request.query_params.get("yyyyMm")
        movie_id = request.query_params.get("movie_id")
        target_filter = request.query_params.get("target", "전체극장")

        if not yyyy_mm or not movie_id:
            return HttpResponse("년월과 영화를 선택해주세요.", status=400)

        # 1. 영화 정보 및 배급사 정보 로드
        try:
            movie = Movie.objects.select_related(
                "distributor").get(id=movie_id)
            dist = movie.distributor
        except Movie.DoesNotExist:
            return HttpResponse("영화를 찾을 수 없습니다.", status=404)

        # [품목명 결정을 위한 대표 영화 및 제목 정제 로직]
        raw_title = movie.title_ko
        # 대표 영화가 아니고 대표 영화 코드가 존재하는 경우 대표 영화 제목 사용
        if not movie.is_primary_movie and movie.primary_movie_code:
            clean_parent_code = movie.primary_movie_code.replace(" ", "")
            primary_movie = Movie.objects.annotate(
                clean_code=Replace(F('movie_code'), Value(' '), Value(''))
            ).filter(clean_code=clean_parent_code, is_primary_movie=True).first()
            if primary_movie:
                raw_title = primary_movie.title_ko

        # 괄호 및 포맷 정보 제거: "(" 문자가 있으면 그 앞부분만 추출하여 공백 제거
        display_movie_title = raw_title.split(
            "(")[0].strip() if "(" in raw_title else raw_title

        # 2. 정산 데이터 조회 및 합산 (극장별 세금계산서 1장을 위한 집계)
        raw_items = self.get_processed_data(yyyy_mm, movie_id, target_filter)
        if not raw_items:
            return HttpResponse("조회된 데이터가 없습니다.", status=404)

        aggregated_for_esero = {}
        for item in raw_items:
            if item.get("is_subtotal"):
                continue

            biz_no = (item.get("사업자 등록번호") or "").replace("-", "")
            sub_biz_no = item.get("종사업장번호") or ""
            key = (biz_no, sub_biz_no)

            if key not in aggregated_for_esero:
                aggregated_for_esero[key] = item.copy()
            else:
                target = aggregated_for_esero[key]
                target["인원"] += item.get("인원", 0)
                target["공급가액"] += item.get("공급가액", 0)
                target["부가세"] += item.get("부가세", 0)
                target["영화사 지급금"] += item.get("영화사 지급금", 0)

                if item.get("날짜(To)", "") > target.get("날짜(To)", ""):
                    target["날짜(To)"] = item.get("날짜(To)", "")

        # 3. 엑셀 템플릿 로드 (건수에 따라 템플릿 선택)
        is_over_100 = len(aggregated_for_esero) > 100
        template_name = "esero_over100.xlsx" if is_over_100 else "esero_under100.xlsx"
        
        template_path = os.path.join(
            settings.BASE_DIR, "settlement", "excel_from", template_name)
        
        if not os.path.exists(template_path):
            return HttpResponse(f"템플릿 파일({template_name})을 찾을 수 없습니다.", status=404)

        wb = openpyxl.load_workbook(template_path)
        ws = wb.active

        # 말일자 계산
        yyyy, mm = map(int, yyyy_mm.split("-"))
        last_day_val = calendar.monthrange(yyyy, mm)[1]
        last_day_str = f"{yyyy}{mm:02d}{last_day_val:02d}"

        row_idx = 7
        for (biz_no, sub_biz), item in aggregated_for_esero.items():
            # A: 종류 (01: 일반) / B: 작성일자
            ws.cell(row=row_idx, column=1, value="01")

            multi_type = item.get("멀티구분") or ""
            clss = item.get("classification") or ""
            to_date_raw = item.get("날짜(To)", "")

            if ("롯데" in multi_type and (clss == "직영" or clss == "위탁")) or \
               ("메가" in multi_type and clss == "직영"):
                write_date = last_day_str
            else:
                write_date = to_date_raw.replace(
                    "-", "") if to_date_raw else last_day_str
            ws.cell(row=row_idx, column=2, value=write_date)

            supply_val = item.get("공급가액", 0)
            vat_val = item.get("부가세", 0)
            item_name = f"[{display_movie_title}]{mm}월 극장부금"

            if not is_over_100:
                # [CASE 1] 100건 이하 (공급자 정보 포함)
                if dist:
                    ws.cell(row=row_idx, column=3, value=(dist.business_registration_number or "").replace("-", ""))
                    ws.cell(row=row_idx, column=4, value=dist.business_operator or "")
                    ws.cell(row=row_idx, column=5, value=dist.business_name or "")
                    ws.cell(row=row_idx, column=6, value=dist.representative_name or "")
                    ws.cell(row=row_idx, column=7, value=dist.business_address or "")
                    ws.cell(row=row_idx, column=8, value=dist.business_category or "")
                    ws.cell(row=row_idx, column=9, value=dist.business_industry or "")
                    ws.cell(row=row_idx, column=10, value=dist.invoice_email_address or "")

                # K~S: 공급받는자(극장) 정보
                ws.cell(row=row_idx, column=11, value=biz_no)
                ws.cell(row=row_idx, column=12, value=sub_biz)
                ws.cell(row=row_idx, column=13, value=item.get("공급받는자 상호", ""))
                ws.cell(row=row_idx, column=14, value=item.get("공급받는자 성명", ""))
                ws.cell(row=row_idx, column=15, value=item.get("사업장 소재", ""))
                ws.cell(row=row_idx, column=16, value=item.get("업태", ""))
                ws.cell(row=row_idx, column=17, value=item.get("업종", ""))
                ws.cell(row=row_idx, column=18, value=item.get("수신자이메일", ""))
                ws.cell(row=row_idx, column=19, value=item.get("수신자이메일2", ""))

                # T~V: 합계 및 비고
                ws.cell(row=row_idx, column=20, value=supply_val)
                ws.cell(row=row_idx, column=21, value=vat_val)
                ws.cell(row=row_idx, column=22, value=item.get("극장명", ""))

                # W, X: 일자 및 품목명
                ws.cell(row=row_idx, column=23, value=write_date[-2:])
                ws.cell(row=row_idx, column=24, value=item_name)

                # AA~AC: 단가(합계), 공급가액, 세액
                ws.cell(row=row_idx, column=27, value=supply_val + vat_val)
                ws.cell(row=row_idx, column=28, value=supply_val)
                ws.cell(row=row_idx, column=29, value=vat_val)
            else:
                # [CASE 2] 101건 이상 (공급자 정보 미표기, 극장정보가 C열부터)
                ws.cell(row=row_idx, column=3, value=biz_no)
                ws.cell(row=row_idx, column=4, value=sub_biz)
                ws.cell(row=row_idx, column=5, value=item.get("공급받는자 상호", ""))
                ws.cell(row=row_idx, column=6, value=item.get("공급받는자 성명", ""))
                ws.cell(row=row_idx, column=7, value=item.get("사업장 소재", ""))
                ws.cell(row=row_idx, column=8, value=item.get("업태", ""))
                ws.cell(row=row_idx, column=9, value=item.get("업종", ""))
                ws.cell(row=row_idx, column=10, value=item.get("수신자이메일", ""))
                ws.cell(row=row_idx, column=11, value=item.get("수신자이메일2", ""))

                # L, M, N: 합계(공급가액, 세액) 및 비고
                ws.cell(row=row_idx, column=12, value=supply_val)
                ws.cell(row=row_idx, column=13, value=vat_val)
                ws.cell(row=row_idx, column=14, value=item.get("극장명", ""))

                # O, P: 일자 및 품목명
                ws.cell(row=row_idx, column=15, value=write_date[-2:])
                ws.cell(row=row_idx, column=16, value=item_name)

                # T, U: 품목1 공급가액, 세액
                ws.cell(row=row_idx, column=19, value=supply_val + vat_val)
                ws.cell(row=row_idx, column=20, value=supply_val)
                ws.cell(row=row_idx, column=21, value=vat_val)

            row_idx += 1

        # 4. 파일 응답 생성
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        filename_prefix = "Esero_Over100" if is_over_100 else "Esero_Under100"
        filename = f"{filename_prefix}_{yyyy_mm}_{datetime.now().strftime('%Y%m%d')}"
        response["Content-Disposition"] = f'attachment; filename="{filename}.xlsx"'
        wb.save(response)
        return response
